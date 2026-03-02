#!/usr/bin/env python3
"""
Atualiza o PDI da Ayla Lima com os dados do Matcher Comportamental (Sólides).
Adiciona seção no Resumo PDI e atualiza a aba Competências Referência.
Reconstrói o arquivo completo para garantir consistência.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Cores CIH ──
TEAL = '0097A7'
DARK_TEAL = '00796B'
WHITE = 'FFFFFF'
LIGHT_TEAL = 'E0F2F1'
LIGHT_GRAY = 'F5F5F5'
LIGHT_GREEN = 'E8F5E9'
LIGHT_YELLOW = 'FFFDE7'
LIGHT_RED = 'FFEBEE'
LIGHT_ORANGE = 'FFF3E0'
LIGHT_BLUE = 'E3F2FD'
PALE_YELLOW = 'FFF8E1'
LIGHT_PURPLE = 'F3E5F5'
GRAY_TEXT = '757575'
GREEN_BG = '43A047'
YELLOW_BG = 'FFC107'
RED_BG = 'E53935'
DARK_GRAY = '424242'
ORANGE_TEXT = 'F57C00'
RED_TEXT = 'E53935'

def hfill(c):
    return PatternFill(start_color=c, end_color=c, fill_type='solid')

header_fill = hfill(TEAL)
header_font = Font(bold=True, size=11, color=WHITE)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
title_font = Font(bold=True, size=14, color=TEAL)
title_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
subtitle_font = Font(bold=True, size=12, color=DARK_TEAL)
label_fill = hfill(LIGHT_TEAL)
label_font = Font(bold=True, size=10)
data_font = Font(size=10)
data_align_l = Alignment(horizontal='left', vertical='center', wrap_text=True)
data_align_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
ctx_fill = hfill(LIGHT_GRAY)
ctx_font = Font(bold=True, size=10)
ctx_data_align = Alignment(vertical='top', wrap_text=True)
src_fill = hfill(LIGHT_GRAY)
bold_font = Font(bold=True, size=10)
section_fill = hfill(LIGHT_TEAL)
section_font = Font(bold=True, size=11, color=DARK_TEAL)
tip_font = Font(size=10, italic=True, color=GRAY_TEXT)
example_font = Font(size=10, color='37474F')

thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

def set_header_row(ws, row, cols):
    for ci, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = thin_border

def set_data_cell(ws, row, col, val, font=None, fill=None, align=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font or data_font
    if fill:
        c.fill = fill
    c.alignment = align or data_align_l
    c.border = thin_border
    return c

def cell(ws, row, col, val, font=None, fill=None, align=None, merge_end=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font or data_font
    if fill:
        c.fill = fill
    c.alignment = align or data_align_l
    c.border = thin_border
    if merge_end:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end)
    return c

def level_fill(level):
    if level <= 1: return hfill(LIGHT_RED)
    if level == 2: return hfill(LIGHT_YELLOW)
    return hfill(LIGHT_GREEN)

def type_fill(tipo):
    if '70%' in tipo: return hfill(LIGHT_BLUE)
    if '20%' in tipo: return hfill(PALE_YELLOW)
    return hfill(LIGHT_PURPLE)

def section_title(ws, row, text, end_col=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = section_fill
    c.font = section_font
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = thin_border
    ws.row_dimensions[row].height = 28
    return row + 1

# ═══════════════════════════════════════════════
# DADOS
# ═══════════════════════════════════════════════
COMPETENCIAS = [
    (1, 'Organização e Gestão de Rotina', 1, 3, '25%'),
    (2, 'Atenção aos Detalhes e Saída do Automático', 1, 3, '25%'),
    (3, 'Comunicação Profissional e Formalidade', 2, 3, '20%'),
    (4, 'Pensamento Analítico e Gestão de Clientes', 1, 3, '20%'),
    (5, 'Autonomia e Senso de Dono', 2, 3, '10%'),
]

ACOES = [
    (1,'Organização e Gestão','70% Experiência','Implementar planner/agenda física',
     'Comprar e utilizar um planner ou caderno para organizar a semana. Dividir atividades por turno (manhã/tarde). Listar tarefas pendentes, prazos de aprovação e entregas do dia. Substituir o grupo de WhatsApp individual.',
     'Mar/2026','Jun/2026','Diário','Planner preenchido 5x/semana','Ayla','Devolutiva'),
    (2,'Organização e Gestão','70% Experiência','Checklist diário por turno',
     'Criar checklist estruturado com tarefas obrigatórias de cada turno: verificar aprovações pendentes, enviar conteúdos programados, conferir Trello, enviar pendências ao final do dia.',
     'Mar/2026','Jun/2026','Diário','100% checklists preenchidos','Ayla','Briefing liderança'),
    (3,'Organização e Gestão','70% Experiência','Verificação do Trello 2x por turno',
     'Abrir o Trello obrigatoriamente no início e fim de cada turno. Verificar: cards vencidos, exclusões recentes, alterações de status, conteúdos aguardando aprovação. Anotar no planner.',
     'Mar/2026','Jun/2026','Diário','4 verificações/dia registradas','Ayla','Devolutiva + Briefing'),
    (4,'Organização e Gestão','70% Experiência','Rotina de encerramento do dia',
     'Antes de finalizar o expediente: enviar lista de pendências para o grupo, verificar aprovações atrasadas, conferir se todos os posts foram enviados. Criar hábito de "fechar o dia".',
     'Mar/2026','Jun/2026','Diário','Pendências enviadas 5x/semana','Ayla','Devolutiva'),
    (5,'Organização e Gestão','20% Social','Acompanhamento quinzenal com Larissa',
     '15 min quinzenais para revisar juntas: uso do planner, checklist, organização do Trello. Larissa dá feedback sobre evolução da organização e pontualidade nas entregas.',
     'Mar/2026','Jun/2026','Quinzenal','2 conversas/mês registradas','Ayla + Larissa','Briefing'),
    (6,'Organização e Gestão','10% Formal','Estudo sobre gestão de tempo',
     'Assistir conteúdo sobre técnicas de organização pessoal (Pomodoro, GTD ou similar). Escolher 1 técnica e implementar no dia a dia. Discutir com Eleine na sessão.',
     'Mar/2026','Abr/2026','Única','1 técnica escolhida e implementada','Ayla','Devolutiva'),
    (7,'Atenção aos Detalhes','70% Experiência','Protocolo "PARE, CONFIRA, ENVIE"',
     'Antes de qualquer envio (post, aprovação, mensagem para cliente): PARAR, CONFERIR (grupo correto? conteúdo correto? aprovação dada?), e só então ENVIAR. Fixar regra no monitor como lembrete visual.',
     'Mar/2026','Jun/2026','Contínuo','Zero erros de envio por semana','Ayla','Devolutiva'),
    (8,'Atenção aos Detalhes','70% Experiência','Diário de erros e aprendizados',
     'Registrar cada erro cometido: o que aconteceu, por que aconteceu, como evitar. Revisar semanalmente para identificar padrões. Foco: sair do automático.',
     'Mar/2026','Mai/2026','Contínuo','Redução progressiva de erros','Ayla','DISC + Devolutiva'),
    (9,'Atenção aos Detalhes','70% Experiência','Revisão dupla de postagens',
     'Antes de publicar qualquer conteúdo, verificar: 1) Está no card correto do Trello? 2) Foi aprovado? 3) É o grupo/rede certa? 4) Texto e imagem estão corretos? Criar checklist mental de 4 pontos.',
     'Mar/2026','Jun/2026','Contínuo','0 postagens erradas/semana','Ayla','Briefing'),
    (10,'Atenção aos Detalhes','20% Social','Feedback semanal com Larissa sobre erros',
     'Larissa reporta erros detectados na semana. Ayla analisa se eram evitáveis com o protocolo. Ajustar processo conforme necessário.',
     'Mar/2026','Jun/2026','Semanal','Redução semanal de erros','Ayla + Larissa','Briefing'),
    (11,'Comunicação Profissional','70% Experiência','Criar banco de mensagens padrão',
     'Criar templates para: cobranças de aprovação, comunicados de feriado, onboarding de novos clientes, solicitação de materiais, follow-up. Usar linguagem formal e detalhista.',
     'Mar/2026','Abr/2026','Única','5+ templates criados e aprovados','Ayla','Devolutiva'),
    (12,'Comunicação Profissional','70% Experiência','Adaptar linguagem por perfil de cliente',
     'Identificar quais clientes preferem comunicação direta (executor) vs detalhada (analista). Ajustar tom, nível de detalhe e formalidade conforme cada cliente.',
     'Mar/2026','Jun/2026','Contínuo','0 reclamações de tom/linguagem','Ayla','DISC'),
    (13,'Comunicação Profissional','20% Social','Revisão de comunicações com Larissa',
     'Submeter comunicações importantes para revisão antes do envio. Larissa indica ajustes de tom, formalidade e detalhamento. Reduzir gradualmente a necessidade de revisão.',
     'Mar/2026','Mai/2026','Semanal','Aprovação em 1ª tentativa ≥80%','Ayla + Larissa','Devolutiva'),
    (14,'Comunicação Profissional','10% Formal','Estudo do resultado DISC completo',
     'Ler o material DISC completo (38 páginas). Entender perfis Comunicador, Executor, Analista e Planejador. Aplicar conceitos de comunicação adaptada no dia a dia.',
     'Mar/2026','Mar/2026','Única','Leitura concluída + 3 insights','Ayla','DISC'),
    (15,'Pensamento Analítico','70% Experiência','Análise semanal de clientes no Trello',
     'Dedicar 30 min por semana para analisar cada cliente: cards atrasados, aprovações pendentes, gargalos. Mapear situação e antecipar problemas.',
     'Mar/2026','Jun/2026','Semanal','Análise de 100% dos clientes/semana','Ayla','Briefing'),
    (16,'Pensamento Analítico','70% Experiência','Relatório "Top 3 Clientes" quinzenal',
     'Trazer para Larissa os 3 clientes mais críticos: qual o problema, possível causa, sugestão de ação. Desenvolver pensamento analítico aplicado.',
     'Mar/2026','Jun/2026','Quinzenal','2 relatórios/mês com análise','Ayla','Briefing'),
    (17,'Pensamento Analítico','70% Experiência','Registro de soluções no playbook pessoal',
     'Ao resolver um problema, registrar: situação, solução aplicada, resultado. Usar como base para situações futuras similares. Construir playbook pessoal.',
     'Mar/2026','Jun/2026','Contínuo','3+ soluções registradas/mês','Ayla','Briefing + Devolutiva'),
    (18,'Pensamento Analítico','20% Social','Discussão com Larissa sobre leitura de cenários',
     'Apresentar análise dos clientes para Larissa e discutir: o que Ayla viu vs o que não percebeu. Desenvolver visão holística e antecipação de problemas.',
     'Mar/2026','Jun/2026','Quinzenal','2 discussões/mês com insights','Ayla + Larissa','Devolutiva'),
    (19,'Autonomia e Senso de Dono','70% Experiência','Cobrar proativamente a equipe',
     'Não esperar que design atrase para cobrar. Criar rotina de follow-up: verificar status das demandas, cobrar com assertividade e profissionalismo. Escalar para Larissa se necessário.',
     'Mar/2026','Jun/2026','Contínuo','0 atrasos por falta de cobrança','Ayla','Devolutiva'),
    (20,'Autonomia e Senso de Dono','70% Experiência','Pedir ajuda cedo em vez de travar',
     'Quando travar em uma tarefa por mais de 15 min, pedir ajuda. Registrar: o que travou, quanto tempo ficou travada. Meta: reduzir tempo travada de 30min+ para <15min.',
     'Mar/2026','Jun/2026','Contínuo','Tempo médio até pedir ajuda <15min','Ayla','DISC + Devolutiva'),
]

DISC_INDICADORES = [
    ('Perfil Analista (A)', '19,48%', 'Baixo', 'Gap principal — cargo exige detalhismo', 'Foco do PDI em organização e atenção'),
    ('Autoestima (IAE)', 'Baixa', 'Atenção', 'Autocrítica elevada, risco de desânimo', 'Feedback positivo e celebração de conquistas'),
    ('Aproveitamento (IA)', 'Normal Baixo', 'Atenção', 'Sente habilidades subutilizadas', 'Dar mais responsabilidade e desafios'),
    ('Autoconfiança (IAC)', 'Normal Baixa', 'Atenção', 'Sente necessidade de mudar muito', 'Evolução gradual com metas curtas'),
    ('Energia (EN)', 'Alta', 'Positivo', 'Disposição e motivação presentes', 'Canalizar em organização e análise'),
    ('Área de Talento', 'Diplomata', 'Extremamente Alto', 'Excelente em relações interpessoais', 'Usar como alavanca na comunicação'),
]

# ═══════════════════════════════════════════════
# DADOS DO MATCHER (NOVOS)
# ═══════════════════════════════════════════════
MATCHER_GAPS = [
    ('Executor', '↑ 7,07%', 'Ayla tem MAIS do que o cargo exige', hfill(LIGHT_GREEN), 'Excesso pode gerar impulsividade — canalizar em proatividade'),
    ('Comunicador', '↑ 6,25%', 'Ayla tem MAIS do que o cargo exige', hfill(LIGHT_GREEN), 'Excesso gera informalidade — treinar formalidade sem perder a naturalidade'),
    ('Planejador', '↓ 4,22%', 'Ayla tem MENOS do que o cargo exige', hfill(LIGHT_YELLOW), 'Desenvolver via planner, checklist e rotina de encerramento do dia'),
    ('Analista', '↓ 9,09%', 'Ayla tem MENOS do que o cargo exige', hfill(LIGHT_RED), 'MAIOR GAP — foco em atenção, detalhismo, organização e análise'),
]

MATCHER_COMPETENCIAS_GAP = [
    ('Planejamento', 'Normal Alto', 'Alto', 'Cargo exige mais planejamento do que Ayla demonstra naturalmente', 'Competências 1 e 4'),
    ('Empatia', 'Normal Alto', 'Alto', 'Cargo exige mais leitura emocional dos clientes', 'Competência 3'),
    ('Capacidade de ouvir', 'Normal Alto', 'Alto', 'Cargo exige escuta ativa mais apurada para entender demandas', 'Competência 3'),
    ('Concentração', 'Normal Alto', 'Alto', 'Cargo exige foco sustentado — modo automático prejudica', 'Competência 2'),
    ('Condescendência', 'Normal Alto', 'Alto', 'Cargo exige mais paciência e flexibilidade com os clientes', 'Competência 3'),
    ('Perfil Técnico', 'Normal Alto', 'Alto', 'Cargo exige domínio técnico mais aprofundado dos processos', 'Competência 4'),
    ('Organização', 'Normal Alto', 'Alto', 'Cargo exige organização formal superior à natural de Ayla', 'Competência 1'),
    ('Detalhismo', 'Normal Alto', 'Alto', 'Cargo exige mais atenção a detalhes — confirma gap do Analista', 'Competência 2'),
]

# ═══════════════════════════════════════════════
wb = openpyxl.Workbook()
output = '/home/headless/workspace/cih/operacional/execução/valeassessoria/aylalima/PDI_Ayla_Lima_ValeAssessoria.xlsx'

# ═══════════════════════════════════════════════
# ABA 1 — RESUMO PDI (COM MATCHER)
# ═══════════════════════════════════════════════
ws = wb.active
ws.title = 'Resumo PDI'
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 45
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15

ws.merge_cells('A1:H1')
c = ws.cell(row=1, column=1, value='PDI — AYLA JESUS LIMA | VALE ASSESSORIA')
c.font = title_font; c.alignment = title_align

ws.merge_cells('A2:H2')
c = ws.cell(row=2, column=1, value='Plano de Desenvolvimento Individual — Ciclo Março a Junho/2026 (v2 — atualizado com Matcher Comportamental)')
c.font = Font(bold=True, size=12, color=DARK_TEAL); c.alignment = title_align

ficha = [
    ('Colaboradora', 'Ayla Jesus Lima'),
    ('Cargo', 'Estagiária → Analista de Operações e Atendimento de Contas (em transição)'),
    ('Empresa', 'Vale Assessoria'),
    ('Tempo de Casa', '~1 ano'),
    ('Perfil DISC', 'Comunicador Executor (CE) — Diplomata'),
    ('Match Comportamental', '86% de compatibilidade com o cargo (Matcher Sólides — 01/03/2026)'),
    ('Perfil Exigido pelo Cargo', 'Analista Planejador (AP) — cargo exige mais análise e planejamento'),
    ('Liderança Direta', 'Larissa Carinhanha (gestora — responsável pelo acompanhamento do PDI)'),
    ('Consultora CIH', 'Eleine Passos'),
    ('E-mail', 'aylalima93@gmail.com'),
    ('Ciclo PDI', 'Março/2026 a Junho/2026 (4 meses)'),
    ('Objetivo', 'Reduzir riscos na promoção e acelerar maturidade comportamental e profissional'),
]
for i, (lab, val) in enumerate(ficha, 4):
    cl = ws.cell(row=i, column=1, value=lab)
    cl.fill = label_fill; cl.font = label_font; cl.border = thin_border
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    cv = ws.cell(row=i, column=2, value=val)
    cv.font = data_font; cv.border = thin_border
    # Destaque visual para Match e Perfil Exigido
    if 'Match' in lab:
        cv.font = Font(bold=True, size=10, color=TEAL)
    elif 'Perfil Exigido' in lab:
        cv.font = Font(bold=True, size=10, color=RED_TEXT)

r = 17
ws.merge_cells(f'A{r}:H{r}')
ws.cell(row=r, column=1, value='CONTEXTO DA PROMOÇÃO (BRIEFING COM LIDERANÇA)').font = subtitle_font

contexto = [
    ('Papel definido', 'Analista de Operações e Atendimento de Contas (PJ). Responsável por gestão de materiais, comunicação com clientes, organização interna, onboarding e apoio a projetos.'),
    ('Expectativa da liderança', 'Sair do modo automático, reduzir erros de envio/postagem, ser mais organizada, formal na comunicação com clientes e analítica. Cobrar equipe com assertividade.'),
    ('Gap identificado', 'Perfil Analista mais baixo (19,48%) vs alta exigência do cargo (perfil AP). Matcher confirma: Analista ↓9,09% é o maior gap. Modo automático gera erros frequentes.'),
    ('Ferramentas em uso', 'Trello (gestão de tarefas), WhatsApp (grupos de clientes), Canva (criação emergencial), Google Drive. Sugestão: planner físico + checklist diário.'),
    ('Risco na promoção', 'Sem desenvolvimento, erros operacionais podem aumentar com nova responsabilidade. Falta de organização e detalhismo compromete a entrega ao cliente.'),
    ('Compromisso 50/50', 'Desenvolvimento é 50% da colaboradora e 50% da empresa. Liderança fornecerá feedback constante e tempo para implementação das mudanças.'),
]
for j, (lab, val) in enumerate(contexto, r+1):
    cl = ws.cell(row=j, column=1, value=lab)
    cl.fill = ctx_fill; cl.font = ctx_font; cl.border = thin_border
    ws.merge_cells(start_row=j, start_column=2, end_row=j, end_column=8)
    cv = ws.cell(row=j, column=2, value=val)
    cv.font = data_font; cv.alignment = ctx_data_align; cv.border = thin_border
    ws.row_dimensions[j].height = 35

# Competências priorizadas
r = 25
ws.merge_cells(f'A{r}:H{r}')
ws.cell(row=r, column=1, value='COMPETÊNCIAS PRIORIZADAS (5 COMPETÊNCIAS)').font = subtitle_font

r = 26
set_header_row(ws, r, ['#','Competência','Nível Atual','Meta','Peso','Status Mês 1','Status Mês 2','Status Mês 3'])

for ci, (num, nome, atual, meta, peso) in enumerate(COMPETENCIAS, 0):
    row = 27 + ci
    set_data_cell(ws, row, 1, num, align=data_align_c)
    set_data_cell(ws, row, 2, nome, align=data_align_l)
    set_data_cell(ws, row, 3, atual, fill=level_fill(atual), align=data_align_c)
    set_data_cell(ws, row, 4, meta, fill=hfill(LIGHT_GREEN), align=data_align_c)
    set_data_cell(ws, row, 5, peso, align=data_align_c)

# Indicadores DISC
r = 33
ws.merge_cells(f'A{r}:H{r}')
ws.cell(row=r, column=1, value='INDICADORES DISC — PONTOS DE ATENÇÃO').font = subtitle_font

r = 34
set_header_row(ws, r, ['Indicador','Valor','Classificação','Impacto','Ação Recomendada','','',''])

for j, (ind, val, classif, impacto, acao) in enumerate(DISC_INDICADORES, r+1):
    set_data_cell(ws, j, 1, ind, align=data_align_l)
    set_data_cell(ws, j, 2, val, align=data_align_c)
    classif_fill = hfill(LIGHT_ORANGE) if classif in ('Baixo','Atenção') else hfill(LIGHT_GREEN)
    set_data_cell(ws, j, 3, classif, fill=classif_fill, align=data_align_c)
    set_data_cell(ws, j, 4, impacto, align=data_align_l)
    set_data_cell(ws, j, 5, acao, align=data_align_l)

# ═══════════════════════════════════════════════
# NOVA SEÇÃO: MATCH COMPORTAMENTAL
# ═══════════════════════════════════════════════
r = 42
ws.merge_cells(f'A{r}:H{r}')
ws.cell(row=r, column=1, value='MATCH COMPORTAMENTAL — PERFIL AYLA (CE) vs. CARGO (AP)').font = subtitle_font

r += 1
ws.merge_cells(f'A{r}:H{r}')
c = ws.cell(row=r, column=1, value='Fonte: Matcher Sólides — 01/03/2026 | Match geral: 86% | Perfil Ayla: CE (Comunicador Executor) | Perfil do Cargo: AP (Analista Planejador)')
c.font = Font(size=10, color=GRAY_TEXT); c.alignment = data_align_c

r += 1
# Destaque do score
ws.merge_cells(f'A{r}:B{r}')
c = ws.cell(row=r, column=1, value='MATCH: 86%')
c.font = Font(bold=True, size=16, color=TEAL); c.alignment = data_align_c
ws.merge_cells(f'C{r}:H{r}')
c = ws.cell(row=r, column=3, value='Boa compatibilidade geral, mas com gaps específicos nas dimensões Analista e Planejador que justificam as 5 competências do PDI.')
c.font = Font(size=10); c.alignment = data_align_l
ws.row_dimensions[r].height = 35

# Gaps DISC
r += 2
set_header_row(ws, r, ['Dimensão DISC','Gap','Direção','','Impacto no PDI','','',''])
r += 1

for dim, gap, direcao, gap_fill, impacto in MATCHER_GAPS:
    set_data_cell(ws, r, 1, dim, font=Font(bold=True, size=10), align=data_align_c)
    set_data_cell(ws, r, 2, gap, fill=gap_fill, font=Font(bold=True, size=10), align=data_align_c)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    set_data_cell(ws, r, 3, direcao, align=data_align_l)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
    set_data_cell(ws, r, 5, impacto, align=data_align_l)
    ws.row_dimensions[r].height = 28
    r += 1

# Competências com gap
r += 1
ws.merge_cells(f'A{r}:H{r}')
ws.cell(row=r, column=1, value='COMPETÊNCIAS ONDE O CARGO EXIGE MAIS DO QUE AYLA DEMONSTRA (MATCHER)').font = subtitle_font
r += 1

set_header_row(ws, r, ['Competência','Nível Ayla','Nível Exigido','Gap / Impacto','Conecta com PDI','','',''])
r += 1

for comp, nivel_ayla, nivel_cargo, impacto, pdi_link in MATCHER_COMPETENCIAS_GAP:
    set_data_cell(ws, r, 1, comp, font=Font(bold=True, size=10), align=data_align_l)
    set_data_cell(ws, r, 2, nivel_ayla, fill=hfill(LIGHT_YELLOW), align=data_align_c)
    set_data_cell(ws, r, 3, nivel_cargo, fill=hfill(LIGHT_ORANGE), align=data_align_c)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    set_data_cell(ws, r, 4, impacto, align=data_align_l)
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
    set_data_cell(ws, r, 7, pdi_link, font=Font(bold=True, size=10, color=TEAL), align=data_align_c)
    ws.row_dimensions[r].height = 28
    r += 1

# Nota de interpretação
r += 1
ws.merge_cells(f'A{r}:H{r}')
c = ws.cell(row=r, column=1, value='INTERPRETAÇÃO: O Matcher confirma que o PDI está corretamente direcionado. Os 3 maiores gaps (Analista ↓9,09%, Organização Normal Alto→Alto, Detalhismo Normal Alto→Alto) estão cobertos pelas Competências 1, 2 e 4. As competências com excesso (Comunicador ↑6,25%, Executor ↑7,07%) são ativos — devem ser canalizados, não reprimidos.')
c.font = Font(size=10, italic=True, color=DARK_TEAL); c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.row_dimensions[r].height = 45

# ═══════════════════════════════════════════════
# ABA 2 — PLANO DE AÇÕES (mantido)
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet('Plano de Ações')
ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 35
ws2.column_dimensions['E'].width = 60
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 12
ws2.column_dimensions['H'].width = 12
ws2.column_dimensions['I'].width = 38
ws2.column_dimensions['J'].width = 15
ws2.column_dimensions['K'].width = 18
ws2.column_dimensions['L'].width = 12

ws2.merge_cells('A1:L1')
ws2.cell(row=1, column=1, value='PLANO DE AÇÕES — PDI AYLA LIMA (20 ações, 5 competências)').font = title_font
ws2['A1'].alignment = title_align

set_header_row(ws2, 3, ['#','Competência','Tipo 70-20-10','Ação','Descrição Detalhada',
                         'Prazo Início','Prazo Fim','Frequência','Indicador de Sucesso',
                         'Responsável','Fonte','Status'])

for i, a in enumerate(ACOES):
    row = 4 + i
    num, comp, tipo, acao, desc, ini, fim, freq, indic, resp, fonte = a
    set_data_cell(ws2, row, 1, num, align=data_align_c)
    set_data_cell(ws2, row, 2, comp, align=data_align_l)
    set_data_cell(ws2, row, 3, tipo, fill=type_fill(tipo), align=data_align_l)
    set_data_cell(ws2, row, 4, acao, align=data_align_l)
    set_data_cell(ws2, row, 5, desc, align=data_align_l)
    set_data_cell(ws2, row, 6, ini, align=data_align_c)
    set_data_cell(ws2, row, 7, fim, align=data_align_c)
    set_data_cell(ws2, row, 8, freq, align=data_align_c)
    set_data_cell(ws2, row, 9, indic, align=data_align_l)
    set_data_cell(ws2, row, 10, resp, align=data_align_c)
    set_data_cell(ws2, row, 11, fonte, fill=src_fill, align=data_align_l)
    set_data_cell(ws2, row, 12, 'Pendente', align=data_align_c)

# ═══════════════════════════════════════════════
# ABA 3 — ACOMPANHAMENTO SEMANAL (mantido)
# ═══════════════════════════════════════════════
ws3 = wb.create_sheet('Acompanhamento Semanal')
ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 55
ws3.column_dimensions['D'].width = 14
ws3.column_dimensions['E'].width = 14
ws3.column_dimensions['F'].width = 14
ws3.column_dimensions['G'].width = 14
ws3.column_dimensions['H'].width = 14
ws3.column_dimensions['I'].width = 35
ws3.column_dimensions['J'].width = 35
ws3.column_dimensions['K'].width = 30
ws3.column_dimensions['L'].width = 30

ws3.merge_cells('A1:L1')
ws3.cell(row=1, column=1, value='ACOMPANHAMENTO SEMANAL — MARÇO A JUNHO/2026').font = title_font
ws3['A1'].alignment = title_align

set_header_row(ws3, 3, ['Semana','Período','Foco Principal',
    'Organização\n(Nota 1-5)','Atenção\n(Nota 1-5)','Comunicação\n(Nota 1-5)',
    'Pens. Analítico\n(Nota 1-5)','Autonomia\n(Nota 1-5)',
    'Ações Realizadas','Dificuldades / Bloqueios','Próximos Passos','Observações do Consultor'])

semanas = [
    ('Semana 1','03/03 a 07/03','Implementar planner + 1° checklist + Protocolo PARE-CONFIRA-ENVIE + Verificação Trello',False),
    ('Semana 2','10/03 a 14/03','Início banco de mensagens + Análise clientes Trello + Diário de erros + Leitura DISC',False),
    ('Semana 3','17/03 a 21/03','1° Top 3 Clientes + Feedback Larissa + Adaptar linguagem + Cobranças assertivas',False),
    ('Semana 4','24/03 a 28/03','CHECKPOINT MÊS 1 — Avaliar primeiras semanas + Estudo gestão de tempo',True),
    ('Semana 5','31/03 a 04/04','Consolidar organização + Aprofundar análise de clientes + Templates de mensagens',False),
    ('Semana 6','07/04 a 11/04','2° Top 3 Clientes + Revisão comunicações + Playbook pessoal',False),
    ('Semana 7','14/04 a 18/04','Ajustar protocolo de revisão + Feedback Larissa + Verificar erros recorrentes',False),
    ('Semana 8','21/04 a 25/04','CHECKPOINT MÊS 2 — Avaliar evolução + Larissa avalia organização + Revisão PDI',True),
    ('Semana 9','28/04 a 02/05','Consolidar comunicação formal + Aprofundar análise de cenários + Cobrança equipe',False),
    ('Semana 10','05/05 a 09/05','3° Top 3 Clientes + Avaliar playbook + Autonomia em decisões',False),
    ('Semana 11','12/05 a 16/05','Verificar hábitos consolidados + Planner e checklist automáticos?',False),
    ('Semana 12','19/05 a 23/05','Preparar evidências de evolução + Consolidar competências',False),
    ('Semana 13','26/05 a 30/05','CHECKPOINT MÊS 3 — Revisão geral + Avaliar prontidão para promoção',True),
    ('Semana 14','02/06 a 06/06','Consolidação final + Testar autonomia total',False),
    ('Semana 15','09/06 a 13/06','Ajustes finais + Preparar parecer de evolução',False),
    ('Semana 16','16/06 a 20/06','CHECKPOINT FINAL — Avaliação final + Parecer de prontidão para promoção',True),
]

for i, (sem, per, foco, is_cp) in enumerate(semanas, 4):
    fill_row = hfill(LIGHT_ORANGE) if is_cp else None
    set_data_cell(ws3, i, 1, sem, font=Font(bold=True, size=10), fill=fill_row, align=data_align_c)
    set_data_cell(ws3, i, 2, per, fill=fill_row, align=data_align_c)
    set_data_cell(ws3, i, 3, foco, fill=fill_row, align=Alignment(vertical='top', wrap_text=True))
    ws3.row_dimensions[i].height = 45

# ═══════════════════════════════════════════════
# ABA 4 — RESULTADO E EVOLUÇÃO (mantido)
# ═══════════════════════════════════════════════
ws4 = wb.create_sheet('Resultado e Evolução')
ws4.column_dimensions['A'].width = 48
for col in 'BCDEFGH':
    ws4.column_dimensions[col].width = 15

ws4.merge_cells('A1:H1')
ws4.cell(row=1, column=1, value='RESULTADO DO PDI — EVOLUÇÃO DAS COMPETÊNCIAS').font = title_font
ws4['A1'].alignment = title_align

set_header_row(ws4, 3, ['Competência','Nível Inicial','Meta','Mês 1','Mês 2','Mês 3','Nível Final','Atingiu?'])

comp_names_full = [
    'Organização e Gestão de Rotina',
    'Atenção aos Detalhes e Saída do Automático',
    'Comunicação Profissional e Formalidade',
    'Pensamento Analítico e Gestão de Clientes',
    'Autonomia e Senso de Dono',
]
comp_levels = [(1,3),(1,3),(2,3),(1,3),(2,3)]

for i, (nome, (ini, meta)) in enumerate(zip(comp_names_full, comp_levels), 4):
    set_data_cell(ws4, i, 1, nome, align=data_align_l)
    set_data_cell(ws4, i, 2, ini, fill=level_fill(ini), align=data_align_c)
    set_data_cell(ws4, i, 3, meta, fill=hfill(LIGHT_GREEN), align=data_align_c)

r = 10
ws4.merge_cells(f'A{r}:H{r}')
ws4.cell(row=r, column=1, value='INDICADORES DE ACOMPANHAMENTO').font = subtitle_font

r = 11
set_header_row(ws4, r, ['Indicador','Baseline','Meta Mês 1','Real Mês 1','Meta Mês 2','Real Mês 2','Meta Mês 3','Real Mês 3'])

indicadores = [
    ('Checklist diário preenchido (%)', 0, '80%', '', '90%', '', '100%', ''),
    ('Verificações Trello por dia', 0, 4, '', 4, '', 4, ''),
    ('Erros de envio por semana (↓)', '—', 'Registrar', '', 'Reduzir 30%', '', 'Reduzir 60%', ''),
    ('Templates de mensagens criados', 0, 5, '', 'Usar', '', 'Ajustar', ''),
    ('Análises semanais de clientes', 0, 4, '', 4, '', 4, ''),
    ('Relatórios Top 3 Clientes', 0, 2, '', 2, '', 2, ''),
    ('Cobranças assertivas sem atraso', '—', 'Registrar', '', '80%+', '', '90%+', ''),
    ('Protocolo PARE-CONFIRA-ENVIE aplicado', '—', 'Sempre', '', 'Automático', '', 'Consolidado', ''),
    ('Reuniões feedback com Larissa', 0, 2, '', 2, '', 2, ''),
    ('Soluções registradas no playbook', 0, 3, '', 3, '', 3, ''),
    ('Planner utilizado diariamente', 'Não', '80%', '', '90%', '', '100%', ''),
    ('Ações do PDI concluídas (acumulado)', '0/20', '7/20', '', '14/20', '', '18/20', ''),
]

for i, ind in enumerate(indicadores, r+1):
    for j, val in enumerate(ind):
        set_data_cell(ws4, i, j+1, val, align=data_align_c if j > 0 else data_align_l)

r2 = r + 1 + len(indicadores) + 1
ws4.merge_cells(f'A{r2}:H{r2}')
ws4.cell(row=r2, column=1, value='FEEDBACK DA LIDERANÇA (COLETAS COM LARISSA)').font = subtitle_font

r2 += 1
set_header_row(ws4, r2, ['Pergunta','1ª Coleta (Mês 1)\nNota 1-5','Comentários',
                          '2ª Coleta (Mês 2)\nNota 1-5','Comentários','Evolução','',''])

perguntas_fb = [
    'Como você avalia a organização da Ayla nas entregas?',
    'Ayla tem reduzido erros de envio e postagem?',
    'A comunicação com clientes está mais profissional e formal?',
    'Ayla está usando o Trello de forma consistente?',
    'Ayla pede ajuda quando precisa ou trava sozinha?',
    'As cobranças à equipe estão mais assertivas e profissionais?',
    'Ayla demonstra pensamento analítico sobre os clientes?',
    'Ayla demonstra mais autonomia e senso de responsabilidade?',
    'Nota geral para a evolução da Ayla (1 a 5)',
]

for i, p in enumerate(perguntas_fb, r2+1):
    set_data_cell(ws4, i, 1, p, align=Alignment(vertical='top', wrap_text=True))

r3 = r2 + 1 + len(perguntas_fb) + 1
ws4.merge_cells(f'A{r3}:H{r3}')
ws4.cell(row=r3, column=1, value='RESUMO EXECUTIVO — RESULTADO DO CICLO').font = subtitle_font

resumo_items = [
    'Principais conquistas:',
    'Competências com maior evolução:',
    'Competências que precisam de mais tempo:',
    'Impacto percebido nos clientes:',
    'Impacto na operação (organização, comunicação, entregas):',
    'Percepção de Larissa sobre maturidade de Ayla:',
    'Evolução na redução de erros:',
    'Evolução dos gaps do Matcher (Analista ↓9,09%, Planejador ↓4,22%):',
    'Recomendações para o próximo ciclo:',
    'Nota geral do PDI (consultor):',
]

for i, item in enumerate(resumo_items, r3+1):
    ws4.cell(row=i, column=1, value=item).font = Font(bold=True, size=10)
    ws4.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
    ws4.row_dimensions[i].height = 30

# ═══════════════════════════════════════════════
# ABA 5 — CHECKPOINTS (mantido)
# ═══════════════════════════════════════════════
ws5 = wb.create_sheet('Checkpoints')
ws5.column_dimensions['A'].width = 8
ws5.column_dimensions['B'].width = 55
ws5.column_dimensions['C'].width = 25
for col in 'DEFG':
    ws5.column_dimensions[col].width = 15

ws5.merge_cells('A1:G1')
ws5.cell(row=1, column=1, value='CHECKPOINTS MENSAIS — REGISTRO DE SESSÕES').font = title_font
ws5['A1'].alignment = title_align

def checkpoint_block(ws, r, titulo, perguntas):
    ws.merge_cells(f'A{r}:G{r}')
    c = ws.cell(row=r, column=1, value=titulo)
    c.fill = hfill(LIGHT_TEAL); c.font = Font(bold=True, size=12, color=DARK_TEAL)
    ws.cell(row=r+1, column=1, value='Participantes:').font = Font(bold=True, size=10)
    ws.cell(row=r+1, column=2, value='Eleine + Ayla + Larissa').font = data_font
    ws.cell(row=r+2, column=1, value='Data realizada:').font = Font(bold=True, size=10)
    ws.cell(row=r+2, column=2, value='___/___/______').font = data_font
    r += 3
    set_header_row(ws, r, ['#','Pergunta / Item','Resposta / Evidência','','','',''])
    for i, p in enumerate(perguntas):
        row = r + 1 + i
        set_data_cell(ws, row, 1, i+1, align=data_align_c)
        set_data_cell(ws, row, 2, p, align=Alignment(vertical='top', wrap_text=True))
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        ws.row_dimensions[row].height = 30
    return r + 1 + len(perguntas) + 1

r = 3
r = checkpoint_block(ws5, r, 'CHECKPOINT MÊS 1 — Fim de Março/2026', [
    'Ayla está usando o planner diariamente?',
    'Checklist por turno: está preenchendo e seguindo?',
    'Verificação do Trello: está fazendo 2x por turno?',
    'Protocolo PARE-CONFIRA-ENVIE: está aplicando antes de cada envio?',
    'Diário de erros: quantos erros registrou? Quais padrões?',
    'Banco de mensagens padrão: quantos templates criou?',
    'Comunicação com clientes: está mais formal e detalhista?',
    'Análise de clientes: fez análise semanal no Trello?',
    'Top 3 Clientes: entregou o 1° relatório?',
    'Cobranças à equipe: está cobrando proativamente?',
    'Pedir ajuda: tempo médio até pedir ajuda diminuiu?',
    'Leitura DISC: concluiu? Quais insights?',
    'Nível de motivação de Ayla (autoavaliação 1-5):',
    'Percepção de Larissa: Ayla está evoluindo? O que falta?',
    'Ajustes necessários no PDI:',
])

r = checkpoint_block(ws5, r, 'CHECKPOINT MÊS 2 — Fim de Abril/2026', [
    'Organização: planner e checklist viraram hábito?',
    'Erros de envio: houve redução significativa?',
    'Protocolo de revisão: está automático ou ainda precisa lembrar?',
    'Templates de mensagem: estão sendo usados? São eficazes?',
    'Comunicação: Larissa percebe mais formalidade e profissionalismo?',
    'Análise de clientes: Ayla antecipa problemas ou apenas reage?',
    'Top 3 Clientes: qualidade da análise está melhorando?',
    'Autonomia: Ayla resolve mais sozinha ou ainda depende de passo a passo?',
    'Cobranças: equipe responde melhor?',
    'Playbook de soluções: está construindo e usando?',
    'Percepção de Larissa: confiança na promoção aumentou?',
    'Feedback de clientes (informal): perceberam mudança?',
    'Gaps do Matcher: Ayla está fechando os gaps de Analista e Planejador?',
    'Ajustes necessários no PDI:',
])

r = checkpoint_block(ws5, r, 'CHECKPOINT MÊS 3 — Fim de Maio/2026', [
    'Evolução geral nas 5 competências (nota 1-5 cada):',
    'Organização: é consistente sem supervisão?',
    'Atenção: erros recorrentes foram eliminados?',
    'Comunicação: padrão profissional consolidado?',
    'Pensamento analítico: traz análise sem ser solicitada?',
    'Autonomia: resolve problemas e cobra equipe com naturalidade?',
    'Planner e checklist: rotina automatizada?',
    'Feedback de Larissa: pronta para a promoção?',
    'Feedback de clientes: percebem evolução?',
    'O que funcionou melhor no PDI?',
    'O que precisa mudar para consolidar?',
    'Matcher: as 8 competências com gap (Planejamento, Detalhismo, Organização, etc.) evoluíram?',
    'Parecer de prontidão para promoção:',
    'Recomendações de desenvolvimento futuro:',
])

# ═══════════════════════════════════════════════
# ABA 6 — SEMÁFORO (mantido)
# ═══════════════════════════════════════════════
ws6 = wb.create_sheet('Semáforo')
ws6.column_dimensions['A'].width = 22
ws6.column_dimensions['B'].width = 40
ws6.column_dimensions['C'].width = 18
ws6.column_dimensions['D'].width = 15
ws6.column_dimensions['E'].width = 15
ws6.column_dimensions['F'].width = 15

ws6.merge_cells('A1:F1')
ws6.cell(row=1, column=1, value='SEMÁFORO DE PROGRESSO — VISÃO EXECUTIVA').font = title_font
ws6['A1'].alignment = title_align

ws6.merge_cells('A2:F2')
c = ws6.cell(row=2, column=1, value='Verde = No caminho | Amarelo = Atenção | Vermelho = Intervenção necessária')
c.font = Font(size=10, color=GRAY_TEXT); c.alignment = data_align_c

set_header_row(ws6, 4, ['Dimensão','Indicador-Chave','Meta','Mês 1','Mês 2','Mês 3'])

semaforo_items = [
    ('Organização', 'Planner utilizado diariamente', 'Diário'),
    ('Organização', 'Checklist preenchido por turno', 'Diário'),
    ('Organização', 'Trello verificado 2x/turno', 'Diário'),
    ('Atenção', 'Erros de envio por semana (↓)', '0/semana'),
    ('Atenção', 'Protocolo PARE-CONFIRA-ENVIE aplicado', '100%'),
    ('Atenção', 'Pendências enviadas ao final do dia', 'Diário'),
    ('Comunicação', 'Templates criados e em uso', '5+'),
    ('Comunicação', 'Comunicação formal com clientes', '0 reclamações'),
    ('Comunicação', 'Adaptação por perfil de cliente', 'Contínuo'),
    ('Pens. Analítico', 'Análise semanal de clientes', 'Semanal'),
    ('Pens. Analítico', 'Relatório Top 3 Clientes', 'Quinzenal'),
    ('Pens. Analítico', 'Soluções registradas no playbook', '3+/mês'),
    ('Autonomia', 'Cobranças assertivas realizadas', '0 atrasos'),
    ('Autonomia', 'Tempo até pedir ajuda', '<15 min'),
    ('Autonomia', 'Iniciativa sem passo a passo', 'Crescente'),
    ('Matcher', 'Gap Analista reduzido (↓9,09%)', 'Evoluir'),
    ('Matcher', 'Gap Planejador reduzido (↓4,22%)', 'Evoluir'),
    ('Geral', 'Ações do PDI concluídas', '90%+'),
    ('Geral', 'Reuniões com Larissa realizadas', '2/mês'),
    ('Geral', 'Motivação Ayla (1-5)', '≥ 4'),
    ('Geral', 'Larissa confia na promoção?', 'Sim'),
]

for i, (dim, ind, meta) in enumerate(semaforo_items, 5):
    set_data_cell(ws6, i, 1, dim, align=data_align_l)
    set_data_cell(ws6, i, 2, ind, align=data_align_l)
    set_data_cell(ws6, i, 3, meta, align=data_align_c)

r = 5 + len(semaforo_items) + 1
ws6.cell(row=r, column=1, value='LEGENDA:').font = Font(bold=True, size=10)

r += 1
c = ws6.cell(row=r, column=1, value='VERDE')
c.fill = hfill(GREEN_BG); c.font = Font(bold=True, size=10, color=WHITE); c.alignment = data_align_c
ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
ws6.cell(row=r, column=2, value='70%+ das ações concluídas, indicadores no caminho, Larissa percebe evolução, clientes satisfeitos').font = data_font

r += 1
c = ws6.cell(row=r, column=1, value='AMARELO')
c.fill = hfill(YELLOW_BG); c.font = Font(bold=True, size=10, color=DARK_GRAY); c.alignment = data_align_c
ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
ws6.cell(row=r, column=2, value='40-69% das ações, alguns indicadores atrasados, erros ainda frequentes, organização inconsistente').font = data_font

r += 1
c = ws6.cell(row=r, column=1, value='VERMELHO')
c.fill = hfill(RED_BG); c.font = Font(bold=True, size=10, color=WHITE); c.alignment = data_align_c
ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
ws6.cell(row=r, column=2, value='<40% das ações, erros não diminuíram, modo automático persiste, Larissa não percebe evolução, promoção em risco').font = data_font

# ═══════════════════════════════════════════════
# ABA 7 — COMPETÊNCIAS REFERÊNCIA (ATUALIZADA COM MATCHER)
# ═══════════════════════════════════════════════
ws7 = wb.create_sheet('Competências Referência')
ws7.column_dimensions['A'].width = 22
ws7.column_dimensions['B'].width = 32
ws7.column_dimensions['C'].width = 55
ws7.column_dimensions['D'].width = 22
ws7.column_dimensions['E'].width = 20
ws7.column_dimensions['F'].width = 18

ws7.merge_cells('A1:F1')
ws7.cell(row=1, column=1, value='MAPEAMENTO DE COMPETÊNCIAS — BASE PARA AVALIAÇÃO').font = title_font
ws7['A1'].alignment = title_align

ws7.merge_cells('A2:F2')
c = ws7.cell(row=2, column=1, value='Fonte: 190 Competências + Contexto Vale Assessoria + Matcher Comportamental Sólides (01/03/2026)')
c.font = Font(size=10, color=GRAY_TEXT); c.alignment = data_align_c

set_header_row(ws7, 4, ['Categoria','Competência','Relevância para Ayla','Nível Atual\nEstimado (1-5)','Prioridade\nno PDI','Gap Matcher'])

comps_ref = [
    # (cat, comp, rel, nivel, nivel_color, rel_color, prio, matcher_gap)
    ('Gestão de Operações','Organização e Planejamento','CRÍTICA — esquece tarefas, não usa ferramentas adequadas. Matcher: Normal Alto → Alto exigido',1,'FFEBEE','E53935','Competência 1','↓ Cargo exige Alto'),
    ('Gestão de Operações','Gestão do Tempo','CRÍTICA — não prioriza, não fecha o dia com revisão. Planejador ↓4,22%',1,'FFEBEE','E53935','Competência 1','↓ Planejador'),
    ('Gestão de Operações','Atenção ao Detalhe','CRÍTICA — modo automático, Analista 19,48%. Matcher: Normal Alto → Alto exigido',1,'FFEBEE','E53935','Competência 2','↓ Cargo exige Alto'),
    ('Gestão de Operações','Controle de Qualidade','ALTA — erros de envio, postagens em grupos errados',2,'FFFDE7','F57C00','Competência 2','—'),
    ('Gestão de Operações','Gestão de Processos','ALTA — não segue fluxos de aprovação consistentemente',2,'FFFDE7','F57C00','Competência 1','—'),
    ('Comunicação','Comunicação Escrita','ALTA — precisa ser mais formal e detalhista',2,'FFFDE7','F57C00','Competência 3','—'),
    ('Comunicação','Comunicação Verbal','MÉDIA — boa oral, mas falta formalidade. Comunicador ↑6,25% = excesso',3,'FFF3E0',None,'Competência 3','↑ Excesso'),
    ('Comunicação','Capacidade de Ouvir','ALTA — no automático, perde informações. Matcher: Normal Alto → Alto exigido',2,'FFFDE7','F57C00','Competência 3','↓ Cargo exige Alto'),
    ('Comunicação','Empatia','ALTA — Matcher: Normal Alto → Alto exigido. Desenvolver leitura dos clientes',2,'FFFDE7','F57C00','Competência 3','↓ Cargo exige Alto'),
    ('Comunicação','Técnicas de Negociação','MÉDIA — precisa cobrar equipe e clientes assertivamente',2,'FFFDE7',None,'Competência 5','—'),
    ('Raciocínio Lógico','Pensamento Analítico','CRÍTICA — Analista ↓9,09% (maior gap do Matcher). Falta raciocínio sequencial',1,'FFEBEE','E53935','Competência 4','↓ Maior gap'),
    ('Raciocínio Lógico','Resolução de Problemas','ALTA — trava sozinha em vez de pedir ajuda',2,'FFFDE7','F57C00','Competência 5','—'),
    ('Raciocínio Lógico','Abordagem Metódica','ALTA — Matcher: Perfil Técnico Normal Alto → Alto exigido',1,'FFEBEE','F57C00','Competência 4','↓ Cargo exige Alto'),
    ('Raciocínio Lógico','Concentração','ALTA — Matcher: Normal Alto → Alto exigido. Modo automático prejudica foco',2,'FFFDE7','F57C00','Competência 2','↓ Cargo exige Alto'),
    ('Comportamental','Proatividade','ALTA — já possui, ponto forte. Executor ↑7,07% = excesso positivo',4,'E8F5E9',None,'Ativo existente','↑ Excesso positivo'),
    ('Comportamental','Autonomia','ALTA — precisa reduzir dependência de passo a passo',2,'FFFDE7','F57C00','Competência 5','—'),
    ('Comportamental','Assertividade','ALTA — dificuldade em cobrar equipe profissionalmente',2,'FFFDE7','F57C00','Competência 5','—'),
    ('Comportamental','Condescendência','MÉDIA — Matcher: Normal Alto → Alto exigido. Mais paciência com clientes',2,'FFFDE7',None,'Competência 3','↓ Cargo exige Alto'),
    ('Comportamental','Adaptabilidade','MÉDIA — demora para incorporar novos processos',2,'FFFDE7',None,'Competência 2','—'),
    ('Atendimento','Foco no Cliente','ALTA — já tem feeling, precisa mais análise',3,'FFF3E0',None,'Competência 4','—'),
    ('Atendimento','Relacionamento Interpessoal','ALTA — ponto forte (Diplomata), usar como alavanca',4,'E8F5E9',None,'Ativo existente','Alinhado'),
    ('Atendimento','Sociabilidade','MÉDIA — boa naturalmente, manter e direcionar',3,'FFF3E0',None,'Ativo existente','Alinhado'),
    ('Desenvolvimento','Vontade de Aprender','ALTA — nunca recusou atividade, busca entender',4,'E8F5E9',None,'Ativo existente','Alinhado'),
    ('Desenvolvimento','Potencial de Crescimento','ALTA — liderança vê potencial real para promoção',3,'FFF3E0',None,'Ativo existente','Alinhado'),
    ('Desenvolvimento','Comprometimento','ALTA — valoriza trabalhar na Vale, quer crescer',4,'E8F5E9',None,'Ativo existente','Alinhado'),
]

for i, (cat, comp, rel, nivel, nivel_color, rel_color, prio, matcher) in enumerate(comps_ref, 5):
    set_data_cell(ws7, i, 1, cat, align=data_align_l)
    set_data_cell(ws7, i, 2, comp, align=data_align_l)
    rel_font = Font(bold=True, size=10, color=rel_color) if rel_color else data_font
    set_data_cell(ws7, i, 3, rel, font=rel_font, align=data_align_l)
    set_data_cell(ws7, i, 4, nivel, fill=hfill(nivel_color), align=data_align_c)
    set_data_cell(ws7, i, 5, prio, align=data_align_c)
    # Matcher gap column
    matcher_fill = None
    matcher_font = data_font
    if '↓' in matcher and 'Maior' in matcher:
        matcher_fill = hfill(LIGHT_RED)
        matcher_font = Font(bold=True, size=10, color=RED_TEXT)
    elif '↓' in matcher:
        matcher_fill = hfill(LIGHT_ORANGE)
        matcher_font = Font(bold=True, size=10, color=ORANGE_TEXT)
    elif '↑' in matcher:
        matcher_fill = hfill(LIGHT_GREEN)
        matcher_font = Font(size=10, color='43A047')
    elif 'Alinhado' in matcher:
        matcher_fill = hfill(LIGHT_GREEN)
    set_data_cell(ws7, i, 6, matcher, font=matcher_font, fill=matcher_fill, align=data_align_c)

# Legenda Matcher
r = 5 + len(comps_ref) + 1
ws7.merge_cells(f'A{r}:F{r}')
ws7.cell(row=r, column=1, value='LEGENDA MATCHER:').font = Font(bold=True, size=10)
r += 1
legendas = [
    ('↓ Cargo exige Alto', LIGHT_ORANGE, 'Competência que o cargo exige em nível Alto, mas Ayla está em Normal Alto — gap a desenvolver'),
    ('↓ Maior gap', LIGHT_RED, 'Maior gap identificado pelo Matcher (Analista ↓9,09%) — prioridade máxima no PDI'),
    ('↑ Excesso', LIGHT_GREEN, 'Ayla tem mais do que o cargo exige — ativo a canalizar, não reprimir'),
    ('Alinhado', LIGHT_GREEN, 'Perfil de Ayla e exigência do cargo estão alinhados'),
]
for leg, leg_color, desc in legendas:
    set_data_cell(ws7, r, 1, leg, font=Font(bold=True, size=10), fill=hfill(leg_color), align=data_align_c)
    ws7.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    set_data_cell(ws7, r, 2, desc, align=data_align_l)
    r += 1

# ═══════════════════════════════════════════════
# ABA 8 — GUIA 1-1 (recriada do script anterior)
# ═══════════════════════════════════════════════
ws8 = wb.create_sheet('Guia 1-1')
ws8.column_dimensions['A'].width = 6
ws8.column_dimensions['B'].width = 32
ws8.column_dimensions['C'].width = 50
ws8.column_dimensions['D'].width = 50
ws8.column_dimensions['E'].width = 22
ws8.column_dimensions['F'].width = 22
ws8.column_dimensions['G'].width = 22
ws8.column_dimensions['H'].width = 22

ws8.merge_cells('A1:H1')
c = ws8.cell(row=1, column=1, value='GUIA DE REUNIÕES 1:1 — LARISSA ↔ AYLA')
c.font = title_font; c.alignment = title_align

ws8.merge_cells('A2:H2')
c = ws8.cell(row=2, column=1, value='Instruções, modelos e exemplos práticos para acompanhamento do PDI via reuniões one-on-one')
c.font = Font(size=10, color=GRAY_TEXT); c.alignment = data_align_c
ws8.row_dimensions[1].height = 32; ws8.row_dimensions[2].height = 22

r = 4
r = section_title(ws8, r, '1. O QUE É A REUNIÃO 1:1 E POR QUE FAZER')

items_oq = [
    ('Definição','É um encontro individual e recorrente entre líder e colaboradora, com foco no desenvolvimento, alinhamento e acompanhamento do PDI. Não é cobrança — é conexão.'),
    ('Quem participa','Larissa (líder) + Ayla (colaboradora). Eventualmente, Eleine pode participar nos checkpoints mensais.'),
    ('Frequência recomendada','Quinzenal (a cada 15 dias), com duração de 20-30 minutos. Pode ser semanal no primeiro mês para criar o hábito.'),
    ('Onde realizar','Preferencialmente presencial, em local reservado e sem interrupções. Se remoto: câmera ligada, sem multitarefa.'),
    ('Regra de ouro','A reunião é 90% da Ayla e 10% da Larissa. Larissa escuta, faz perguntas e orienta. Ayla traz a pauta, os desafios e as conquistas.'),
    ('Benefícios comprovados','Reduz rotatividade em 15%, aumenta engajamento em 30%, acelera o desenvolvimento, previne problemas e fortalece a confiança mútua.'),
]

for lab, desc in items_oq:
    cell(ws8, r, 1, '', fill=hfill(LIGHT_TEAL))
    cell(ws8, r, 2, lab, font=bold_font, fill=hfill(LIGHT_TEAL))
    ws8.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    cell(ws8, r, 3, desc, align=data_align_l)
    ws8.row_dimensions[r].height = 40; r += 1

r += 1
r = section_title(ws8, r, '2. BOAS PRÁTICAS PARA A 1:1')
set_header_row(ws8, r, ['', 'Fazer', 'Evitar', '', '', '', '', '']); r += 1

boas_praticas = [
    ('Agendar com antecedência e nunca cancelar sem reagendar','Cancelar em cima da hora ou "deixar pra depois"'),
    ('Começar com algo positivo — reconhecer uma conquista','Começar listando erros ou problemas'),
    ('Fazer perguntas abertas ("Como você se sentiu sobre...")','Fazer interrogatório ou perguntas fechadas ("Você fez X?")'),
    ('Escutar ativamente — 80% escuta, 20% fala (para Larissa)','Monopolizar a conversa ou dar sermão'),
    ('Anotar combinados e compromissos ao final','Sair da reunião sem ações definidas'),
    ('Conectar a conversa com as competências do PDI','Falar só de tarefas operacionais sem foco em desenvolvimento'),
    ('Criar ambiente seguro — é ok errar e dizer que está difícil','Usar a 1:1 como reunião de cobrança ou punição'),
    ('Celebrar pequenas vitórias e progressos','Só falar de problemas e gaps'),
]

for fazer, evitar in boas_praticas:
    cell(ws8, r, 1, '✓', font=Font(bold=True, size=10, color='43A047'), align=data_align_c)
    cell(ws8, r, 2, fazer, align=data_align_l, merge_end=4)
    cell(ws8, r, 5, '✗', font=Font(bold=True, size=10, color='E53935'), align=data_align_c)
    ws8.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    cell(ws8, r, 6, evitar, align=data_align_l)
    ws8.row_dimensions[r].height = 35; r += 1

r += 1
r = section_title(ws8, r, '3. MODELO DE PAUTA — ADAPTADO PARA O PDI DA AYLA')
ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cell(ws8, r, 1, 'Modelo "8 Áreas Principais" adaptado — organiza a conversa em blocos temáticos conectados às 5 competências do PDI.', font=tip_font, align=data_align_l)
r += 1

set_header_row(ws8, r, ['#', 'Bloco da Pauta', 'O que abordar', 'Exemplo de Pergunta para Ayla', 'Tempo', 'Competência PDI', '', '']); r += 1

blocos = [
    (1,'Abertura e Check-in','Como Ayla está se sentindo? Motivação, energia, dificuldades pessoais que impactam o trabalho.','"Ayla, como você está se sentindo essa quinzena? O que te animou e o que te frustrou?"','3 min','Todas'),
    (2,'Conquistas e Progressos','O que deu certo desde a última 1:1? Celebrar vitórias, mesmo pequenas.','"Me conta uma coisa que você fez diferente essa quinzena e que deu certo."','4 min','Todas'),
    (3,'Organização e Rotina','Uso do planner, checklist por turno, verificação do Trello, rotina de encerramento do dia.','"Você está conseguindo usar o planner todos os dias? Quais dias foram mais difíceis e por quê?"','4 min','Competência 1'),
    (4,'Atenção e Erros','Protocolo PARE-CONFIRA-ENVIE, diário de erros, revisão dupla de postagens.','"Quantos erros você registrou no diário essa quinzena? Notou algum padrão?"','4 min','Competência 2'),
    (5,'Comunicação com Clientes','Uso dos templates, formalidade, adaptação por perfil de cliente.','"Teve alguma situação com cliente que você sentiu dificuldade em ser mais formal? Como lidou?"','3 min','Competência 3'),
    (6,'Análise e Gestão de Clientes','Análise semanal do Trello, Top 3 Clientes, playbook de soluções.','"Quais são os 3 clientes mais críticos agora? O que você acha que pode resolver?"','4 min','Competência 4'),
    (7,'Autonomia e Cobranças','Cobrança à equipe, pedidos de ajuda, iniciativa própria.','"Nessa quinzena, teve algum momento que você resolveu algo sozinha sem precisar de passo a passo?"','3 min','Competência 5'),
    (8,'Fechamento e Compromissos','Resumir os pontos principais, definir 2-3 ações específicas até a próxima 1:1.','"Então para as próximas 2 semanas, combinamos: 1) ___, 2) ___, 3) ___. Concordamos?"','3 min','Todas'),
]

for num, bloco, abordagem, pergunta, tempo, comp in blocos:
    cell(ws8, r, 1, num, font=Font(bold=True, size=11, color=TEAL), align=data_align_c)
    cell(ws8, r, 2, bloco, font=bold_font, align=data_align_l)
    cell(ws8, r, 3, abordagem, align=data_align_l)
    cell(ws8, r, 4, pergunta, font=example_font, align=data_align_l)
    cell(ws8, r, 5, tempo, align=data_align_c)
    ws8.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    cell(ws8, r, 6, comp, align=data_align_c)
    ws8.row_dimensions[r].height = 55; r += 1

ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cell(ws8, r, 1, 'Tempo total estimado: 25-30 minutos | A pauta pode ser simplificada — o importante é manter a regularidade.', font=tip_font, align=data_align_l)
r += 1

# Seção 4 — Perguntas por competência (resumido)
r += 1
r = section_title(ws8, r, '4. BANCO DE PERGUNTAS POR COMPETÊNCIA DO PDI')
ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cell(ws8, r, 1, 'Larissa pode escolher 2-3 perguntas por reunião. Não precisa usar todas de uma vez.', font=tip_font, align=data_align_l)
r += 1

perguntas_por_comp = [
    ('Organização e Gestão de Rotina', hfill(LIGHT_BLUE), [
        'Você está conseguindo dividir suas atividades por turno (manhã/tarde)?',
        'Como está o uso do planner? Algum dia da semana é mais difícil de manter?',
        'Está verificando o Trello 2x por turno? O que encontra quando verifica?',
        'Está enviando a lista de pendências antes de sair? Isso ajudou a organizar?',
        'Qual o maior desafio de organização que você enfrentou essa quinzena?',
    ]),
    ('Atenção aos Detalhes e Saída do Automático', hfill(LIGHT_YELLOW), [
        'Quantos erros você registrou no diário essa quinzena? Teve algum padrão?',
        'O protocolo PARE-CONFIRA-ENVIE já está ficando automático ou ainda precisa lembrar?',
        'Conte um momento que você "parou" antes de agir e isso evitou um erro.',
        'Qual erro mais te incomodou e o que você fez para que não aconteça de novo?',
        'Você percebe quando está no "modo automático"? O que faz para sair?',
    ]),
    ('Comunicação Profissional e Formalidade', hfill(LIGHT_GREEN), [
        'Usou algum template de mensagem essa quinzena? Funcionou bem?',
        'Teve alguma situação onde sentiu que a comunicação poderia ter sido mais formal?',
        'Algum cliente reagiu diferente ao seu novo tom de comunicação?',
        'Você está conseguindo adaptar a linguagem conforme o perfil do cliente?',
        'Larissa precisou ajustar alguma comunicação sua essa quinzena?',
    ]),
    ('Pensamento Analítico e Gestão de Clientes', hfill(LIGHT_ORANGE), [
        'Fez a análise semanal de clientes no Trello? O que descobriu?',
        'Quais são os Top 3 Clientes mais críticos agora e por quê?',
        'Registrou alguma solução no seu playbook pessoal? Qual?',
        'Conseguiu antecipar algum problema antes que ele virasse urgência?',
        'Se eu (Larissa) perguntasse agora: "como está o cliente X?" — você saberia responder com dados?',
    ]),
    ('Autonomia e Senso de Dono', hfill(LIGHT_PURPLE), [
        'Essa quinzena, resolveu alguma coisa sozinha que antes precisaria de ajuda?',
        'Cobrou o design/equipe de forma assertiva? Como foi a reação?',
        'Teve algum momento que travou em uma tarefa? Quanto tempo levou para pedir ajuda?',
        'O que você faria diferente se pudesse refazer essa quinzena?',
        'Se eu não estivesse disponível por 1 dia inteiro, como seria o dia da operação?',
    ]),
]

for comp_name, comp_fill, perguntas in perguntas_por_comp:
    ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    cell(ws8, r, 1, comp_name, font=Font(bold=True, size=10, color=DARK_TEAL), fill=comp_fill, align=Alignment(horizontal='left', vertical='center'))
    ws8.row_dimensions[r].height = 24; r += 1
    for p in perguntas:
        cell(ws8, r, 1, '•', align=data_align_c)
        ws8.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        cell(ws8, r, 2, p, align=data_align_l)
        ws8.row_dimensions[r].height = 22; r += 1
    r += 1

# Seção 5 — Feedbacks
r = section_title(ws8, r, '5. EXEMPLOS DE FEEDBACK — BASEADOS NO CONTEXTO DA AYLA')
ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cell(ws8, r, 1, 'Modelos de feedback construtivo e positivo que Larissa pode adaptar. Sempre ser específica, não genérica.', font=tip_font, align=data_align_l)
r += 1
set_header_row(ws8, r, ['', 'Situação', 'Tipo', 'Exemplo de Feedback', '', '', '', '']); r += 1

feedbacks = [
    ('Ayla usou o planner a semana toda sem falhar','Positivo','"Ayla, percebi que você usou o planner toda a semana e nenhuma tarefa ficou esquecida. Isso mostra que o hábito está se formando. Continue assim — a organização está fazendo diferença nas entregas."'),
    ('Ayla postou conteúdo no grupo errado de um cliente','Construtivo','"Ayla, vi que o post do cliente X foi para o grupo errado. Entendo que foi correria, mas lembra do protocolo PARE-CONFIRA-ENVIE? O que você acha que faltou nessa hora? Como podemos evitar que aconteça de novo?"'),
    ('Ayla cobrou o design proativamente e evitou atraso','Positivo','"Ayla, gostei muito que você cobrou o design antes do prazo apertar. Isso é senso de dono na prática. O cliente nem percebeu que poderia ter atrasado. Esse é o tipo de atitude que faz diferença."'),
    ('Comunicação com cliente foi muito informal','Construtivo','"Ayla, a mensagem para o cliente Y ficou um pouco informal demais — parecia conversa entre amigas. Para esse tipo de comunicação, use o template de follow-up que criamos. Quer que a gente revise juntas como adaptar?"'),
    ('Ayla trouxe análise Top 3 Clientes com sugestões','Positivo','"Ayla, sua análise dos Top 3 Clientes ficou excelente. Você identificou o gargalo do cliente Z antes de eu precisar perguntar. Isso é pensamento analítico — exatamente o que o seu PDI está desenvolvendo."'),
    ('Ayla ficou travada 40 min sem pedir ajuda','Construtivo','"Ayla, vi que você ficou quase 40 minutos tentando resolver sozinha o problema do Trello. Lembra que combinamos que após 15 minutos sem avançar, é hora de pedir ajuda? Não é fraqueza — é eficiência. O que te impediu de chamar?"'),
    ('Ayla não enviou pendências ao final do dia','Construtivo','"Ayla, ontem não recebi a lista de pendências no final do dia. Isso me deixa sem visibilidade do que ficou em aberto. Conseguimos pensar num lembrete — alarme no celular ou item fixo no checklist — para garantir que isso não escape?"'),
    ('Ayla adaptou a linguagem para cliente executor','Positivo','"Ayla, percebi que com o cliente W você foi mais direta e objetiva na mensagem — sem rodeios, direto ao ponto. Isso é adaptar comunicação ao perfil do cliente, exatamente o que discutimos no DISC. Muito bom!"'),
    ('Ayla identificou sozinha um problema recorrente','Positivo','"Ayla, quando você me trouxe que o cliente V tem sempre o mesmo gargalo na aprovação e sugeriu mudar o dia de envio, isso mostrou que você está pensando de forma analítica. Está construindo seu playbook na prática."'),
    ('Erros diminuíram mas ainda acontecem por pressa','Construtivo + Incentivo','"Ayla, seus erros reduziram bastante — de 5 por semana para 2. Isso é evolução real. Mas os que ainda acontecem parecem ser por pressa. O que acha de nos próximos 15 dias focar especificamente no momento de maior pressa do dia e aplicar o protocolo ali?"'),
]

for sit, tipo, fb in feedbacks:
    tipo_fill = hfill(LIGHT_GREEN) if 'Positivo' in tipo else hfill(LIGHT_YELLOW)
    cell(ws8, r, 1, '', fill=tipo_fill)
    cell(ws8, r, 2, sit, font=bold_font, align=data_align_l)
    cell(ws8, r, 3, tipo, fill=tipo_fill, font=Font(bold=True, size=10), align=data_align_c)
    ws8.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
    cell(ws8, r, 4, fb, font=example_font, align=data_align_l)
    ws8.row_dimensions[r].height = 65; r += 1

# Seção 6 — Registro
r += 1
r = section_title(ws8, r, '6. REGISTRO DAS REUNIÕES 1:1 (PREENCHER A CADA ENCONTRO)')
set_header_row(ws8, r, ['#', 'Data', 'Competência em foco', 'Principais pontos discutidos', 'Compromissos assumidos', 'Prazo', 'Status', 'Observações']); r += 1

for i in range(1, 9):
    cell(ws8, r, 1, i, font=Font(bold=True, size=10, color=TEAL), align=data_align_c)
    cell(ws8, r, 2, '___/___/___', align=data_align_c)
    for col in range(3, 9): cell(ws8, r, col, '', align=data_align_l)
    ws8.row_dimensions[r].height = 50; r += 1

# Seção 7 — Calendário
r += 1
r = section_title(ws8, r, '7. CALENDÁRIO SUGERIDO DE 1:1 — MARÇO A JUNHO/2026')
set_header_row(ws8, r, ['#', 'Data sugerida', 'Foco principal', 'Competência PDI', 'Perguntas sugeridas', 'Tipo de feedback', '', '']); r += 1

calendario = [
    (1,'2ª semana Mar','Onboarding da 1:1 + como Ayla está se sentindo com o PDI','Todas','Check-in geral, expectativas, medos e motivações','Incentivo e acolhimento'),
    (2,'4ª semana Mar','Organização e planner + primeiros erros registrados','Comp. 1 e 2','Uso do planner, checklist, diário de erros','Construtivo + positivo'),
    (3,'2ª semana Abr','Comunicação + primeiros templates + Top 3 Clientes #1','Comp. 3 e 4','Templates em uso, formalidade, análise de clientes','Positivo sobre evolução'),
    (4,'4ª semana Abr','CHECKPOINT MÊS 2 — revisão geral + autonomia','Todas + Comp. 5','Evolução geral, cobranças, iniciativa','Balanço construtivo'),
    (5,'2ª semana Mai','Consolidação de hábitos + pensamento analítico','Comp. 1, 2 e 4','Hábitos automáticos? Antecipa problemas?','Celebrar hábitos formados'),
    (6,'4ª semana Mai','CHECKPOINT MÊS 3 — prontidão para promoção','Todas','Visão de Larissa, clientes, maturidade','Feedback formal de progresso'),
    (7,'2ª semana Jun','Autonomia total + consolidação','Comp. 5','Funciona sozinha por 1 dia? Cobra equipe?','Positivo ou plano de ação'),
    (8,'4ª semana Jun','FECHAMENTO — avaliação final + parecer de promoção','Todas','Nota geral, evolução, parecer final','Relatório final'),
]

for num, data, foco, comp, pergs, fb_tipo in calendario:
    is_cp = 'CHECKPOINT' in foco or 'FECHAMENTO' in foco
    row_fill = hfill(LIGHT_ORANGE) if is_cp else None
    cell(ws8, r, 1, num, font=Font(bold=True, size=10, color=TEAL), fill=row_fill, align=data_align_c)
    cell(ws8, r, 2, data, font=bold_font, fill=row_fill, align=data_align_c)
    cell(ws8, r, 3, foco, fill=row_fill, align=data_align_l)
    cell(ws8, r, 4, comp, fill=row_fill, align=data_align_c)
    cell(ws8, r, 5, pergs, fill=row_fill, align=data_align_l)
    ws8.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    cell(ws8, r, 6, fb_tipo, fill=row_fill, align=data_align_l)
    ws8.row_dimensions[r].height = 45; r += 1

# Seção 8 — Dicas
r += 1
r = section_title(ws8, r, '8. DICAS RÁPIDAS PARA LARISSA')

dicas = [
    'Antes da reunião: revise as ações combinadas na 1:1 anterior e anote 2-3 pontos que observou na quinzena.',
    'No início: comece com algo positivo — reconheça um esforço, mesmo que pequeno. Isso ativa a abertura emocional.',
    'Durante: faça mais perguntas do que afirmações. "O que você acha?" é mais poderoso que "Você deveria...".',
    'Sobre erros: não julgue — explore. Pergunte "o que aconteceu?" antes de "por que você fez isso?".',
    'Sobre o DISC: lembre que Ayla é Comunicadora Executora. Ela responde bem a desafios e reconhecimento, mas não a críticas secas.',
    'Autoestima BAIXA: cuidado com feedback apenas negativo. Sempre equilibre: para cada ajuste, reconheça uma conquista.',
    'Matcher 86%: os gaps de Analista e Planejador são os focos do PDI. Celebre quando Ayla demonstrar essas competências.',
    'Registre tudo: os registros desta aba alimentam os checkpoints e o relatório final de evolução.',
    'Consistência > perfeição: é melhor uma 1:1 de 15 minutos do que nenhuma. Nunca pule sem reagendar.',
]

for dica in dicas:
    cell(ws8, r, 1, '→', font=Font(bold=True, size=11, color=TEAL), align=data_align_c)
    ws8.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    cell(ws8, r, 2, dica, align=data_align_l)
    ws8.row_dimensions[r].height = 28; r += 1

r += 1
ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cell(ws8, r, 1, 'Guia elaborado pela CIH Consultoria em Inteligência Humana | Consultora: Eleine Passos | Fontes: Convenia, Feedz, Qulture.Rocks',
     font=Font(size=9, italic=True, color=GRAY_TEXT), align=data_align_c)

# ═══════════════════════════════════════════════
# SALVAR
# ═══════════════════════════════════════════════
wb.save(output)
print(f'PDI v2 salvo com sucesso em: {output}')
print('Atualizações com Matcher:')
print('  1. Resumo PDI: adicionados Match 86%, perfil exigido AP, gaps DISC, 8 competências com gap')
print('  2. Contexto: gap atualizado com dados do Matcher (Analista ↓9,09%)')
print('  3. Checkpoints: perguntas sobre gaps do Matcher nos meses 2 e 3')
print('  4. Semáforo: 2 novos indicadores de Matcher (Analista, Planejador)')
print('  5. Competências Referência: nova coluna "Gap Matcher" + legenda + 3 novas competências')
print('  6. Guia 1-1: dica sobre Matcher 86% adicionada')
print('  7. Resultado e Evolução: pergunta sobre evolução dos gaps do Matcher')
