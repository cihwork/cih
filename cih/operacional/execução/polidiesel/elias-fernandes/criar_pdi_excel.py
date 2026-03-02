import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

wb = openpyxl.Workbook()

# ========== CORES E ESTILOS ==========
TEAL = "0097A7"
DARK_TEAL = "00796B"
LIGHT_TEAL = "E0F2F1"
ORANGE = "F57C00"
LIGHT_ORANGE = "FFF3E0"
RED = "E53935"
LIGHT_RED = "FFEBEE"
GREEN = "43A047"
LIGHT_GREEN = "E8F5E9"
YELLOW = "FFC107"
LIGHT_YELLOW = "FFFDE7"
GRAY = "F5F5F5"
DARK_GRAY = "424242"
WHITE = "FFFFFF"
BLUE_70 = "E3F2FD"
YELLOW_20 = "FFF8E1"
PURPLE_10 = "F3E5F5"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
subheader_font = Font(name="Calibri", bold=True, color=DARK_GRAY, size=10)
subheader_fill = PatternFill(start_color=LIGHT_TEAL, end_color=LIGHT_TEAL, fill_type="solid")
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)
title_font = Font(name="Calibri", bold=True, color=TEAL, size=14)
subtitle_font = Font(name="Calibri", bold=True, color=DARK_TEAL, size=12)

thin_border = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
wrap_align = Alignment(wrap_text=True, vertical="top")

green_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
yellow_fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
red_fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
orange_fill = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid")
gray_fill = PatternFill(start_color=GRAY, end_color=GRAY, fill_type="solid")


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def style_data_cell(ws, row, col, font=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col)
    cell.font = font or normal_font
    if fill:
        cell.fill = fill
    cell.alignment = align or left_align
    cell.border = thin_border
    return cell


# =============================================
# ABA 1: RESUMO DO PDI (REVISADO)
# =============================================
ws1 = wb.active
ws1.title = "Resumo PDI"
ws1.sheet_properties.tabColor = TEAL

ws1.merge_cells("A1:H1")
ws1["A1"].value = "PDI - ELIAS FERNANDES DOS SANTOS | POLIDIESEL"
ws1["A1"].font = title_font
ws1["A1"].alignment = center_align

ws1.merge_cells("A2:H2")
ws1["A2"].value = "Plano de Desenvolvimento Individual — Ciclo Março a Agosto/2026 (v2 — revisado com kick-off)"
ws1["A2"].font = subtitle_font
ws1["A2"].alignment = center_align

info = [
    ("Colaborador", "Elias Fernandes dos Santos"),
    ("Cargo", "Responsável Técnico / Supervisor de Mecânica (em transição)"),
    ("Empresa", "Polidiesel"),
    ("Tempo de Casa", "~20 anos"),
    ("Perfil DISC", "Executor (D) — Autocrata"),
    ("Liderança Direta", "Caio Freitas (sócio — responsável pelo acompanhamento do PDI)"),
    ("Consultora CIH", "Eleine Passos"),
    ("Equipe", "7 colaboradores (área mecânica)"),
    ("Ciclo PDI", "Março/2026 a Agosto/2026 (6 meses)"),
    ("Confiança de Caio", "10/10 em confiança | 9/10 em expectativa de evolução"),
]

row = 4
for label, value in info:
    ws1.cell(row=row, column=1, value=label).font = bold_font
    ws1.cell(row=row, column=1).fill = PatternFill(start_color=LIGHT_TEAL, end_color=LIGHT_TEAL, fill_type="solid")
    ws1.cell(row=row, column=1).border = thin_border
    ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws1.cell(row=row, column=2, value=value).font = normal_font
    ws1.cell(row=row, column=2).border = thin_border
    row += 1

# Contexto do Kick-off
row += 1
ws1.merge_cells(f"A{row}:H{row}")
ws1.cell(row=row, column=1, value="CONTEXTO DA TRANSIÇÃO (KICK-OFF COM CAIO)").font = subtitle_font
row += 1

contexto_items = [
    ("Papel definido por Caio", "Responsável técnico direcionado ao administrativo de mecânica. Foco em DIAGNÓSTICO, não execução. Intermediar recepção e área técnica."),
    ("Expectativa de Caio", "Elias conduzir reuniões de produção, ser ponto focal técnico, revisar/assinar relatórios diários, dar direcionamento de tarefas ao time."),
    ("Gap identificado", "Área mecânica ficou sem liderança após redução de pessoal. Diagnósticos errados e informações perdidas sem um responsável."),
    ("Ferramentas em uso", "Planilha de gerenciamento para serviços externos (em teste há 3 meses). Plano de TV para gestão à vista. Sugestão de tablet para campo."),
    ("Risco de dependência", "Caio reconhece risco: Elias é insubstituível tecnicamente. Há 2 anos quase saiu. Precisa formar pelo menos 1 backup."),
    ("Responsabilidade 50/50", "Desenvolvimento é 50% do colaborador e 50% da empresa (Eleine). Caio fornecerá suporte (acesso, tempo, recursos)."),
]

for label, value in contexto_items:
    ws1.cell(row=row, column=1, value=label).font = bold_font
    ws1.cell(row=row, column=1).border = thin_border
    ws1.cell(row=row, column=1).fill = gray_fill
    ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    ws1.cell(row=row, column=2, value=value).font = normal_font
    ws1.cell(row=row, column=2).border = thin_border
    ws1.cell(row=row, column=2).alignment = wrap_align
    ws1.row_dimensions[row].height = 35
    row += 1

# Competências resumo - agora 4
row += 1
ws1.merge_cells(f"A{row}:H{row}")
ws1.cell(row=row, column=1, value="COMPETÊNCIAS PRIORIZADAS (4 COMPETÊNCIAS)").font = subtitle_font
row += 1

headers_comp = ["#", "Competência", "Nível Atual", "Meta", "Peso", "Status Mês 1", "Status Mês 2", "Status Mês 3"]
for col, h in enumerate(headers_comp, 1):
    ws1.cell(row=row, column=col, value=h)
style_header_row(ws1, row, 8)

competencias = [
    (1, "Escuta Ativa e Comunicação Adaptada ao Perfil da Equipe", 2, 3, "30%", "", "", ""),
    (2, "Transição de Executor para Líder (Soltar a Execução)", 1, 3, "30%", "", "", ""),
    (3, "Gestão Técnica e Formação da Equipe", 1, 3, "25%", "", "", ""),
    (4, "Inteligência Emocional e Gestão de Energia", 2, 3, "15%", "", "", ""),
]

for comp in competencias:
    row += 1
    for col, val in enumerate(comp, 1):
        cell = style_data_cell(ws1, row, col)
        cell.value = val
        if col == 3 and isinstance(val, int):
            cell.alignment = center_align
            cell.fill = red_fill if val <= 1 else yellow_fill
        elif col == 4 and isinstance(val, int):
            cell.alignment = center_align
            cell.fill = green_fill
        elif col == 5:
            cell.alignment = center_align
        elif col >= 6:
            cell.alignment = center_align

# Indicadores DISC
row += 2
ws1.merge_cells(f"A{row}:H{row}")
ws1.cell(row=row, column=1, value="INDICADORES DISC — PONTOS DE ATENÇÃO").font = subtitle_font
row += 1

headers_disc = ["Indicador", "Valor", "Classificação", "Impacto", "Ação Recomendada"]
for col, h in enumerate(headers_disc, 1):
    ws1.cell(row=row, column=col, value=h)
style_header_row(ws1, row, 5)

disc_data = [
    ("Energia (EN)", 6.9, "Baixa", "Capacidade reduzida de absorver mudanças", "Mapa de energia + rituais de recarga"),
    ("Moral (MRL)", 1.2, "Normal Baixo", "Sente que mudanças devem ocorrer", "Alinhar expectativas com Caio"),
    ("Flexibilidade (IF)", -4.76, "Normal Baixo", "Dificuldade para adaptar comportamento", "Exposição gradual a novas abordagens"),
    ("Energia do Perfil (ENP)", 72.22, "Extremamente Alta", "Potencial enorme represado", "Canalizar em desafios de liderança"),
    ("Exigência do Meio (IEM)", 66.76, "Normal Alto", "Ambiente cobra adaptações", "Reduzir gap com desenvolvimento"),
    ("Positividade (IP)", 25.0, "Extremamente Alto", "Autoestima forte — recurso positivo", "Usar como alavanca para desenvolvimento"),
]

for d in disc_data:
    row += 1
    for col, val in enumerate(d, 1):
        cell = style_data_cell(ws1, row, col)
        cell.value = val
        if col == 3:
            if "Baixa" in str(val) or "Baixo" in str(val):
                cell.fill = orange_fill
            elif "Alta" in str(val) or "Alto" in str(val):
                cell.fill = green_fill

ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 55
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 12
ws1.column_dimensions["E"].width = 45
ws1.column_dimensions["F"].width = 15
ws1.column_dimensions["G"].width = 15
ws1.column_dimensions["H"].width = 15

# =============================================
# ABA 2: PLANO DE AÇÕES DETALHADO (REVISADO - 20 AÇÕES)
# =============================================
ws2 = wb.create_sheet("Plano de Ações")
ws2.sheet_properties.tabColor = DARK_TEAL

ws2.merge_cells("A1:L1")
ws2["A1"].value = "PLANO DE AÇÕES — PDI ELIAS FERNANDES (v2 — 20 ações, 4 competências)"
ws2["A1"].font = title_font
ws2["A1"].alignment = center_align

row = 3
headers_acoes = [
    "#", "Competência", "Tipo 70-20-10", "Ação", "Descrição Detalhada",
    "Prazo Início", "Prazo Fim", "Frequência", "Indicador de Sucesso",
    "Responsável", "Fonte", "Status"
]
for col, h in enumerate(headers_acoes, 1):
    ws2.cell(row=row, column=col, value=h)
style_header_row(ws2, row, 12)

acoes = [
    # === Competência 1: Escuta Ativa e Comunicação ===
    (1, "Escuta Ativa e Comunicação", "70% Experiência", "Reuniões individuais mensais",
     "20-30 min com cada um dos 7 membros. Roteiro: Como você está? O que está difícil? O que posso fazer? Mais escutar, menos falar.",
     "Mar/2026", "Ago/2026", "Mensal", "7 reuniões/mês com anotações", "Elias", "DISC + Devolutiva", "Pendente"),
    (2, "Escuta Ativa e Comunicação", "70% Experiência", "Conduzir reunião diária de produção",
     "Assumir a condução das reuniões de produção com o time mecânico. Descrever ordens de serviço em aberto, definir prioridades, dar direcionamento de tarefas. Caio observa e depois dá feedback.",
     "Mar/2026", "Ago/2026", "Diário", "Reunião conduzida por Elias 5x/semana", "Elias", "Kick-off (Caio)", "Pendente"),
    (3, "Escuta Ativa e Comunicação", "70% Experiência", "Mapa de perfis do time",
     "Com apoio de Eleine, identificar perfil DISC de cada membro. Criar cola com estilo de comunicação ideal por pessoa.",
     "Mar/2026", "Abr/2026", "Única", "Mapa completo dos 7 perfis", "Elias + Eleine", "DISC", "Pendente"),
    (4, "Escuta Ativa e Comunicação", "70% Experiência", "Regra dos 10 segundos",
     "Após fazer pergunta em reunião, contar até 10 mentalmente antes de responder. Dar espaço para o time. Especialmente importante porque Elias é 'de poucas palavras' e tende a encurtar diálogos.",
     "Mar/2026", "Ago/2026", "Diário", "Auto-avaliação semanal", "Elias", "DISC + Kick-off", "Pendente"),
    (5, "Escuta Ativa e Comunicação", "20% Social", "Mentoria semanal com Caio",
     "15-20 min para discutir situações com a equipe: como eu lidaria vs. como devo lidar. Caio compartilha como ele próprio aprendeu a se comunicar.",
     "Mar/2026", "Ago/2026", "Semanal", "4+ conversas/mês registradas", "Elias + Caio", "DISC + Kick-off", "Pendente"),
    (6, "Escuta Ativa e Comunicação", "10% Formal", "Estudo dos perfis DISC",
     "Ler material completo do resultado DISC (Caio imprime e entrega). Entender diferenças entre Comunicador, Executor, Analista e Planejador.",
     "Mar/2026", "Mar/2026", "Única", "Leitura concluída + 3 insights", "Elias", "DISC + Devolutiva", "Pendente"),

    # === Competência 2: Transição Executor → Líder ===
    (7, "Transição Executor → Líder", "70% Experiência", "Sabatina semanal com o time",
     "Apresentar demanda/problema e pedir que o time traga soluções. Elias observa, questiona e orienta — NÃO executa. Foco: fazer o time pensar e resolver.",
     "Mar/2026", "Ago/2026", "Semanal", "1 sabatina/semana com registro", "Elias", "Devolutiva", "Pendente"),
    (8, "Transição Executor → Líder", "70% Experiência", "Diário 'Quase Meti a Mão'",
     "Registrar toda vez que sentir o impulso de pegar na execução. Anotar: situação, sentimento, o que fez em vez de executar. Foco em DIAGNÓSTICO, não execução (conforme definido por Caio no kick-off).",
     "Mar/2026", "Mai/2026", "Diário", "Redução progressiva dos episódios", "Elias", "DISC + Kick-off", "Pendente"),
    (9, "Transição Executor → Líder", "70% Experiência", "Delegação com ensino",
     "Ao delegar: 1) Explicar O QUÊ; 2) Perguntar COMO faria; 3) Complementar; 4) Acompanhar sem fazer. Se resultado não for perfeito, usar como oportunidade de ensino — não refazer.",
     "Mar/2026", "Ago/2026", "Contínuo", "3 delegações/semana neste formato", "Elias", "Devolutiva", "Pendente"),
    (10, "Transição Executor → Líder", "70% Experiência", "Foco no diagnóstico técnico",
     "Assumir o papel de ponto focal técnico: receber as demandas da recepção, diagnosticar, e DIRECIONAR o mecânico adequado para a execução. Encurtar a distância recepção ↔ área técnica, sem executar o serviço.",
     "Mar/2026", "Ago/2026", "Contínuo", "Elias é consultado em 80%+ dos diagnósticos", "Elias", "Kick-off (Caio)", "Pendente"),
    (11, "Transição Executor → Líder", "20% Social", "Conversa com Caio sobre delegação",
     "Discutir situações onde é difícil soltar a execução. Buscar feedback de como Caio aprendeu a delegar. Caio compartilha quais tarefas já pode 'soltar' para Elias decidir.",
     "Mar/2026", "Ago/2026", "Quinzenal", "2-3 tarefas delegadas por mês", "Elias + Caio", "Kick-off", "Pendente"),
    (12, "Transição Executor → Líder", "10% Formal", "Sessão Líder Multiplicador",
     "Eleine apresenta em sessão de 30min: líder não faz melhor que o time — faz o time fazer melhor. Referência: 'O trabalho do líder é trabalhar o trabalho' (Liz Wiseman).",
     "Mar/2026", "Mar/2026", "Única", "Sessão realizada + 3 princípios", "Eleine", "Devolutiva", "Pendente"),

    # === Competência 3: Gestão Técnica e Formação da Equipe (NOVA) ===
    (13, "Gestão Técnica e Formação", "70% Experiência", "Revisão e assinatura de relatórios diários",
     "Revisar e assinar os relatórios diários de cada mecânico. Garantir que está a par de todos os serviços. Usar como instrumento de acompanhamento e gestão, não apenas burocracia.",
     "Mar/2026", "Ago/2026", "Diário", "100% dos relatórios revisados e assinados", "Elias", "Kick-off (Caio)", "Pendente"),
    (14, "Gestão Técnica e Formação", "70% Experiência", "Avaliação técnica e comportamental da equipe",
     "Avaliar cada mecânico em 2 dimensões: competência técnica e postura/atitude. Criar ranking para alocar recursos de forma eficaz. Usar termo 'avaliação' (não 'julgamento', conforme orientação de Eleine).",
     "Abr/2026", "Mai/2026", "Mensal", "Ficha de avaliação preenchida para os 7", "Elias + Eleine", "Kick-off (Caio + Eleine)", "Pendente"),
    (15, "Gestão Técnica e Formação", "70% Experiência", "Identificar e desenvolver um potencial sucessor",
     "Elias identifica pelo menos 1 mecânico com potencial para cobrir suas funções (inicialmente para férias). Começa a envolvê-lo em diagnósticos e decisões. Formar alguém reduz o risco de dependência técnica que preocupa Caio.",
     "Abr/2026", "Ago/2026", "Contínuo", "1 pessoa sendo desenvolvida ativamente", "Elias + Caio", "Kick-off (Eleine)", "Pendente"),
    (16, "Gestão Técnica e Formação", "70% Experiência", "Gestão da planilha de serviços",
     "Assumir a atualização e análise da planilha de gerenciamento de serviços externos. Usar para tomada de decisão sobre alocação de mão de obra, priorização e prazos. Reportar para Caio com análise, não apenas dados.",
     "Mar/2026", "Ago/2026", "Diário", "Planilha atualizada diariamente com análise semanal", "Elias", "Kick-off (Caio)", "Pendente"),
    (17, "Gestão Técnica e Formação", "20% Social", "Observar Caio fazendo gestão",
     "Acompanhar Caio em momentos de avaliação de desempenho, tomada de decisão sobre equipe, e feedback. Aprender o 'olhar de gestão' observando quem já faz. Eleine orienta: 'A melhor forma de aprender é observando Caio fazer com ele.'",
     "Mar/2026", "Mai/2026", "Quinzenal", "4+ situações observadas com reflexão", "Elias + Caio", "Kick-off (Eleine)", "Pendente"),

    # === Competência 4: Inteligência Emocional ===
    (18, "Inteligência Emocional", "70% Experiência", "Regra Pause & Reframe",
     "Em frustração/conflito, pausar e perguntar: 'Como quero que essa pessoa saia se sentindo dessa conversa?' Reformular se necessário.",
     "Mar/2026", "Ago/2026", "Diário", "3+ situações/semana registradas", "Elias", "DISC", "Pendente"),
    (19, "Inteligência Emocional", "70% Experiência", "Mapa de energia semanal",
     "Registrar atividades que dão energia (+) e drenam (-). Identificar padrões em 4 semanas. Ajustar rotina para proteger energia (EN=6,9 — baixa).",
     "Mar/2026", "Abr/2026", "Semanal", "Mapa com 3+ padrões identificados", "Elias", "DISC", "Pendente"),
    (20, "Inteligência Emocional", "20% Social", "Feedback anônimo do time",
     "Eleine coleta anonimamente a cada 2 meses: O que melhorou? O que ainda incomoda? Perguntas sobre comunicação, escuta, clima e liderança.",
     "Abr/2026", "Ago/2026", "Bimestral", "Evolução entre 1ª e 2ª coleta", "Eleine", "DISC + Devolutiva", "Pendente"),
]

for a in acoes:
    row += 1
    for col, val in enumerate(a, 1):
        cell = style_data_cell(ws2, row, col)
        cell.value = val
        if col == 3:
            if "70%" in str(val):
                cell.fill = PatternFill(start_color=BLUE_70, end_color=BLUE_70, fill_type="solid")
            elif "20%" in str(val):
                cell.fill = PatternFill(start_color=YELLOW_20, end_color=YELLOW_20, fill_type="solid")
            elif "10%" in str(val):
                cell.fill = PatternFill(start_color=PURPLE_10, end_color=PURPLE_10, fill_type="solid")
        if col == 11:
            cell.fill = gray_fill
        if col == 12:
            cell.alignment = center_align

ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 35
ws2.column_dimensions["E"].width = 60
ws2.column_dimensions["F"].width = 12
ws2.column_dimensions["G"].width = 12
ws2.column_dimensions["H"].width = 12
ws2.column_dimensions["I"].width = 38
ws2.column_dimensions["J"].width = 15
ws2.column_dimensions["K"].width = 18
ws2.column_dimensions["L"].width = 12

# =============================================
# ABA 3: ACOMPANHAMENTO SEMANAL (3 MESES) — REVISADO
# =============================================
ws3 = wb.create_sheet("Acompanhamento Semanal")
ws3.sheet_properties.tabColor = ORANGE

ws3.merge_cells("A1:K1")
ws3["A1"].value = "ACOMPANHAMENTO SEMANAL — 3 MESES (MARÇO A MAIO/2026)"
ws3["A1"].font = title_font
ws3["A1"].alignment = center_align

row = 3
headers_acomp = [
    "Semana", "Período", "Foco Principal",
    "Escuta Ativa\n(Nota 1-5)", "Transição Líder\n(Nota 1-5)",
    "Gestão Técnica\n(Nota 1-5)", "Intelig. Emocional\n(Nota 1-5)",
    "Ações Realizadas", "Dificuldades / Bloqueios",
    "Próximos Passos", "Observações do Consultor"
]
for col, h in enumerate(headers_acomp, 1):
    ws3.cell(row=row, column=col, value=h)
style_header_row(ws3, row, 11)

start_date = datetime(2026, 3, 2)
focos = {
    1: "Leitura DISC + Conduzir 1ª reunião de produção + Início reuniões individuais + Mapa de energia",
    2: "1ª sabatina + Regra dos 10s + Delegação com ensino + Revisão de relatórios diários",
    3: "Sessão Líder Multiplicador + Foco em diagnóstico (não execução) + Gestão da planilha",
    4: "CHECKPOINT MÊS 1 — Avaliar primeiras semanas com Eleine + Caio avalia condução de reuniões",
    5: "Consolidar sabatinas + Aprofundar reuniões individuais + Iniciar avaliação técnica do time",
    6: "Mapa de perfis completo + Início leitura Goleman + Identificar potencial sucessor",
    7: "1ª coleta feedback anônimo do time + Observar Caio fazendo gestão",
    8: "CHECKPOINT MÊS 2 — Avaliar evolução com Eleine + Caio + Avaliar planilha e relatórios",
    9: "Avaliar padrões do mapa de energia + Ficha de avaliação dos 7 mecânicos",
    10: "Ajustar ações que drenam energia + Aprofundar formação do sucessor",
    11: "Aprofundar delegação — identificar tarefas operacionais para soltar definitivamente",
    12: "Consolidar gestão técnica + Preparar evidências de evolução",
    13: "CHECKPOINT MÊS 3 — Revisão geral + Ajuste do PDI + Feedback de Caio sobre condução de reuniões",
}

for week in range(1, 14):
    row += 1
    week_start = start_date + timedelta(weeks=week - 1)
    week_end = week_start + timedelta(days=4)
    periodo = f"{week_start.strftime('%d/%m')} a {week_end.strftime('%d/%m')}"
    foco = focos.get(week, "")

    ws3.cell(row=row, column=1, value=f"Semana {week}").font = bold_font
    ws3.cell(row=row, column=1).alignment = center_align
    ws3.cell(row=row, column=1).border = thin_border
    ws3.cell(row=row, column=2, value=periodo).font = normal_font
    ws3.cell(row=row, column=2).alignment = center_align
    ws3.cell(row=row, column=2).border = thin_border
    ws3.cell(row=row, column=3, value=foco).font = normal_font
    ws3.cell(row=row, column=3).alignment = wrap_align
    ws3.cell(row=row, column=3).border = thin_border

    for col in range(4, 12):
        cell = ws3.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = center_align if col <= 7 else wrap_align

    if week in [4, 8, 13]:
        for col in range(1, 12):
            ws3.cell(row=row, column=col).fill = PatternFill(
                start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid"
            )

    ws3.row_dimensions[row].height = 45

ws3.column_dimensions["A"].width = 12
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 55
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 14
ws3.column_dimensions["F"].width = 14
ws3.column_dimensions["G"].width = 14
ws3.column_dimensions["H"].width = 35
ws3.column_dimensions["I"].width = 35
ws3.column_dimensions["J"].width = 30
ws3.column_dimensions["K"].width = 30

# =============================================
# ABA 4: RESULTADO E EVOLUÇÃO — REVISADO
# =============================================
ws4 = wb.create_sheet("Resultado e Evolução")
ws4.sheet_properties.tabColor = GREEN

ws4.merge_cells("A1:H1")
ws4["A1"].value = "RESULTADO DO PDI — EVOLUÇÃO DAS COMPETÊNCIAS"
ws4["A1"].font = title_font
ws4["A1"].alignment = center_align

row = 3
headers_evo = ["Competência", "Nível Inicial", "Meta", "Mês 1", "Mês 2", "Mês 3", "Nível Final", "Atingiu?"]
for col, h in enumerate(headers_evo, 1):
    ws4.cell(row=row, column=col, value=h)
style_header_row(ws4, row, 8)

comp_evo = [
    ("Escuta Ativa e Comunicação Adaptada", 2, 3, "", "", "", "", ""),
    ("Transição Executor → Líder", 1, 3, "", "", "", "", ""),
    ("Gestão Técnica e Formação da Equipe", 1, 3, "", "", "", "", ""),
    ("Inteligência Emocional e Gestão de Energia", 2, 3, "", "", "", "", ""),
]

for c in comp_evo:
    row += 1
    for col, val in enumerate(c, 1):
        cell = style_data_cell(ws4, row, col)
        cell.value = val
        if col >= 2:
            cell.alignment = center_align
        if col == 2 and isinstance(val, int):
            cell.fill = red_fill if val <= 1 else yellow_fill
        if col == 3 and isinstance(val, int):
            cell.fill = green_fill

# Indicadores de acompanhamento - REVISADOS
row += 2
ws4.merge_cells(f"A{row}:H{row}")
ws4.cell(row=row, column=1, value="INDICADORES DE ACOMPANHAMENTO").font = subtitle_font
row += 1

headers_ind = ["Indicador", "Baseline", "Meta Mês 1", "Real Mês 1", "Meta Mês 2", "Real Mês 2", "Meta Mês 3", "Real Mês 3"]
for col, h in enumerate(headers_ind, 1):
    ws4.cell(row=row, column=col, value=h)
style_header_row(ws4, row, 8)

indicadores = [
    ("Reuniões individuais realizadas (/7)", 0, 7, "", 7, "", 7, ""),
    ("Reuniões de produção conduzidas por Elias", 0, "20/mês", "", "20/mês", "", "20/mês", ""),
    ("Sabatinas semanais realizadas", 0, 4, "", 4, "", 4, ""),
    ("Delegações com ensino por semana", 0, 3, "", 3, "", 3, ""),
    ("Episódios 'Quase Meti a Mão' (menor=melhor)", "—", "Registrar", "", "Reduzir 30%", "", "Reduzir 50%", ""),
    ("Relatórios diários revisados e assinados", 0, "100%", "", "100%", "", "100%", ""),
    ("Planilha de serviços atualizada", "—", "Diário", "", "Diário", "", "Diário + análise", ""),
    ("Diagnósticos feitos por Elias (sem executar)", "—", "Registrar", "", "80%+ dos casos", "", "90%+ dos casos", ""),
    ("Trocas com Caio por mês", 0, 4, "", 4, "", 4, ""),
    ("Potencial sucessor em desenvolvimento", 0, "Identificar", "", "Envolver", "", "Cobrir 1 ausência", ""),
    ("Avaliação técnica da equipe", "—", "—", "", "1ª ficha", "", "Atualizar", ""),
    ("Nota feedback anônimo do time (1-5)", "—", "—", "", "Coletar 1ª", "", "Comparar", ""),
    ("Situações Pause & Reframe por semana", 0, 3, "", 3, "", 3, ""),
    ("Ações do PDI concluídas (acumulado)", "0/20", "7/20", "", "14/20", "", "18/20", ""),
]

for ind in indicadores:
    row += 1
    for col, val in enumerate(ind, 1):
        cell = style_data_cell(ws4, row, col)
        cell.value = val
        if col >= 2:
            cell.alignment = center_align

# Feedback do time - REVISADO
row += 2
ws4.merge_cells(f"A{row}:H{row}")
ws4.cell(row=row, column=1, value="FEEDBACK ANÔNIMO DO TIME (COLETAS BIMESTRAIS)").font = subtitle_font
row += 1

headers_fb = ["Pergunta", "1ª Coleta (Mês 2)\nNota 1-5", "Comentários", "2ª Coleta (Mês 4)\nNota 1-5", "Comentários", "Evolução"]
for col, h in enumerate(headers_fb, 1):
    ws4.cell(row=row, column=col, value=h)
style_header_row(ws4, row, 6)

perguntas_fb = [
    "Como você avalia a comunicação do Elias com você?",
    "Você sente que o Elias ouve sua opinião?",
    "O Elias tem dado espaço para você executar e aprender?",
    "Como você se sente nas interações com o Elias?",
    "O Elias tem ajudado no seu desenvolvimento profissional?",
    "Elias conduz bem as reuniões de produção?",
    "Elias dá direcionamento técnico claro?",
    "Você sabe a quem recorrer para questões técnicas?",
    "Nota geral para a liderança do Elias (1 a 5)",
]

for p in perguntas_fb:
    row += 1
    ws4.cell(row=row, column=1, value=p).font = normal_font
    ws4.cell(row=row, column=1).alignment = wrap_align
    ws4.cell(row=row, column=1).border = thin_border
    for col in range(2, 7):
        ws4.cell(row=row, column=col).border = thin_border
        ws4.cell(row=row, column=col).alignment = center_align

# Resumo executivo
row += 2
ws4.merge_cells(f"A{row}:H{row}")
ws4.cell(row=row, column=1, value="RESUMO EXECUTIVO — RESULTADO DO CICLO").font = subtitle_font
row += 1

resumo_items = [
    ("Principais conquistas:", ""),
    ("Competências com maior evolução:", ""),
    ("Competências que precisam de mais tempo:", ""),
    ("Impacto percebido na equipe:", ""),
    ("Impacto na operação (diagnósticos, planilha, reuniões):", ""),
    ("Percepção de Caio sobre autonomia de Elias:", ""),
    ("Evolução do potencial sucessor:", ""),
    ("Recomendações para o próximo ciclo:", ""),
    ("Nota geral do PDI (consultor):", ""),
]

for label, val in resumo_items:
    ws4.cell(row=row, column=1, value=label).font = bold_font
    ws4.cell(row=row, column=1).border = thin_border
    ws4.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    ws4.cell(row=row, column=2, value=val).border = thin_border
    ws4.cell(row=row, column=2).alignment = wrap_align
    ws4.row_dimensions[row].height = 30
    row += 1

ws4.column_dimensions["A"].width = 48
ws4.column_dimensions["B"].width = 15
ws4.column_dimensions["C"].width = 15
ws4.column_dimensions["D"].width = 15
ws4.column_dimensions["E"].width = 15
ws4.column_dimensions["F"].width = 15
ws4.column_dimensions["G"].width = 15
ws4.column_dimensions["H"].width = 15

# =============================================
# ABA 5: CHECKPOINTS — REVISADO
# =============================================
ws5 = wb.create_sheet("Checkpoints")
ws5.sheet_properties.tabColor = "7B1FA2"

ws5.merge_cells("A1:G1")
ws5["A1"].value = "CHECKPOINTS MENSAIS — REGISTRO DE SESSÕES (v2)"
ws5["A1"].font = title_font
ws5["A1"].alignment = center_align

checkpoints = [
    ("CHECKPOINT MÊS 1 — Fim de Março/2026", "Eleine + Elias + Caio", [
        "Elias leu o material DISC completo?",
        "Quantas reuniões individuais realizou?",
        "Elias está conduzindo as reuniões diárias de produção? Como está a qualidade?",
        "Sabatinas: quantas fez? Como foi a participação do time?",
        "Diário 'Quase Meti a Mão': quantos episódios registrados?",
        "Regra dos 10 segundos: conseguiu praticar?",
        "Elias está revisando e assinando os relatórios diários?",
        "Planilha de serviços: está atualizando? Usa para tomada de decisão?",
        "Trocas com Caio: quantas? Foram úteis?",
        "Mapa de energia: iniciou o registro?",
        "Sessão Líder Multiplicador realizada?",
        "Foco no diagnóstico: Elias está direcionando sem executar?",
        "Nível de energia e motivação de Elias (auto-avaliação 1-5):",
        "Percepção de Caio: Elias está assumindo o papel? O que falta?",
        "Ajustes necessários no PDI:",
    ]),
    ("CHECKPOINT MÊS 2 — Fim de Abril/2026", "Eleine + Elias + Caio", [
        "Mapa de perfis do time está completo?",
        "Elias está adaptando a comunicação por perfil?",
        "Condução de reuniões: evolução da qualidade e do engajamento do time?",
        "Sabatinas: evolução da participação do time?",
        "Delegação com ensino: 3/semana sendo feitas?",
        "Diário: houve redução dos episódios de 'meter a mão'?",
        "Avaliação técnica e comportamental: Elias preencheu a ficha dos 7?",
        "Potencial sucessor identificado? Quem? Já envolveu em diagnósticos?",
        "1ª coleta de feedback anônimo do time realizada?",
        "Resultado da 1ª coleta (resumo):",
        "Planilha: Elias está fazendo análise ou apenas preenchendo?",
        "Percepção de Caio: Elias está aliviando a carga dele? Em que medida?",
        "Leitura Goleman: iniciou?",
        "Ajustes necessários no PDI:",
    ]),
    ("CHECKPOINT MÊS 3 — Fim de Maio/2026", "Eleine + Elias + Caio", [
        "Evolução geral nas 4 competências (nota 1-5 cada):",
        "Mapa de energia: padrões identificados? Ajustes feitos?",
        "Sabatinas: o time está mais engajado? Traz soluções sozinho?",
        "Delegação: Elias conseguiu soltar tarefas operacionais?",
        "Gestão técnica: relatórios, planilha, diagnósticos — está consolidado?",
        "Sucessor: pessoa está evoluindo? Cobriu alguma ausência de Elias?",
        "Pause & Reframe: está usando? Exemplos?",
        "Trocas com Caio: estão acontecendo semanalmente?",
        "Leitura concluída? Insights aplicados?",
        "Percepção geral da equipe (informal):",
        "Caio sente que está com mais liberdade para focar em comercial/gestão?",
        "O que funcionou melhor até agora?",
        "O que precisa mudar para os próximos 3 meses?",
    ]),
]

row = 3
for title, participantes, perguntas in checkpoints:
    ws5.merge_cells(f"A{row}:G{row}")
    ws5.cell(row=row, column=1, value=title).font = subtitle_font
    ws5.cell(row=row, column=1).fill = PatternFill(start_color=LIGHT_TEAL, end_color=LIGHT_TEAL, fill_type="solid")
    row += 1

    ws5.cell(row=row, column=1, value="Participantes:").font = bold_font
    ws5.cell(row=row, column=2, value=participantes).font = normal_font
    ws5.cell(row=row, column=1).border = thin_border
    ws5.cell(row=row, column=2).border = thin_border
    row += 1

    ws5.cell(row=row, column=1, value="Data realizada:").font = bold_font
    ws5.cell(row=row, column=2, value="___/___/______").font = normal_font
    ws5.cell(row=row, column=1).border = thin_border
    ws5.cell(row=row, column=2).border = thin_border
    row += 1

    headers_ck = ["#", "Pergunta / Item", "Resposta / Evidência"]
    for col, h in enumerate(headers_ck, 1):
        ws5.cell(row=row, column=col, value=h)
    style_header_row(ws5, row, 3)
    row += 1

    for i, p in enumerate(perguntas, 1):
        ws5.cell(row=row, column=1, value=i).font = normal_font
        ws5.cell(row=row, column=1).alignment = center_align
        ws5.cell(row=row, column=1).border = thin_border
        ws5.cell(row=row, column=2, value=p).font = normal_font
        ws5.cell(row=row, column=2).alignment = wrap_align
        ws5.cell(row=row, column=2).border = thin_border
        ws5.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        ws5.cell(row=row, column=3).border = thin_border
        ws5.cell(row=row, column=3).alignment = wrap_align
        ws5.row_dimensions[row].height = 30
        row += 1

    row += 1

ws5.column_dimensions["A"].width = 8
ws5.column_dimensions["B"].width = 55
ws5.column_dimensions["C"].width = 25
ws5.column_dimensions["D"].width = 15
ws5.column_dimensions["E"].width = 15
ws5.column_dimensions["F"].width = 15
ws5.column_dimensions["G"].width = 15

# =============================================
# ABA 6: SEMÁFORO VISUAL — REVISADO
# =============================================
ws6 = wb.create_sheet("Semáforo")
ws6.sheet_properties.tabColor = YELLOW

ws6.merge_cells("A1:F1")
ws6["A1"].value = "SEMÁFORO DE PROGRESSO — VISÃO EXECUTIVA (v2)"
ws6["A1"].font = title_font
ws6["A1"].alignment = center_align

ws6.merge_cells("A2:F2")
ws6["A2"].value = "Verde = No caminho | Amarelo = Atenção | Vermelho = Intervenção necessária"
ws6["A2"].font = Font(name="Calibri", italic=True, size=10, color="757575")
ws6["A2"].alignment = center_align

row = 4
headers_sem = ["Dimensão", "Indicador-Chave", "Meta", "Mês 1", "Mês 2", "Mês 3"]
for col, h in enumerate(headers_sem, 1):
    ws6.cell(row=row, column=col, value=h)
style_header_row(ws6, row, 6)

semaforo_data = [
    ("Escuta Ativa", "Reuniões individuais realizadas", "7/mês", "", "", ""),
    ("Escuta Ativa", "Condução de reuniões de produção", "Diário", "", "", ""),
    ("Escuta Ativa", "Time se sente ouvido (feedback)", "≥ 3,5/5", "", "", ""),
    ("Transição Líder", "Sabatinas realizadas", "4/mês", "", "", ""),
    ("Transição Líder", "Tarefas operacionais delegadas", "≥ 2 novas/mês", "", "", ""),
    ("Transição Líder", "Diagnósticos feitos sem executar", "80%+ dos casos", "", "", ""),
    ("Transição Líder", "Episódios 'mão na massa' (↓)", "Reduzir 50%", "", "", ""),
    ("Gestão Técnica", "Relatórios diários revisados/assinados", "100%", "", "", ""),
    ("Gestão Técnica", "Planilha de serviços atualizada", "Diário", "", "", ""),
    ("Gestão Técnica", "Avaliação técnica da equipe preenchida", "Sim/Não", "", "", ""),
    ("Gestão Técnica", "Sucessor em desenvolvimento", "1 pessoa ativa", "", "", ""),
    ("Intelig. Emocional", "Situações Pause & Reframe", "3/semana", "", "", ""),
    ("Intelig. Emocional", "Nota do time sobre clima", "≥ 3,5/5", "", "", ""),
    ("Geral", "Ações do PDI concluídas", "90%+", "", "", ""),
    ("Geral", "Trocas com Caio realizadas", "4/mês", "", "", ""),
    ("Geral", "Energia/motivação Elias (1-5)", "≥ 4", "", "", ""),
    ("Geral", "Caio sente alívio operacional?", "Sim", "", "", ""),
]

for s in semaforo_data:
    row += 1
    for col, val in enumerate(s, 1):
        cell = style_data_cell(ws6, row, col)
        cell.value = val
        if col >= 4:
            cell.alignment = center_align

ws6.column_dimensions["A"].width = 22
ws6.column_dimensions["B"].width = 40
ws6.column_dimensions["C"].width = 18
ws6.column_dimensions["D"].width = 15
ws6.column_dimensions["E"].width = 15
ws6.column_dimensions["F"].width = 15

# Legenda
row += 2
ws6.cell(row=row, column=1, value="LEGENDA:").font = bold_font
row += 1
legends = [
    ("VERDE", "70%+ das ações concluídas, indicadores no caminho, time reporta evolução, Caio percebe alívio operacional", GREEN, WHITE),
    ("AMARELO", "40-69% das ações, alguns indicadores atrasados, 1 checkpoint com pendências, Elias ainda executa mais que diagnostica", YELLOW, DARK_GRAY),
    ("VERMELHO", "<40% das ações, indicadores estagnados, time não percebe mudança, energia caindo, sem avanço na formação de sucessor", RED, WHITE),
]
for label, desc, color, font_color in legends:
    cell1 = ws6.cell(row=row, column=1, value=label)
    cell1.font = Font(name="Calibri", bold=True, color=font_color, size=10)
    cell1.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell1.alignment = center_align
    cell1.border = thin_border
    ws6.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws6.cell(row=row, column=2, value=desc).font = normal_font
    ws6.cell(row=row, column=2).border = thin_border
    row += 1

# =============================================
# ABA 7: COMPETÊNCIAS DE REFERÊNCIA (NOVA)
# =============================================
ws7 = wb.create_sheet("Competências Referência")
ws7.sheet_properties.tabColor = "5C6BC0"

ws7.merge_cells("A1:E1")
ws7["A1"].value = "MAPEAMENTO DE COMPETÊNCIAS — BASE PARA AVALIAÇÃO"
ws7["A1"].font = title_font
ws7["A1"].alignment = center_align

ws7.merge_cells("A2:E2")
ws7["A2"].value = "Fonte: Planilha 190 Exemplos de Competências para Avaliação de Desempenho + Contexto Polidiesel"
ws7["A2"].font = Font(name="Calibri", italic=True, size=10, color="757575")
ws7["A2"].alignment = center_align

row = 4
headers_ref = ["Categoria", "Competência", "Relevância para Elias", "Nível Atual Estimado (1-5)", "Prioridade no PDI"]
for col, h in enumerate(headers_ref, 1):
    ws7.cell(row=row, column=col, value=h)
style_header_row(ws7, row, 5)

ref_data = [
    ("Gestão de Pessoas", "Coaching e Mentoria", "CRÍTICA — precisa formar sucessor e desenvolver o time", 1, "Competência 3"),
    ("Gestão de Pessoas", "Gerenciamento de Desempenho", "ALTA — avaliar mecânicos tecnicamente e por postura", 1, "Competência 3"),
    ("Gestão de Pessoas", "Trabalho em Equipe", "ALTA — transitar de executor para facilitador do time", 2, "Competência 2"),
    ("Gestão de Pessoas", "Treinamento e Desenvolvimento", "MÉDIA — compartilhar conhecimento técnico com a equipe", 2, "Competência 3"),
    ("Gestão de Pessoas", "Adaptabilidade e Flexibilidade", "ALTA — IF=-4,76, transição de papel exige flexibilidade", 2, "Competências 2 e 4"),
    ("Comunicação", "Habilidades de Escuta", "CRÍTICA — DISC mostra 'Capacidade de Ouvir' muito baixa", 2, "Competência 1"),
    ("Comunicação", "Habilidades de Fala", "MÉDIA — Elias é 'de poucas palavras' (Caio no kick-off)", 2, "Competência 1"),
    ("Comunicação", "Técnicas de Negociação", "MÉDIA — intermediar recepção ↔ área técnica com clientes", 2, "Competência 1"),
    ("Liderança", "Gestão Estratégica", "ALTA — visão do quadro geral, não apenas execução", 1, "Competência 2"),
    ("Liderança", "Responsabilidade e Confiabilidade", "ALTA — já possui (Caio nota 10/10), manter e formalizar", 4, "Ativo existente"),
    ("Liderança", "Persuasão e Influência", "MÉDIA — time ainda o vê como par, precisa construir autoridade", 2, "Competência 2"),
    ("Desenvolvimento Pessoal", "Potencial de Crescimento", "ALTA — Caio vê 9/10 de potencial", 3, "Ativo existente"),
    ("Desenvolvimento Pessoal", "Mudança de Liderança", "CRÍTICA — lidar com impactos da transição no time", 1, "Competência 2"),
    ("Desenvolvimento Pessoal", "Foco nos Resultados", "ALTA — já possui (perfil Executor), direcionar para gestão", 4, "Ativo existente"),
    ("Raciocínio Lógico", "Tomada de Decisões", "ALTA — papel de diagnóstico e direcionamento técnico", 3, "Competência 3"),
    ("Raciocínio Lógico", "Resolução de Problemas", "ALTA — ponto focal técnico, resolver sem executar", 3, "Competência 2"),
    ("Raciocínio Lógico", "Abordagem Metódica", "MÉDIA — uso de planilha, relatórios, gestão organizada", 2, "Competência 3"),
    ("Comportamental", "Empatia", "CRÍTICA — DISC mostra empatia muito baixa", 1, "Competência 4"),
    ("Comportamental", "Habilidades Interpessoais", "ALTA — construir relações de liderança, não só de execução", 2, "Competências 1 e 4"),
    ("Comportamental", "Gestão do Tempo", "MÉDIA — equilibrar novo papel administrativo com operacional residual", 2, "Competência 2"),
    ("Competências Técnicas", "Conhecimento de Informática", "MÉDIA — planilha, possível uso de tablet no futuro", 2, "Competência 3"),
    ("Transferível", "Planejamento e Organização", "ALTA — gestão diária, alocação de recursos, priorização", 2, "Competência 3"),
]

for d in ref_data:
    row += 1
    for col, val in enumerate(d, 1):
        cell = style_data_cell(ws7, row, col)
        cell.value = val
        if col == 4:
            cell.alignment = center_align
            if isinstance(val, int):
                if val <= 1:
                    cell.fill = red_fill
                elif val == 2:
                    cell.fill = yellow_fill
                elif val == 3:
                    cell.fill = orange_fill
                elif val >= 4:
                    cell.fill = green_fill
        if col == 3:
            if "CRÍTICA" in str(val):
                cell.font = Font(name="Calibri", bold=True, color=RED, size=10)
            elif "ALTA" in str(val):
                cell.font = Font(name="Calibri", bold=True, color=ORANGE, size=10)
        if col == 5:
            cell.alignment = center_align

ws7.column_dimensions["A"].width = 22
ws7.column_dimensions["B"].width = 32
ws7.column_dimensions["C"].width = 55
ws7.column_dimensions["D"].width = 22
ws7.column_dimensions["E"].width = 20

# =============================================
# SALVAR
# =============================================
output_path = "/home/headless/workspace/cih/operacional/execução/polidiesel/elias-fernandes/PDI_Elias_Fernandes_Polidiesel.xlsx"
wb.save(output_path)
print(f"Arquivo salvo em: {output_path}")
print(f"Abas criadas: {wb.sheetnames}")
print(f"Total de ações: 20")
print(f"Total de competências: 4")
print(f"Total de indicadores: 14")
print(f"Total de perguntas feedback: 9")
print(f"Total de itens semáforo: 17")
print(f"Total de competências mapeadas (referência): 22")
