#!/usr/bin/env python3
"""Adiciona aba 'Guia 1-1' ao PDI da Ayla Lima com instruções e exemplos personalizados."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Cores CIH ──
TEAL = '0097A7'
DARK_TEAL = '00796B'
WHITE = 'FFFFFF'
LIGHT_TEAL = 'E0F2F1'
LIGHT_GRAY = 'F5F5F5'
LIGHT_GREEN = 'E8F5E9'
LIGHT_YELLOW = 'FFFDE7'
LIGHT_BLUE = 'E3F2FD'
LIGHT_ORANGE = 'FFF3E0'
LIGHT_PURPLE = 'F3E5F5'
GRAY_TEXT = '757575'
DARK_GRAY = '424242'

# ── Estilos ──
def hfill(c):
    return PatternFill(start_color=c, end_color=c, fill_type='solid')

header_fill = hfill(TEAL)
header_font = Font(bold=True, size=11, color=WHITE)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
title_font = Font(bold=True, size=14, color=TEAL)
title_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
subtitle_font = Font(bold=True, size=12, color=DARK_TEAL)
subtitle_align = Alignment(horizontal='left', vertical='center')
section_fill = hfill(LIGHT_TEAL)
section_font = Font(bold=True, size=11, color=DARK_TEAL)
label_font = Font(bold=True, size=10)
data_font = Font(size=10)
data_align_l = Alignment(horizontal='left', vertical='top', wrap_text=True)
data_align_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
tip_font = Font(size=10, italic=True, color=GRAY_TEXT)
bold_font = Font(bold=True, size=10)
example_font = Font(size=10, color='37474F')

thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

def cell(ws, row, col, val, font=None, fill=None, align=None, merge_end=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font or data_font
    if fill:
        c.fill = fill
    c.alignment = align or data_align_l
    c.border = thin_border
    if merge_end:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end)
    return c

def header_row(ws, row, cols):
    for ci, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = thin_border

def section_title(ws, row, text, end_col=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = section_fill
    c.font = section_font
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = thin_border
    ws.row_dimensions[row].height = 28
    return row + 1

# ═══════════════════════════════════════════════
# Abrir PDI existente
# ═══════════════════════════════════════════════
path = '/home/headless/workspace/cih/operacional/execução/valeassessoria/aylalima/PDI_Ayla_Lima_ValeAssessoria.xlsx'
wb = openpyxl.load_workbook(path)

ws = wb.create_sheet('Guia 1-1')

# Colunas
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 50
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 22
ws.column_dimensions['G'].width = 22
ws.column_dimensions['H'].width = 22

# ═══════════════════════════════════════════════
# TÍTULO
# ═══════════════════════════════════════════════
ws.merge_cells('A1:H1')
c = ws.cell(row=1, column=1, value='GUIA DE REUNIÕES 1:1 — LARISSA ↔ AYLA')
c.font = title_font; c.alignment = title_align

ws.merge_cells('A2:H2')
c = ws.cell(row=2, column=1, value='Instruções, modelos e exemplos práticos para acompanhamento do PDI via reuniões one-on-one')
c.font = Font(size=10, color=GRAY_TEXT); c.alignment = data_align_c
ws.row_dimensions[1].height = 32
ws.row_dimensions[2].height = 22

# ═══════════════════════════════════════════════
# SEÇÃO 1 — O QUE É E POR QUE FAZER
# ═══════════════════════════════════════════════
r = 4
r = section_title(ws, r, '1. O QUE É A REUNIÃO 1:1 E POR QUE FAZER')

items_oq = [
    ('Definição',
     'É um encontro individual e recorrente entre líder e colaboradora, com foco no desenvolvimento, alinhamento e acompanhamento do PDI. Não é cobrança — é conexão.'),
    ('Quem participa',
     'Larissa (líder) + Ayla (colaboradora). Eventualmente, Eleine pode participar nos checkpoints mensais.'),
    ('Frequência recomendada',
     'Quinzenal (a cada 15 dias), com duração de 20-30 minutos. Pode ser semanal no primeiro mês para criar o hábito.'),
    ('Onde realizar',
     'Preferencialmente presencial, em local reservado e sem interrupções. Se remoto: câmera ligada, sem multitarefa.'),
    ('Regra de ouro',
     'A reunião é 90% da Ayla e 10% da Larissa. Larissa escuta, faz perguntas e orienta. Ayla traz a pauta, os desafios e as conquistas.'),
    ('Benefícios comprovados',
     'Reduz rotatividade em 15%, aumenta engajamento em 30%, acelera o desenvolvimento, previne problemas e fortalece a confiança mútua.'),
]

for lab, desc in items_oq:
    cell(ws, r, 1, '', fill=hfill(LIGHT_TEAL))
    cell(ws, r, 2, lab, font=bold_font, fill=hfill(LIGHT_TEAL))
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    cell(ws, r, 3, desc, align=data_align_l)
    ws.row_dimensions[r].height = 40
    r += 1

# ═══════════════════════════════════════════════
# SEÇÃO 2 — BOAS PRÁTICAS
# ═══════════════════════════════════════════════
r += 1
r = section_title(ws, r, '2. BOAS PRÁTICAS PARA A 1:1')

header_row(ws, r, ['', 'Fazer', 'Evitar', '', '', '', '', ''])
r += 1

boas_praticas = [
    ('Agendar com antecedência e nunca cancelar sem reagendar',
     'Cancelar em cima da hora ou "deixar pra depois"'),
    ('Começar com algo positivo — reconhecer uma conquista',
     'Começar listando erros ou problemas'),
    ('Fazer perguntas abertas ("Como você se sentiu sobre...")',
     'Fazer interrogatório ou perguntas fechadas ("Você fez X?")'),
    ('Escutar ativamente — 80% escuta, 20% fala (para Larissa)',
     'Monopolizar a conversa ou dar sermão'),
    ('Anotar combinados e compromissos ao final',
     'Sair da reunião sem ações definidas'),
    ('Conectar a conversa com as competências do PDI',
     'Falar só de tarefas operacionais sem foco em desenvolvimento'),
    ('Criar ambiente seguro — é ok errar e dizer que está difícil',
     'Usar a 1:1 como reunião de cobrança ou punição'),
    ('Celebrar pequenas vitórias e progressos',
     'Só falar de problemas e gaps'),
]

for fazer, evitar in boas_praticas:
    cell(ws, r, 1, '✓', font=Font(bold=True, size=10, color='43A047'), align=data_align_c)
    cell(ws, r, 2, fazer, align=data_align_l, merge_end=4)
    cell(ws, r, 5, '✗', font=Font(bold=True, size=10, color='E53935'), align=data_align_c)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    cell(ws, r, 6, evitar, align=data_align_l)
    ws.row_dimensions[r].height = 35
    r += 1

# ═══════════════════════════════════════════════
# SEÇÃO 3 — MODELO DE PAUTA RECOMENDADO
# ═══════════════════════════════════════════════
r += 1
r = section_title(ws, r, '3. MODELO DE PAUTA — ADAPTADO PARA O PDI DA AYLA')

ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cell(ws, r, 1, 'Modelo "8 Áreas Principais" adaptado — organiza a conversa em blocos temáticos conectados às 5 competências do PDI.',
     font=tip_font, align=data_align_l)
ws.row_dimensions[r].height = 22
r += 1

header_row(ws, r, ['#', 'Bloco da Pauta', 'O que abordar', 'Exemplo de Pergunta para Ayla', 'Tempo', 'Competência PDI', '', ''])
r += 1

blocos_pauta = [
    (1, 'Abertura e Check-in',
     'Como Ayla está se sentindo? Motivação, energia, dificuldades pessoais que impactam o trabalho.',
     '"Ayla, como você está se sentindo essa quinzena? O que te animou e o que te frustrou?"',
     '3 min', 'Todas'),
    (2, 'Conquistas e Progressos',
     'O que deu certo desde a última 1:1? Celebrar vitórias, mesmo pequenas.',
     '"Me conta uma coisa que você fez diferente essa quinzena e que deu certo."',
     '4 min', 'Todas'),
    (3, 'Organização e Rotina',
     'Uso do planner, checklist por turno, verificação do Trello, rotina de encerramento do dia.',
     '"Você está conseguindo usar o planner todos os dias? Quais dias foram mais difíceis e por quê?"',
     '4 min', 'Competência 1'),
    (4, 'Atenção e Erros',
     'Protocolo PARE-CONFIRA-ENVIE, diário de erros, revisão dupla de postagens.',
     '"Quantos erros você registrou no diário essa quinzena? Notou algum padrão?"',
     '4 min', 'Competência 2'),
    (5, 'Comunicação com Clientes',
     'Uso dos templates, formalidade, adaptação por perfil de cliente.',
     '"Teve alguma situação com cliente que você sentiu dificuldade em ser mais formal? Como lidou?"',
     '3 min', 'Competência 3'),
    (6, 'Análise e Gestão de Clientes',
     'Análise semanal do Trello, Top 3 Clientes, playbook de soluções.',
     '"Quais são os 3 clientes mais críticos agora? O que você acha que pode resolver?"',
     '4 min', 'Competência 4'),
    (7, 'Autonomia e Cobranças',
     'Cobrança à equipe, pedidos de ajuda, iniciativa própria.',
     '"Nessa quinzena, teve algum momento que você resolveu algo sozinha sem precisar de passo a passo?"',
     '3 min', 'Competência 5'),
    (8, 'Fechamento e Compromissos',
     'Resumir os pontos principais, definir 2-3 ações específicas até a próxima 1:1.',
     '"Então para as próximas 2 semanas, combinamos: 1) ___, 2) ___, 3) ___. Concordamos?"',
     '3 min', 'Todas'),
]

for num, bloco, abordagem, pergunta, tempo, comp in blocos_pauta:
    cell(ws, r, 1, num, font=Font(bold=True, size=11, color=TEAL), align=data_align_c)
    cell(ws, r, 2, bloco, font=bold_font, align=data_align_l)
    cell(ws, r, 3, abordagem, align=data_align_l)
    cell(ws, r, 4, pergunta, font=example_font, align=data_align_l)
    cell(ws, r, 5, tempo, align=data_align_c)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    cell(ws, r, 6, comp, align=data_align_c)
    ws.row_dimensions[r].height = 55
    r += 1

ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cell(ws, r, 1, 'Tempo total estimado: 25-30 minutos | A pauta pode ser simplificada — o importante é manter a regularidade.',
     font=tip_font, align=data_align_l)
r += 1

# ═══════════════════════════════════════════════
# SEÇÃO 4 — PERGUNTAS POR COMPETÊNCIA
# ═══════════════════════════════════════════════
r += 1
r = section_title(ws, r, '4. BANCO DE PERGUNTAS POR COMPETÊNCIA DO PDI')

ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cell(ws, r, 1, 'Larissa pode escolher 2-3 perguntas por reunião. Não precisa usar todas de uma vez — variar mantém a conversa fresca.',
     font=tip_font, align=data_align_l)
r += 1

perguntas_por_comp = [
    ('Organização e Gestão de Rotina', hfill(LIGHT_BLUE), [
        'Você está conseguindo dividir suas atividades por turno (manhã/tarde)?',
        'Como está o uso do planner? Algum dia da semana é mais difícil de manter?',
        'Está verificando o Trello 2x por turno? O que encontra quando verifica?',
        'Está enviando a lista de pendências antes de sair? Isso ajudou a organizar?',
        'Qual o maior desafio de organização que você enfrentou essa quinzena?',
    ]),
    ('Atenção aos Detalhes e Saída do Automático', hfill(LIGHT_YELLOW), [
        'Quantos erros você registrou no diário essa quinzena? Teve algum padrão?',
        'O protocolo PARE-CONFIRA-ENVIE já está ficando automático ou ainda precisa lembrar?',
        'Conte um momento que você "parou" antes de agir e isso evitou um erro.',
        'Qual erro mais te incomodou e o que você fez para que não aconteça de novo?',
        'Você percebe quando está no "modo automático"? O que faz para sair?',
    ]),
    ('Comunicação Profissional e Formalidade', hfill(LIGHT_GREEN), [
        'Usou algum template de mensagem essa quinzena? Funcionou bem?',
        'Teve alguma situação onde sentiu que a comunicação poderia ter sido mais formal?',
        'Algum cliente reagiu diferente ao seu novo tom de comunicação?',
        'Você está conseguindo adaptar a linguagem conforme o perfil do cliente?',
        'Larissa precisou ajustar alguma comunicação sua essa quinzena?',
    ]),
    ('Pensamento Analítico e Gestão de Clientes', hfill(LIGHT_ORANGE), [
        'Fez a análise semanal de clientes no Trello? O que descobriu?',
        'Quais são os Top 3 Clientes mais críticos agora e por quê?',
        'Registrou alguma solução no seu playbook pessoal? Qual?',
        'Conseguiu antecipar algum problema antes que ele virasse urgência?',
        'Se eu (Larissa) perguntasse agora: "como está o cliente X?" — você saberia responder com dados?',
    ]),
    ('Autonomia e Senso de Dono', hfill(LIGHT_PURPLE), [
        'Essa quinzena, resolveu alguma coisa sozinha que antes precisaria de ajuda?',
        'Cobrou o design/equipe de forma assertiva? Como foi a reação?',
        'Teve algum momento que travou em uma tarefa? Quanto tempo levou para pedir ajuda?',
        'O que você faria diferente se pudesse refazer essa quinzena?',
        'Se eu não estivesse disponível por 1 dia inteiro, como seria o dia da operação?',
    ]),
]

for comp_name, comp_fill, perguntas in perguntas_por_comp:
    # Sub-cabeçalho da competência
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    cell(ws, r, 1, comp_name, font=Font(bold=True, size=10, color=DARK_TEAL), fill=comp_fill,
         align=Alignment(horizontal='left', vertical='center'))
    ws.row_dimensions[r].height = 24
    r += 1
    for p in perguntas:
        cell(ws, r, 1, '•', align=data_align_c)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        cell(ws, r, 2, p, align=data_align_l)
        ws.row_dimensions[r].height = 22
        r += 1
    r += 1  # espaço entre competências

# ═══════════════════════════════════════════════
# SEÇÃO 5 — EXEMPLOS DE FEEDBACK CONTEXTUALIZADO
# ═══════════════════════════════════════════════
r = section_title(ws, r, '5. EXEMPLOS DE FEEDBACK — BASEADOS NO CONTEXTO DA AYLA')

ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cell(ws, r, 1, 'Modelos de feedback construtivo e positivo que Larissa pode adaptar durante as 1:1. Sempre ser específica, não genérica.',
     font=tip_font, align=data_align_l)
r += 1

header_row(ws, r, ['', 'Situação', 'Tipo', 'Exemplo de Feedback', '', '', '', ''])
r += 1

feedbacks = [
    ('Ayla usou o planner a semana toda sem falhar',
     'Positivo',
     '"Ayla, percebi que você usou o planner toda a semana e nenhuma tarefa ficou esquecida. Isso mostra que o hábito está se formando. Continue assim — a organização está fazendo diferença nas entregas."'),
    ('Ayla postou conteúdo no grupo errado de um cliente',
     'Construtivo',
     '"Ayla, vi que o post do cliente X foi para o grupo errado. Entendo que foi correria, mas lembra do protocolo PARE-CONFIRA-ENVIE? O que você acha que faltou nessa hora? Como podemos evitar que aconteça de novo?"'),
    ('Ayla cobrou o design proativamente e evitou atraso',
     'Positivo',
     '"Ayla, gostei muito que você cobrou o design antes do prazo apertar. Isso é senso de dono na prática. O cliente nem percebeu que poderia ter atrasado. Esse é o tipo de atitude que faz diferença."'),
    ('Comunicação com cliente foi muito informal',
     'Construtivo',
     '"Ayla, a mensagem para o cliente Y ficou um pouco informal demais — parecia conversa entre amigas. Para esse tipo de comunicação, use o template de follow-up que criamos. Quer que a gente revise juntas como adaptar?"'),
    ('Ayla trouxe análise Top 3 Clientes com sugestões',
     'Positivo',
     '"Ayla, sua análise dos Top 3 Clientes ficou excelente. Você identificou o gargalo do cliente Z antes de eu precisar perguntar. Isso é pensamento analítico — exatamente o que o seu PDI está desenvolvendo."'),
    ('Ayla ficou travada 40 min sem pedir ajuda',
     'Construtivo',
     '"Ayla, vi que você ficou quase 40 minutos tentando resolver sozinha o problema do Trello. Lembra que combinamos que após 15 minutos sem avançar, é hora de pedir ajuda? Não é fraqueza — é eficiência. O que te impediu de chamar?"'),
    ('Ayla não enviou pendências ao final do dia',
     'Construtivo',
     '"Ayla, ontem não recebi a lista de pendências no final do dia. Isso me deixa sem visibilidade do que ficou em aberto. Conseguimos pensar num lembrete — alarme no celular ou item fixo no checklist — para garantir que isso não escape?"'),
    ('Ayla adaptou a linguagem para cliente executor',
     'Positivo',
     '"Ayla, percebi que com o cliente W você foi mais direta e objetiva na mensagem — sem rodeios, direto ao ponto. Isso é adaptar comunicação ao perfil do cliente, exatamente o que discutimos no DISC. Muito bom!"'),
    ('Ayla identificou sozinha um problema recorrente de um cliente',
     'Positivo',
     '"Ayla, quando você me trouxe que o cliente V tem sempre o mesmo gargalo na aprovação e sugeriu mudar o dia de envio, isso mostrou que você está pensando de forma analítica. Está construindo seu playbook na prática."'),
    ('Erros diminuíram mas ainda acontecem por pressa',
     'Construtivo + Incentivo',
     '"Ayla, seus erros reduziram bastante — de 5 por semana para 2. Isso é evolução real. Mas os que ainda acontecem parecem ser por pressa. O que acha de nos próximos 15 dias focar especificamente no momento de maior pressa do dia e aplicar o protocolo ali?"'),
]

for sit, tipo, fb in feedbacks:
    tipo_fill = hfill(LIGHT_GREEN) if 'Positivo' in tipo else hfill(LIGHT_YELLOW)
    cell(ws, r, 1, '', fill=tipo_fill)
    cell(ws, r, 2, sit, font=bold_font, align=data_align_l)
    cell(ws, r, 3, tipo, fill=tipo_fill, font=Font(bold=True, size=10), align=data_align_c)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
    cell(ws, r, 4, fb, font=example_font, align=data_align_l)
    ws.row_dimensions[r].height = 65
    r += 1

# ═══════════════════════════════════════════════
# SEÇÃO 6 — REGISTRO DAS REUNIÕES
# ═══════════════════════════════════════════════
r += 1
r = section_title(ws, r, '6. REGISTRO DAS REUNIÕES 1:1 (PREENCHER A CADA ENCONTRO)')

header_row(ws, r, ['#', 'Data', 'Competência em foco', 'Principais pontos discutidos',
                    'Compromissos assumidos', 'Prazo', 'Status', 'Observações'])
r += 1

for i in range(1, 9):
    cell(ws, r, 1, i, font=Font(bold=True, size=10, color=TEAL), align=data_align_c)
    cell(ws, r, 2, '___/___/___', align=data_align_c)
    for col in range(3, 9):
        cell(ws, r, col, '', align=data_align_l)
    ws.row_dimensions[r].height = 50
    r += 1

# ═══════════════════════════════════════════════
# SEÇÃO 7 — CALENDÁRIO SUGERIDO
# ═══════════════════════════════════════════════
r += 1
r = section_title(ws, r, '7. CALENDÁRIO SUGERIDO DE 1:1 — MARÇO A JUNHO/2026')

header_row(ws, r, ['#', 'Data sugerida', 'Foco principal', 'Competência PDI',
                    'Perguntas sugeridas', 'Tipo de feedback', '', ''])
r += 1

calendario = [
    (1, '2ª semana Mar', 'Onboarding da 1:1 + como Ayla está se sentindo com o PDI',
     'Todas', 'Check-in geral, expectativas, medos e motivações', 'Incentivo e acolhimento'),
    (2, '4ª semana Mar', 'Organização e planner + primeiros erros registrados',
     'Comp. 1 e 2', 'Uso do planner, checklist, diário de erros', 'Construtivo + positivo'),
    (3, '2ª semana Abr', 'Comunicação + primeiros templates + Top 3 Clientes #1',
     'Comp. 3 e 4', 'Templates em uso, formalidade, análise de clientes', 'Positivo sobre evolução'),
    (4, '4ª semana Abr', 'CHECKPOINT MÊS 2 — revisão geral + autonomia',
     'Todas + Comp. 5', 'Evolução geral, cobranças, iniciativa', 'Balanço construtivo'),
    (5, '2ª semana Mai', 'Consolidação de hábitos + pensamento analítico',
     'Comp. 1, 2 e 4', 'Hábitos automáticos? Antecipa problemas?', 'Celebrar hábitos formados'),
    (6, '4ª semana Mai', 'CHECKPOINT MÊS 3 — prontidão para promoção',
     'Todas', 'Visão de Larissa, clientes, maturidade', 'Feedback formal de progresso'),
    (7, '2ª semana Jun', 'Autonomia total + consolidação',
     'Comp. 5', 'Funciona sozinha por 1 dia? Cobra equipe?', 'Positivo ou plano de ação'),
    (8, '4ª semana Jun', 'FECHAMENTO — avaliação final + parecer de promoção',
     'Todas', 'Nota geral, evolução, parecer final', 'Relatório final'),
]

for num, data, foco, comp, pergs, fb_tipo in calendario:
    is_cp = 'CHECKPOINT' in foco or 'FECHAMENTO' in foco
    row_fill = hfill(LIGHT_ORANGE) if is_cp else None
    cell(ws, r, 1, num, font=Font(bold=True, size=10, color=TEAL), fill=row_fill, align=data_align_c)
    cell(ws, r, 2, data, font=bold_font, fill=row_fill, align=data_align_c)
    cell(ws, r, 3, foco, fill=row_fill, align=data_align_l)
    cell(ws, r, 4, comp, fill=row_fill, align=data_align_c)
    cell(ws, r, 5, pergs, fill=row_fill, align=data_align_l)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    cell(ws, r, 6, fb_tipo, fill=row_fill, align=data_align_l)
    ws.row_dimensions[r].height = 45
    r += 1

# ═══════════════════════════════════════════════
# SEÇÃO 8 — DICAS RÁPIDAS
# ═══════════════════════════════════════════════
r += 1
r = section_title(ws, r, '8. DICAS RÁPIDAS PARA LARISSA')

dicas = [
    'Antes da reunião: revise as ações combinadas na 1:1 anterior e anote 2-3 pontos que observou na quinzena.',
    'No início: comece com algo positivo — reconheça um esforço, mesmo que pequeno. Isso ativa a abertura emocional.',
    'Durante: faça mais perguntas do que afirmações. "O que você acha?" é mais poderoso que "Você deveria...".',
    'Sobre erros: não julgue — explore. Pergunte "o que aconteceu?" antes de "por que você fez isso?".',
    'Sobre o DISC: lembre que Ayla é Comunicadora Executora. Ela responde bem a desafios e reconhecimento, mas não a críticas secas.',
    'Autoestima BAIXA: cuidado com feedback apenas negativo. Sempre equilibre: para cada ajuste, reconheça uma conquista.',
    'Registre tudo: os registros desta aba alimentam os checkpoints e o relatório final de evolução.',
    'Consistência > perfeição: é melhor uma 1:1 de 15 minutos do que nenhuma. Nunca pule sem reagendar.',
]

for dica in dicas:
    cell(ws, r, 1, '→', font=Font(bold=True, size=11, color=TEAL), align=data_align_c)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    cell(ws, r, 2, dica, align=data_align_l)
    ws.row_dimensions[r].height = 28
    r += 1

# Rodapé
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cell(ws, r, 1, 'Guia elaborado pela CIH Consultoria em Inteligência Humana | Consultora: Eleine Passos | Fontes: Convenia, Feedz, Qulture.Rocks',
     font=Font(size=9, italic=True, color=GRAY_TEXT), align=data_align_c)

# ═══════════════════════════════════════════════
# SALVAR
# ═══════════════════════════════════════════════
wb.save(path)
print(f'Aba "Guia 1-1" adicionada com sucesso em: {path}')
