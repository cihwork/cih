#!/usr/bin/env python3
"""
Gera planilha Excel organizada para a equipe de atendimento/prospecao.
356 empresas (39 originais + 317 concorrentes) com formatacao profissional.
"""
import csv
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

BASE = "/home/headless/workspace/cih/clientes/pesquisacliente"

# ============================================================
# 1. READ ALL DATA
# ============================================================
rows = []
with open(f"{BASE}/db_consolidado_completo.csv", "r", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        rows.append(row)

print(f"Total de registros lidos: {len(rows)}")

# ============================================================
# 2. CLASSIFY AND CLEAN DATA
# ============================================================

def classify_setor(grupo, segmento):
    """Agrupa em setores mais limpos para as abas."""
    grupo_lower = grupo.lower()
    seg_lower = segmento.lower() if segmento else ""

    if "petroquimica" in grupo_lower or "petroquimica" in seg_lower or "quimic" in seg_lower:
        return "Petroquimica e Quimica"
    elif "celulose" in grupo_lower or "papel" in grupo_lower or "celulose" in seg_lower or "papel" in seg_lower:
        return "Celulose e Papel"
    elif "fertiliz" in grupo_lower or "fertiliz" in seg_lower:
        return "Fertilizantes"
    elif "refrat" in grupo_lower or "refrat" in seg_lower:
        return "Refratarios"
    elif "mineracao" in grupo_lower or "metalurg" in grupo_lower or "ferro" in seg_lower or "sider" in seg_lower or "miner" in seg_lower or "niob" in seg_lower or "silic" in seg_lower:
        return "Mineracao e Metalurgia"
    elif "energia" in grupo_lower or "energia" in seg_lower or "eletric" in seg_lower:
        return "Energia"
    elif "saneamento" in grupo_lower or "saneamento" in seg_lower:
        return "Saneamento"
    elif "saude" in grupo_lower or "hospital" in seg_lower or "diagnostic" in seg_lower or "plano" in seg_lower or "medic" in seg_lower:
        return "Saude"
    elif "educacao" in grupo_lower or "universidade" in seg_lower or "educacao" in seg_lower:
        return "Educacao"
    elif "varejo" in grupo_lower or "atacado" in grupo_lower or "supermerc" in seg_lower or "atacad" in seg_lower or "hipermercado" in seg_lower or "distribui" in seg_lower:
        return "Varejo e Atacado"
    elif "moveis" in grupo_lower or "movel" in seg_lower or "moveis" in seg_lower or "eletro" in seg_lower or "departamento" in seg_lower:
        return "Varejo Moveis"
    elif "moda" in grupo_lower or "fitness" in seg_lower or "vestuar" in seg_lower or "moda" in seg_lower or "activewear" in seg_lower:
        return "Moda e Vestuario"
    elif "alimento" in grupo_lower or "suplemento" in grupo_lower or "nutri" in seg_lower or "suplemento" in seg_lower or "cafe" in seg_lower:
        return "Alimentos e Suplementos"
    elif "franquia" in grupo_lower or "acai" in seg_lower or "franquia" in seg_lower or "acaiteria" in seg_lower:
        return "Franquias Alimentacao"
    elif "transporte" in grupo_lower or "transporte" in seg_lower or "onibus" in seg_lower:
        return "Transporte"
    elif "textil" in grupo_lower or "textil" in seg_lower or "linhas" in seg_lower or "fio" in seg_lower or "tecido" in seg_lower:
        return "Industria Textil"
    elif "lawtech" in grupo_lower or "juridic" in seg_lower or "lawtech" in seg_lower:
        return "Tecnologia / Lawtech"
    elif "inovacao" in grupo_lower or "aceleradora" in seg_lower or "inovacao" in seg_lower or "hub" in seg_lower:
        return "Tecnologia / Inovacao"
    elif "tech" in grupo_lower or "tecnologia" in grupo_lower or "consultoria" in seg_lower or "digital" in seg_lower:
        return "Tecnologia"
    elif "agricola" in grupo_lower or "agricol" in seg_lower or "insumos" in seg_lower:
        return "Distribuicao Agricola"
    elif grupo in ["INDUSTRIA"]:
        # Re-classify original industry companies
        if "refinaria" in seg_lower or "petroleo" in seg_lower:
            return "Petroquimica e Quimica"
        elif "celulose" in seg_lower:
            return "Celulose e Papel"
        return "Industria (Geral)"
    elif grupo in ["COMERCIO"]:
        if "farmacia" in seg_lower or "drogaria" in seg_lower:
            return "Varejo e Atacado"
        elif "cooperativa" in seg_lower:
            return "Distribuicao Agricola"
        return "Varejo e Atacado"
    elif grupo in ["SERVICOS"]:
        if "saude" in seg_lower or "hospital" in seg_lower or "diagnostic" in seg_lower:
            return "Saude"
        elif "educacao" in seg_lower or "universidade" in seg_lower:
            return "Educacao"
        elif "transporte" in seg_lower:
            return "Transporte"
        return "Servicos"
    elif grupo in ["TECNOLOGIA"]:
        return "Tecnologia"
    else:
        return "Outros"

def is_original(grupo):
    return grupo in ["INDUSTRIA", "COMERCIO", "SERVICOS", "TECNOLOGIA"]

def clean_gestor(gestor_rh, cargo_rh):
    """Limpa o campo de gestor para exibicao."""
    if not gestor_rh:
        return "", ""
    gestor = gestor_rh.strip()
    cargo = cargo_rh.strip() if cargo_rh else ""

    # Remove "Nao identificado" variants
    if "Nao identificado" in gestor or "Não identificado" in gestor:
        # Check if there's useful info in parentheses
        match = re.search(r"\((.+?)\)", gestor)
        if match:
            return match.group(1).strip(), cargo
        return "", cargo
    if "Contato via" in gestor or "A confirmar" in gestor:
        return "", ""
    return gestor, cargo

# Process all rows
processed = []
for row in rows:
    grupo = row.get("GRUPO", "")
    segmento = row.get("SEGMENTO", "")
    setor = classify_setor(grupo, segmento)
    tipo = "EMPRESA ORIGINAL" if is_original(grupo) else "CONCORRENTE"

    gestor, cargo = clean_gestor(row.get("GESTOR_RH", ""), row.get("CARGO_RH", ""))

    processed.append({
        "TIPO": tipo,
        "SETOR": setor,
        "EMPRESA": row.get("EMPRESA", ""),
        "SEGMENTO": segmento,
        "CIDADE_UF": row.get("CIDADE_UF", ""),
        "GESTOR_RH": gestor,
        "CARGO_RH": cargo,
        "LINKEDIN_GESTOR": row.get("LINKEDIN_GESTOR_RH", ""),
        "SITE": row.get("SITE_OFICIAL", ""),
        "LINKEDIN_EMPRESA": row.get("LINKEDIN_EMPRESA", ""),
        "EMAIL": row.get("EMAIL_CONTATO", ""),
        "TELEFONE": row.get("TELEFONE", ""),
        "PORTAL_VAGAS": row.get("PORTAL_VAGAS", ""),
        "GPTW": row.get("GPTW", ""),
        "STATUS": row.get("STATUS_MAPEAMENTO", ""),
        "OBSERVACOES": row.get("OBSERVACOES", ""),
    })

# ============================================================
# 3. CREATE EXCEL WORKBOOK
# ============================================================
wb = openpyxl.Workbook()

# --- STYLES ---
header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
header_fill_main = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_fill_contact = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
header_fill_info = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")

original_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # green tint
concorrente_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
original_font = Font(bold=True, size=10, name="Calibri", color="1F4E79")
normal_font = Font(size=10, name="Calibri")
gestor_font = Font(bold=True, size=10, name="Calibri", color="C00000")
link_font = Font(size=9, name="Calibri", color="0563C1", underline="single")

wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")
center = Alignment(wrap_text=True, vertical="top", horizontal="center")
thin_border = Border(
    left=Side(style='thin', color="D9D9D9"),
    right=Side(style='thin', color="D9D9D9"),
    top=Side(style='thin', color="D9D9D9"),
    bottom=Side(style='thin', color="D9D9D9"),
)

# Column config: (header, width, fill_group)
COLUMNS = [
    ("#", 4, "main"),
    ("TIPO", 14, "main"),
    ("SETOR", 22, "main"),
    ("EMPRESA", 32, "main"),
    ("SEGMENTO", 28, "main"),
    ("CIDADE / UF", 22, "main"),
    ("GESTOR DE RH", 30, "contact"),
    ("CARGO", 32, "contact"),
    ("LINKEDIN GESTOR", 42, "contact"),
    ("SITE OFICIAL", 35, "info"),
    ("LINKEDIN EMPRESA", 42, "info"),
    ("EMAIL / CONTATO", 32, "info"),
    ("TELEFONE", 18, "info"),
    ("PORTAL VAGAS", 35, "info"),
    ("GPTW", 10, "info"),
    ("OBSERVACOES", 45, "info"),
]

def write_sheet(ws, data_rows, sheet_title=None):
    """Write a formatted sheet with data."""
    # Header row
    for col_idx, (col_name, col_width, fill_group) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        if fill_group == "main":
            cell.fill = header_fill_main
        elif fill_group == "contact":
            cell.fill = header_fill_contact
        else:
            cell.fill = header_fill_info
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data rows
    for row_idx, item in enumerate(data_rows, 2):
        is_orig = item["TIPO"] == "EMPRESA ORIGINAL"
        has_gestor = bool(item["GESTOR_RH"])

        values = [
            row_idx - 1,
            item["TIPO"],
            item["SETOR"],
            item["EMPRESA"],
            item["SEGMENTO"],
            item["CIDADE_UF"],
            item["GESTOR_RH"],
            item["CARGO_RH"],
            item["LINKEDIN_GESTOR"],
            item["SITE"],
            item["LINKEDIN_EMPRESA"],
            item["EMAIL"],
            item["TELEFONE"],
            item["PORTAL_VAGAS"],
            item["GPTW"],
            item["OBSERVACOES"],
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val else "")
            cell.alignment = wrap
            cell.border = thin_border

            if is_orig:
                cell.fill = original_fill
                cell.font = original_font if col_idx <= 6 else normal_font
            else:
                cell.fill = concorrente_fill if (row_idx % 2 == 0) else alt_fill
                cell.font = normal_font

            # Highlight gestor name in red/bold
            if col_idx == 7 and has_gestor:
                cell.font = gestor_font

            # Make URLs clickable
            if col_idx in [9, 10, 11, 14] and val and val.startswith("http"):
                cell.font = link_font
                try:
                    cell.hyperlink = val
                except:
                    pass

    # Freeze panes and filters
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(data_rows)+1}"
    ws.sheet_properties.tabColor = "1F4E79"

    # Row height
    for row_idx in range(2, len(data_rows) + 2):
        ws.row_dimensions[row_idx].height = 28

    ws.row_dimensions[1].height = 32

# ============================================================
# 4. SORT DATA
# ============================================================

# Sort: originals first, then by setor, then by empresa name
def sort_key(item):
    tipo_order = 0 if item["TIPO"] == "EMPRESA ORIGINAL" else 1
    return (tipo_order, item["SETOR"], item["EMPRESA"])

processed.sort(key=sort_key)

# ============================================================
# 5. CREATE SHEETS
# ============================================================

# --- Sheet 1: CONSOLIDADO (ALL) ---
ws_all = wb.active
ws_all.title = "TODOS OS CONTATOS"
write_sheet(ws_all, processed)
ws_all.sheet_properties.tabColor = "1F4E79"
print(f"Aba 'TODOS OS CONTATOS': {len(processed)} registros")

# --- Sheet 2: EMPRESAS ORIGINAIS ---
originals = [r for r in processed if r["TIPO"] == "EMPRESA ORIGINAL"]
ws_orig = wb.create_sheet("EMPRESAS ORIGINAIS (39)")
write_sheet(ws_orig, originals)
ws_orig.sheet_properties.tabColor = "548235"
print(f"Aba 'EMPRESAS ORIGINAIS': {len(originals)} registros")

# --- Sheet 3: COM GESTOR IDENTIFICADO ---
with_gestor = [r for r in processed if r["GESTOR_RH"]]
ws_gestor = wb.create_sheet("COM GESTOR RH")
write_sheet(ws_gestor, with_gestor)
ws_gestor.sheet_properties.tabColor = "C00000"
print(f"Aba 'COM GESTOR RH': {len(with_gestor)} registros")

# --- Sheets by Sector ---
setores = {}
for item in processed:
    setor = item["SETOR"]
    if setor not in setores:
        setores[setor] = []
    setores[setor].append(item)

for setor_name in sorted(setores.keys()):
    setor_rows = setores[setor_name]
    # Truncate sheet name to 31 chars and remove invalid chars
    sheet_name = setor_name.replace("/", "-").replace("\\", "-").replace(":", "-")[:31]
    ws = wb.create_sheet(sheet_name)
    write_sheet(ws, setor_rows)
    print(f"Aba '{sheet_name}': {len(setor_rows)} registros")

# ============================================================
# 6. ADD SUMMARY SHEET (RESUMO)
# ============================================================
ws_resumo = wb.create_sheet("RESUMO", 0)
ws_resumo.sheet_properties.tabColor = "FFC000"

# Title
ws_resumo.merge_cells("A1:F1")
cell = ws_resumo["A1"]
cell.value = "BANCO DE DADOS DE PROSPECAO - GESTORES DE RH"
cell.font = Font(bold=True, size=16, color="1F4E79", name="Calibri")
cell.alignment = Alignment(horizontal="center", vertical="center")
ws_resumo.row_dimensions[1].height = 40

ws_resumo.merge_cells("A2:F2")
cell = ws_resumo["A2"]
cell.value = "CIH Consultoria | Data: 24/02/2026"
cell.font = Font(size=11, color="808080", name="Calibri")
cell.alignment = Alignment(horizontal="center")

# Stats
stats_start = 4
stats = [
    ("RESUMO GERAL", "", ""),
    ("Total de empresas mapeadas", len(processed), ""),
    ("Empresas originais (lista Bahia)", len(originals), ""),
    ("Concorrentes pesquisados", len(processed) - len(originals), ""),
    ("Empresas com gestor de RH identificado", len(with_gestor), f"{len(with_gestor)*100//len(processed)}%"),
    ("", "", ""),
    ("POR SETOR", "Qtd", "Com Gestor"),
]

for setor_name in sorted(setores.keys()):
    setor_rows = setores[setor_name]
    com_gestor = len([r for r in setor_rows if r["GESTOR_RH"]])
    stats.append((setor_name, len(setor_rows), com_gestor))

for i, (label, val, extra) in enumerate(stats):
    row = stats_start + i
    cell_a = ws_resumo.cell(row=row, column=1, value=label)
    cell_b = ws_resumo.cell(row=row, column=2, value=val)
    cell_c = ws_resumo.cell(row=row, column=3, value=extra)

    if label in ["RESUMO GERAL", "POR SETOR"]:
        cell_a.font = Font(bold=True, size=12, color="1F4E79", name="Calibri")
        cell_a.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        cell_b.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        cell_c.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    else:
        cell_a.font = Font(size=10, name="Calibri")
        cell_b.font = Font(bold=True, size=10, name="Calibri")
        cell_c.font = Font(size=10, color="808080", name="Calibri")

ws_resumo.column_dimensions["A"].width = 42
ws_resumo.column_dimensions["B"].width = 12
ws_resumo.column_dimensions["C"].width = 14

# Legend
legend_row = stats_start + len(stats) + 2
ws_resumo.cell(row=legend_row, column=1, value="LEGENDA DE CORES:").font = Font(bold=True, size=10, name="Calibri")

cell = ws_resumo.cell(row=legend_row+1, column=1, value="  Verde = Empresa original da lista (Bahia)")
cell.font = Font(size=10, name="Calibri")
cell.fill = original_fill

cell = ws_resumo.cell(row=legend_row+2, column=1, value="  Branco/Cinza = Concorrente pesquisado")
cell.font = Font(size=10, name="Calibri")

cell = ws_resumo.cell(row=legend_row+3, column=1, value="  Vermelho negrito = Gestor de RH identificado")
cell.font = gestor_font

# Navigation
nav_row = legend_row + 5
ws_resumo.cell(row=nav_row, column=1, value="ABAS DA PLANILHA:").font = Font(bold=True, size=10, name="Calibri")
nav_items = [
    "RESUMO - Esta aba (visao geral)",
    "TODOS OS CONTATOS - Base completa com 356 empresas",
    "EMPRESAS ORIGINAIS - 39 empresas da lista original (Bahia)",
    "COM GESTOR RH - Apenas empresas com gestor identificado",
    f"+ {len(setores)} abas por setor (Petroquimica, Saude, Varejo, etc.)",
]
for i, item in enumerate(nav_items):
    ws_resumo.cell(row=nav_row+1+i, column=1, value=f"  {item}").font = Font(size=10, name="Calibri")

# ============================================================
# 7. SAVE
# ============================================================
output = f"{BASE}/PROSPECAO_RH_COMPLETA_CIH.xlsx"
wb.save(output)
print(f"\n{'='*60}")
print(f"PLANILHA SALVA: {output}")
print(f"{'='*60}")
print(f"Total: {len(processed)} empresas")
print(f"Com gestor de RH: {len(with_gestor)}")
print(f"Abas: RESUMO + TODOS + ORIGINAIS + COM GESTOR + {len(setores)} setores")
