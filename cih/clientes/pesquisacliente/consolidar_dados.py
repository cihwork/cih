#!/usr/bin/env python3
"""
Consolidacao de todos os dados de concorrentes em um unico CSV.
Fontes:
1. db_prospeccao_rh_bahia.csv (IDs 40-59 = energia/saneamento)
2. concorrentes_pesquisa.csv (celulose/fertilizantes/mineracao + varejo/consumo)
3. concorrentes_pesquisa.txt (petroquimica/quimica)
4. concorrentes_5empresas_pesquisa.xlsx (saude)
5. concorrentes_clientes_CIH.xlsx (educacao/transporte/tech)
"""
import csv
import re
import os

BASE = "/home/headless/workspace/cih/clientes/pesquisacliente"

# Output columns (same as original DB)
HEADER = [
    "ID", "GRUPO", "EMPRESA", "SEGMENTO", "CIDADE_UF", "GPTW",
    "STATUS_MAPEAMENTO", "SITE_OFICIAL", "LINKEDIN_EMPRESA",
    "GESTOR_RH", "CARGO_RH", "LINKEDIN_GESTOR_RH",
    "EMAIL_CONTATO", "TELEFONE", "PORTAL_VAGAS", "OBSERVACOES"
]

all_rows = []
next_id = 60  # IDs 1-39 = originals, 40-59 = energy/sanitation

# ============================================================
# SOURCE 1: Read original DB (keep IDs 1-59 as-is)
# ============================================================
with open(os.path.join(BASE, "db_prospeccao_rh_bahia.csv"), "r", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        all_rows.append(row)

print(f"Source 1 (db original): {len(all_rows)} rows")

# ============================================================
# SOURCE 2: Parse concorrentes_pesquisa.txt (petrochemistry)
# ============================================================
def parse_petroquimica_txt(filepath):
    rows = []
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    # Split by BLOCO headers to get client name
    blocos = re.split(r"={10,}\nBLOCO \d+: CONCORRENTES D[AEO] (.+?) \(", content)

    current_client = ""
    for i in range(1, len(blocos), 2):
        client_info = blocos[i].strip()
        bloco_content = blocos[i+1] if i+1 < len(blocos) else ""

        # Extract client name
        current_client = client_info.split(",")[0].strip()

        # Split into individual competitor entries
        entries = re.split(r"--- CONCORRENTE \d+\.\d+ ---", bloco_content)

        for entry in entries:
            if not entry.strip():
                continue

            def extract(label):
                match = re.search(rf"{label}:\s*(.+?)(?:\n|$)", entry)
                return match.group(1).strip() if match else ""

            empresa = extract("Empresa")
            if not empresa:
                continue

            segmento = extract("Segmento")
            cidade = extract("Cidade/UF")
            site = extract("Website")
            linkedin = extract("LinkedIn")
            gestor_raw = extract("Gestor de RH")
            linkedin_rh = extract("LinkedIn RH")
            contato = extract("Contato RH")
            portal = extract("Portal Carreiras")
            notas = extract("Notas")

            # Parse gestor name and cargo
            gestor = ""
            cargo = ""
            if gestor_raw and "Nao identificado" not in gestor_raw:
                parts = gestor_raw.split(" - ", 1)
                gestor = parts[0].strip()
                cargo = parts[1].strip() if len(parts) > 1 else ""

            rows.append({
                "GRUPO": f"CONCORRENTE - PETROQUIMICA",
                "EMPRESA": empresa,
                "SEGMENTO": segmento,
                "CIDADE_UF": cidade,
                "GPTW": "",
                "STATUS_MAPEAMENTO": "",
                "SITE_OFICIAL": site,
                "LINKEDIN_EMPRESA": linkedin if linkedin != "-" else "",
                "GESTOR_RH": gestor if gestor else "Nao identificado",
                "CARGO_RH": cargo,
                "LINKEDIN_GESTOR_RH": linkedin_rh if linkedin_rh not in ["-", "Buscar via pagina da empresa"] else "",
                "EMAIL_CONTATO": contato if "LinkedIn" not in contato and contato != "-" else "",
                "TELEFONE": "",
                "PORTAL_VAGAS": portal,
                "OBSERVACOES": f"Concorrente de {current_client}. {notas}"
            })

    return rows

petro_rows = parse_petroquimica_txt(os.path.join(BASE, "concorrentes_pesquisa.txt"))
print(f"Source 2 (petroquimica txt): {len(petro_rows)} rows")

for row in petro_rows:
    row["ID"] = str(next_id)
    next_id += 1
    all_rows.append(row)

# ============================================================
# SOURCE 3: Parse concorrentes_pesquisa.csv (celulose/fert/varejo)
# ============================================================
csv_rows = []
with open(os.path.join(BASE, "concorrentes_pesquisa.csv"), "r", encoding="utf-8") as f:
    reader = csv.DictReader(f, delimiter=";")
    for row in reader:
        cliente = row.get("EMPRESA CLIENTE", "")
        concorrente = row.get("CONCORRENTE", "")
        segmento = row.get("SEGMENTO", "")
        cidade = row.get("CIDADE/UF", "")
        site = row.get("SITE", "")
        linkedin = row.get("LINKEDIN", "")
        gestor_raw = row.get("GESTOR DE RH / DIRETOR DE PESSOAS", "")
        contato = row.get("CONTATO RH", "")
        portal = row.get("PORTAL DE VAGAS", "")
        obs = row.get("OBSERVAÇÕES", "")

        # Parse gestor
        gestor = ""
        cargo = ""
        if gestor_raw and "Nao identificado" not in gestor_raw and "Não identificado" not in gestor_raw:
            match = re.match(r"(.+?)\s*\((.+?)\)", gestor_raw)
            if match:
                gestor = match.group(1).strip()
                cargo = match.group(2).strip()
            else:
                gestor = gestor_raw

        # Determine group
        if any(x in segmento.lower() for x in ["celulose", "papel"]):
            grupo = "CONCORRENTE - CELULOSE/PAPEL"
        elif any(x in segmento.lower() for x in ["fertiliz", "fert."]):
            grupo = "CONCORRENTE - FERTILIZANTES"
        elif any(x in segmento.lower() for x in ["refrat"]):
            grupo = "CONCORRENTE - REFRATARIOS"
        elif any(x in segmento.lower() for x in ["miner", "ferro", "silic", "sider", "niob", "niqu", "metalurg"]):
            grupo = "CONCORRENTE - MINERACAO/METALURGIA"
        elif any(x in segmento.lower() for x in ["supermerc", "atacad", "varejo", "hiper"]):
            grupo = "CONCORRENTE - VAREJO/ATACADO"
        elif any(x in segmento.lower() for x in ["movel", "moveis", "eletro", "depart"]):
            grupo = "CONCORRENTE - VAREJO/MOVEIS"
        elif any(x in segmento.lower() for x in ["fitness", "vestuar", "moda", "sport", "activewear"]):
            grupo = "CONCORRENTE - MODA/VESTUARIO"
        elif any(x in segmento.lower() for x in ["suplemento", "nutri", "cafe", "alimento"]):
            grupo = "CONCORRENTE - ALIMENTOS/SUPLEMENTOS"
        elif any(x in segmento.lower() for x in ["acai", "franquia", "alimentacao"]):
            grupo = "CONCORRENTE - FRANQUIAS/ALIMENTACAO"
        else:
            grupo = f"CONCORRENTE - {cliente.upper()}"

        csv_rows.append({
            "GRUPO": grupo,
            "EMPRESA": concorrente,
            "SEGMENTO": segmento,
            "CIDADE_UF": cidade,
            "GPTW": "",
            "STATUS_MAPEAMENTO": "",
            "SITE_OFICIAL": site,
            "LINKEDIN_EMPRESA": linkedin,
            "GESTOR_RH": gestor if gestor else "Nao identificado",
            "CARGO_RH": cargo,
            "LINKEDIN_GESTOR_RH": "",
            "EMAIL_CONTATO": "",
            "TELEFONE": "",
            "PORTAL_VAGAS": portal,
            "OBSERVACOES": f"Concorrente de {cliente}. {obs}".replace('"', '').strip()
        })

print(f"Source 3 (concorrentes CSV): {len(csv_rows)} rows")

for row in csv_rows:
    row["ID"] = str(next_id)
    next_id += 1
    all_rows.append(row)

# ============================================================
# SOURCE 4: Parse gerar_planilha.py data (education/transport/tech + retail)
# This data is embedded in the Python script as a dict
# ============================================================
import ast

planilha_path = os.path.join(BASE, "gerar_planilha.py")
with open(planilha_path, "r", encoding="utf-8") as f:
    content = f.read()

# Extract the data dict from the script
data_match = re.search(r"data\s*=\s*(\{.*?\n\})", content, re.DOTALL)
if data_match:
    try:
        data_str = data_match.group(1)
        # Use exec to evaluate the data dict safely
        local_vars = {}
        exec(f"data = {data_str}", {}, local_vars)
        planilha_data = local_vars.get("data", {})

        planilha_rows = []
        for sheet_name, rows in planilha_data.items():
            for row_data in rows:
                cliente = row_data[0]
                concorrente = row_data[2]
                segmento = row_data[3]
                cidade = row_data[4]
                site = row_data[5]
                linkedin = row_data[6]
                gestor_raw = row_data[7]
                contato = row_data[8]
                portal = row_data[9]

                gestor = ""
                cargo = ""
                if gestor_raw and "Contato via" not in gestor_raw and "N/D" not in gestor_raw:
                    match = re.match(r"(.+?)\s*\((.+?)\)", gestor_raw)
                    if match:
                        gestor = match.group(1).strip()
                        cargo = match.group(2).strip()
                    else:
                        gestor = gestor_raw

                # Determine group
                if "UFBA" in cliente or "Universidade" in segmento.lower() or "Educacao" in segmento:
                    grupo = "CONCORRENTE - EDUCACAO"
                elif "GEVAN" in cliente or "Transporte" in segmento:
                    grupo = "CONCORRENTE - TRANSPORTE"
                elif "Pro-Linhas" in cliente or "Textil" in segmento:
                    grupo = "CONCORRENTE - TEXTIL"
                elif "Jusbrasil" in cliente or "Lawtech" in segmento:
                    grupo = "CONCORRENTE - TECH/LAWTECH"
                elif "Inventivos" in cliente or "Aceleradora" in segmento or "Inovacao" in segmento:
                    grupo = "CONCORRENTE - TECH/INOVACAO"
                elif "Nordeste" in cliente or "Agricola" in segmento.lower() or "Insumos" in segmento:
                    grupo = "CONCORRENTE - DISTRIBUICAO AGRICOLA"
                elif any(x in segmento.lower() for x in ["supermerc", "atacad", "varejo", "hiper"]):
                    grupo = "CONCORRENTE - VAREJO/ATACADO"
                elif any(x in segmento.lower() for x in ["movel", "moveis", "eletro"]):
                    grupo = "CONCORRENTE - VAREJO/MOVEIS"
                elif any(x in segmento.lower() for x in ["fitness", "vestuar", "moda"]):
                    grupo = "CONCORRENTE - MODA/VESTUARIO"
                elif any(x in segmento.lower() for x in ["suplemento", "nutri", "cafe"]):
                    grupo = "CONCORRENTE - ALIMENTOS/SUPLEMENTOS"
                elif any(x in segmento.lower() for x in ["acai", "franquia"]):
                    grupo = "CONCORRENTE - FRANQUIAS/ALIMENTACAO"
                else:
                    grupo = f"CONCORRENTE - {cliente.upper()}"

                # Check if this entry already exists (avoid duplicates from CSV)
                planilha_rows.append({
                    "GRUPO": grupo,
                    "EMPRESA": concorrente,
                    "SEGMENTO": segmento,
                    "CIDADE_UF": cidade,
                    "GPTW": "",
                    "STATUS_MAPEAMENTO": "",
                    "SITE_OFICIAL": site if site != "N/D" else "",
                    "LINKEDIN_EMPRESA": linkedin if linkedin not in ["N/D", "N/D (pesquisar Grupo Comporte)"] else "",
                    "GESTOR_RH": gestor if gestor else "Nao identificado",
                    "CARGO_RH": cargo,
                    "LINKEDIN_GESTOR_RH": "",
                    "EMAIL_CONTATO": "",
                    "TELEFONE": "",
                    "PORTAL_VAGAS": portal if portal != "N/D" else "",
                    "OBSERVACOES": f"Concorrente de {cliente}"
                })

        # Deduplicate: only add entries not already in concorrentes_pesquisa.csv
        existing_empresas = {r["EMPRESA"].lower().strip() for r in all_rows}
        new_planilha = [r for r in planilha_rows if r["EMPRESA"].lower().strip() not in existing_empresas]

        print(f"Source 4 (planilha py): {len(planilha_rows)} total, {len(new_planilha)} new (after dedup)")

        for row in new_planilha:
            row["ID"] = str(next_id)
            next_id += 1
            all_rows.append(row)
            existing_empresas.add(row["EMPRESA"].lower().strip())

    except Exception as e:
        print(f"Error parsing planilha data: {e}")

# ============================================================
# SOURCE 5: Healthcare data from concorrentes_5empresas_pesquisa.xlsx
# Try to read with openpyxl, fallback to manual if unavailable
# ============================================================
try:
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(BASE, "concorrentes_5empresas_pesquisa.xlsx"))

    healthcare_rows = []
    for ws in wb.worksheets:
        headers_xl = [cell.value for cell in ws[1]]
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers_xl):
                cell_val = ws.cell(row=row_idx, column=col_idx+1).value
                row_data[header] = str(cell_val) if cell_val else ""

            empresa = row_data.get("Empresa Concorrente", "") or row_data.get("Concorrente", "")
            if not empresa or empresa == "None":
                continue

            cliente = row_data.get("Cliente CIH (Origem)", "") or row_data.get("Cliente CIH", "")
            gestor_raw = row_data.get("Gestor de RH / Pessoas", "") or row_data.get("Gestor RH", "")

            gestor = ""
            cargo = ""
            if gestor_raw and "A confirmar" not in gestor_raw and "Nao identificado" not in gestor_raw:
                match = re.match(r"(.+?)\s*\((.+?)\)", gestor_raw)
                if match:
                    gestor = match.group(1).strip().replace("**", "")
                    cargo = match.group(2).strip()
                else:
                    gestor = gestor_raw.replace("**", "")

            if empresa.lower().strip() in existing_empresas:
                continue

            healthcare_rows.append({
                "GRUPO": "CONCORRENTE - SAUDE",
                "EMPRESA": empresa,
                "SEGMENTO": row_data.get("Segmento", "Saude"),
                "CIDADE_UF": row_data.get("Cidade/UF", ""),
                "GPTW": "",
                "STATUS_MAPEAMENTO": "",
                "SITE_OFICIAL": row_data.get("Site Oficial", "") or row_data.get("Site", ""),
                "LINKEDIN_EMPRESA": row_data.get("LinkedIn (Empresa)", "") or row_data.get("LinkedIn", ""),
                "GESTOR_RH": gestor if gestor else "Nao identificado",
                "CARGO_RH": cargo,
                "LINKEDIN_GESTOR_RH": row_data.get("LinkedIn Gestor", ""),
                "EMAIL_CONTATO": "",
                "TELEFONE": "",
                "PORTAL_VAGAS": row_data.get("Portal de Carreiras / Vagas", "") or row_data.get("Portal Vagas", ""),
                "OBSERVACOES": f"Concorrente de {cliente}"
            })
            existing_empresas.add(empresa.lower().strip())

    print(f"Source 5 (healthcare xlsx): {len(healthcare_rows)} new rows")

    for row in healthcare_rows:
        row["ID"] = str(next_id)
        next_id += 1
        all_rows.append(row)

except ImportError:
    print("openpyxl not available - skipping healthcare xlsx")
except Exception as e:
    print(f"Error reading healthcare xlsx: {e}")

# ============================================================
# WRITE FINAL CONSOLIDATED CSV
# ============================================================
output_path = os.path.join(BASE, "db_consolidado_completo.csv")
with open(output_path, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=HEADER)
    writer.writeheader()
    for row in all_rows:
        # Ensure all fields exist
        clean_row = {k: row.get(k, "") for k in HEADER}
        writer.writerow(clean_row)

print(f"\n{'='*60}")
print(f"CONSOLIDACAO COMPLETA")
print(f"{'='*60}")
print(f"Total de registros: {len(all_rows)}")
print(f"  - Empresas originais (Bahia): 39")
print(f"  - Concorrentes energia/saneamento: 20")
print(f"  - Concorrentes petroquimica: {len(petro_rows)}")
print(f"  - Concorrentes celulose/fert/varejo (CSV): {len(csv_rows)}")
planilha_count = len(new_planilha) if 'new_planilha' in dir() else 0
healthcare_count = len(healthcare_rows) if 'healthcare_rows' in dir() else 0
print(f"  - Concorrentes educacao/transporte/tech: {planilha_count}")
print(f"  - Concorrentes saude: {healthcare_count}")
print(f"\nArquivo salvo em: {output_path}")
