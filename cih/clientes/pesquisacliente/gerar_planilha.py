import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Style definitions
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
client_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
client_font = Font(bold=True, size=11, color="1F3864")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

headers = [
    "Cliente CIH (Origem)",
    "#",
    "Empresa Concorrente",
    "Segmento",
    "Cidade/UF",
    "Site Oficial",
    "LinkedIn (Empresa)",
    "Gestor de RH / Pessoas",
    "Contato RH (E-mail / Telefone)",
    "Portal de Carreiras / Vagas"
]

col_widths = [28, 4, 32, 30, 22, 38, 48, 35, 38, 42]

# ===================== DATA =====================

data = {
    "1. UFBA - Universidades Federais": [
        ["UFBA", 1, "UFRJ - Universidade Federal do Rio de Janeiro", "Educacao Superior / Universidade Federal", "Rio de Janeiro/RJ", "https://ufrj.br", "https://www.linkedin.com/school/ufrj/", "Rafael dos Santos Pereira (Sup. Geral de Pessoal / Pro-Reitor Substituto PR4)", "rafaelpereira@pr4.ufrj.br / (21) 3938-1670", "https://concursos.pr4.ufrj.br"],
        ["UFBA", 2, "UFMG - Universidade Federal de Minas Gerais", "Educacao Superior / Universidade Federal", "Belo Horizonte/MG", "https://ufmg.br", "https://www.linkedin.com/school/ufmg/", "Maria Marcia Magela Machado (Pro-Reitora de Recursos Humanos)", "info@prorh.ufmg.br", "https://www.ufmg.br/prorh/"],
        ["UFBA", 3, "UFRGS - Universidade Federal do Rio Grande do Sul", "Educacao Superior / Universidade Federal", "Porto Alegre/RS", "https://www.ufrgs.br", "https://www.linkedin.com/school/ufrgs/", "Marilia Borges Hackmann (Superintendente de Gestao de Pessoas)", "progesp@ufrgs.br", "https://www.ufrgs.br/progesp/"],
        ["UFBA", 4, "UnB - Universidade de Brasilia", "Educacao Superior / Universidade Federal", "Brasilia/DF", "https://www.unb.br", "https://www.linkedin.com/school/universidade-de-brasilia/", "Maria do Socorro Mendes (Decana de Gestao de Pessoas)", "(61) 3107-0578", "https://dgp.unb.br/"],
        ["UFBA", 5, "UFPR - Universidade Federal do Parana", "Educacao Superior / Universidade Federal", "Curitiba/PR", "https://www.ufpr.br", "https://www.linkedin.com/school/ufpr/", "Dulcileia Goncalves (Pro-Reitora de Gestao de Pessoas)", "progepe@ufpr.br / (41) 3360-4500", "https://progepe.ufpr.br/"],
        ["UFBA", 6, "UFPE - Universidade Federal de Pernambuco", "Educacao Superior / Universidade Federal", "Recife/PE", "https://www.ufpe.br", "https://www.linkedin.com/school/ufpe/", "Brunna Carvalho Almeida Granja (Pro-Reitora de Gestao de Pessoas e QV - PROGEPE)", "progepe@ufpe.br", "https://www.ufpe.br/progepe"],
        ["UFBA", 7, "UFSC - Universidade Federal de Santa Catarina", "Educacao Superior / Universidade Federal", "Florianopolis/SC", "https://ufsc.br", "https://www.linkedin.com/school/ufsc/", "Sandra Regina Carrieri de Souza (Pro-Reitora PRODEGESP)", "prodegesp@contato.ufsc.br", "https://prodegesp.ufsc.br/"],
        ["UFBA", 8, "UFC - Universidade Federal do Ceara", "Educacao Superior / Universidade Federal", "Fortaleza/CE", "https://www.ufc.br", "https://www.linkedin.com/school/uaborfc/", "Marilene Feitosa Soares (Pro-Reitora de Gestao de Pessoas - PROGEP)", "(85) 3366-7390", "https://progep.ufc.br/"],
        ["UFBA", 9, "UFF - Universidade Federal Fluminense", "Educacao Superior / Universidade Federal", "Niteroi/RJ", "https://www.uff.br", "https://www.linkedin.com/school/uffoficial/", "Aline da Silva Marques (Pro-Reitora de Gestao de Pessoas - PROGEPE)", "secretaria.progepe@id.uff.br", "https://www.uff.br/?q=pro-reitoria-de-gestao-de-pessoas"],
        ["UFBA", 10, "UFPA - Universidade Federal do Para", "Educacao Superior / Universidade Federal", "Belem/PA", "https://ufpa.br", "https://www.linkedin.com/school/ufpa/", "Icaro Duarte Pastana (Pro-Reitor de Desenv. e Gestao de Pessoal - PROGEP)", "progep@ufpa.br", "https://www.progep.ufpa.br/"],
    ],

    "2. GEVAN - Transporte Urbano": [
        ["Grupo Evangelista/GEVAN", 1, "Sambaiba Transportes Urbanos", "Transporte Urbano / Onibus Municipal", "Sao Paulo/SP", "https://www.sambaibasp.com.br", "https://br.linkedin.com/company/sambaiba-transportes-urbanos-ltda", "Larissa Tigano (Gerente de RH)", "Contato via LinkedIn / site", "https://www.sambaibasp.com.br/trabalhe-conosco"],
        ["Grupo Evangelista/GEVAN", 2, "Grupo JCA (Cometa/1001/Catarinense)", "Transporte Rodoviario e Urbano de Passageiros", "Niteroi/RJ", "https://jcaholding.com.br", "https://br.linkedin.com/company/grupo-jca-holding", "Contato via portal de carreiras", "Contato via Gupy", "https://embarquejca.gupy.io/"],
        ["Grupo Evangelista/GEVAN", 3, "Grupo Guanabara", "Transporte Urbano e Rodoviario", "Recife/PE", "http://www.tguanabara.com.br", "https://br.linkedin.com/company/grupo-guanabara", "Contato via LinkedIn", "Contato via LinkedIn", "https://www.linkedin.com/company/grupo-guanabara/jobs"],
        ["Grupo Evangelista/GEVAN", 4, "Rapido Araguaia", "Transporte Urbano / Onibus Municipal", "Goiania/GO", "https://www.rapidoaraguaia.com.br", "https://br.linkedin.com/company/rapido-araguaia", "Contato via WhatsApp RH: (62) 98523-0289", "(62) 98523-0289", "https://www.rapidoaraguaia.com.br/trabalhe-conosco"],
        ["Grupo Evangelista/GEVAN", 5, "HP Transportes", "Transporte Urbano / Onibus Metropolitano", "Goiania/GO", "https://www.hptransportes.com.br", "https://br.linkedin.com/company/hp-transportes-coletivos", "Contato via portal de carreiras", "Contato via site", "https://redemobconsorcio.rhgestor.com.br/"],
        ["Grupo Evangelista/GEVAN", 6, "Grupo Leblon Transporte", "Transporte Urbano / Onibus Metropolitano", "Fazenda Rio Grande/PR", "https://grupoleblon.com.br", "https://www.linkedin.com/company/grupo-leblon-transporte-passageiros", "Contato via LinkedIn", "Contato via site", "https://grupoleblon.com.br/trabalhe-conosco"],
        ["Grupo Evangelista/GEVAN", 7, "Viacao Pioneira", "Transporte Urbano / Onibus Municipal", "Brasilia/DF", "https://vpioneirabsb.com.br", "https://br.linkedin.com/company/viacao-pioneira", "Contato via LinkedIn", "Contato via site", "https://vpioneirabsb.com.br/contato/"],
        ["Grupo Evangelista/GEVAN", 8, "Grupo Aguia Branca (Viacao)", "Transporte Rodoviario e Urbano / Logistica", "Cariacica/ES", "https://www.aguiabranca.com.br", "https://br.linkedin.com/company/grupo-aguia-branca", "Contato via portal de carreiras GPTW", "Contato via Empregare", "https://aguiabranca.empregare.com/pt-br"],
        ["Grupo Evangelista/GEVAN", 9, "Auto Omnibus Nova Suissa", "Transporte Urbano / Onibus Municipal", "Belo Horizonte/MG", "N/D", "N/D", "Fabio Couto de Araujo Cancado (Diretor)", "(31) 3314-5950", "N/D"],
        ["Grupo Evangelista/GEVAN", 10, "Grupo Comporte", "Transporte Urbano e Rodoviario / Holding", "Sao Bernardo do Campo/SP", "N/D", "N/D (pesquisar Grupo Comporte)", "Contato via empresas do grupo", "Contato via empresas do grupo", "N/D"],
    ],

    "3. Pro-Linhas - Industria Textil": [
        ["Pro-Linhas Nordeste", 1, "Vicunha Textil", "Industria Textil / Fiacao, Tecelagem, Confeccao", "Sao Paulo/SP", "https://www.vicunha.com.br", "https://br.linkedin.com/company/vicunha", "Tereza Lopes (Gerente de RH Corporativo) / Alexandre Ferreira (Dir. Corp. RH)", "Contato via LinkedIn", "https://www.vicunha.com.br/trabalhe-conosco"],
        ["Pro-Linhas Nordeste", 2, "Textil Canatiba", "Industria Textil / Tecelagem de Algodao", "Santa Barbara D'Oeste/SP", "https://canatiba.com.br", "https://br.linkedin.com/company/canatibatextil", "Contato via LinkedIn", "Contato via site: canatiba.com.br/contato", "https://canatiba.com.br/trabalhe-conosco"],
        ["Pro-Linhas Nordeste", 3, "Linhas Corrente (ex-Coats Corrente)", "Industria Textil / Linhas para Costura e Bordado", "Sao Paulo/SP", "https://linhascorrente.com.br", "https://br.linkedin.com/company/linhas-corrente", "Contato via LinkedIn", "Contato via site", "https://linhascorrente.com.br/trabalhe-conosco"],
        ["Pro-Linhas Nordeste", 4, "Circulo Ltda", "Industria Textil / Fios e Linhas para Artesanato", "Gaspar/SC", "https://www.circulo.com.br", "https://br.linkedin.com/company/circulo-s-a", "Contato via LinkedIn", "Contato via site: circulo.com.br/contato", "https://circulo.enlizt.me/"],
        ["Pro-Linhas Nordeste", 5, "Linhanyl Paraguacu", "Industria Textil / Linhas de Nylon e Poliester", "Sorocaba/SP", "https://www.linhanyl.com.br", "https://www.linkedin.com/company/linhanyl-s-a-linhas-para-coser", "Contato via LinkedIn", "Contato via site", "N/D"],
        ["Pro-Linhas Nordeste", 6, "Cedro Textil", "Industria Textil / Denim e Tecidos Profissionais", "Caetanopolis/MG", "https://cedro.com.br", "https://br.linkedin.com/company/cedro-textil", "Contato via LinkedIn", "(31) 3714-7900", "https://cedro.com.br/trabalhe-conosco/"],
        ["Pro-Linhas Nordeste", 7, "Dohler SA", "Industria Textil / Tecidos para Casa e Industria", "Joinville/SC", "https://www.dohler.com.br", "https://br.linkedin.com/company/dohler-sa", "Contato via LinkedIn", "Contato via site", "https://www.dohler.com.br/trabalhe-conosco"],
        ["Pro-Linhas Nordeste", 8, "Linhas Setta", "Industria Textil / Linhas para Costura e Bordado", "Sao Bernardo do Campo/SP", "N/D", "https://www.linkedin.com/company/linhas-setta-ltda", "Contato via LinkedIn", "Contato via LinkedIn", "N/D"],
        ["Pro-Linhas Nordeste", 9, "Coteminas / Santanense", "Industria Textil / Cama Mesa Banho e Uniformes", "Montes Claros/MG", "https://www.coteminas.com.br", "https://br.linkedin.com/company/coteminas", "Contato via LinkedIn", "Contato via site", "N/D"],
        ["Pro-Linhas Nordeste", 10, "Cooperlinhas", "Industria Textil / Cooperativa de Producao de Linhas", "Mogi das Cruzes/SP", "N/D", "N/D", "Contato direto", "N/D", "N/D"],
    ],

    "4. Jusbrasil - Lawtech": [
        ["Jusbrasil", 1, "Projuris (by Softplan)", "Lawtech / Software Juridico para Escritorios e Empresas", "Joinville/SC", "https://www.projuris.com.br", "https://br.linkedin.com/company/projuris", "Alejandra Nadruz (Diretora de Gente e Cultura - Softplan)", "Contato via LinkedIn / Gupy", "https://softplan.gupy.io/"],
        ["Jusbrasil", 2, "Aurum Software", "Lawtech / Software Juridico (Themis e Astrea)", "Florianopolis/SC", "https://www.aurum.com.br", "https://br.linkedin.com/company/aurumbr", "Larissa Baldissarelli (Diretora de Gestao de Pessoas)", "Contato via LinkedIn", "https://aurum.gupy.io/"],
        ["Jusbrasil", 3, "Escavador", "Lawtech / Pesquisa Juridica e Due Diligence", "Salvador/BA", "https://www.escavador.com", "https://www.linkedin.com/company/escavador", "Contato via LinkedIn", "Contato via LinkedIn", "https://www.linkedin.com/company/escavador/jobs"],
        ["Jusbrasil", 4, "netLex", "Lawtech / Gestao de Contratos (CLM)", "Belo Horizonte/MG", "https://netlex.io", "https://br.linkedin.com/company/netlex-io", "Contato via LinkedIn", "Contato via LinkedIn", "https://netlex.io/carreiras"],
        ["Jusbrasil", 5, "Docket", "Lawtech / Plataforma de Documentos e Cartorio", "Sao Paulo/SP", "https://www.docket.com.br", "https://www.linkedin.com/company/docketbrasil", "Contato via LinkedIn", "Contato via LinkedIn", "https://www.linkedin.com/company/docketbrasil/jobs"],
        ["Jusbrasil", 6, "Thomson Reuters (Legal One)", "Lawtech / Pesquisa Juridica e Gestao Legal", "Sao Paulo/SP", "https://www.thomsonreuters.com.br", "https://br.linkedin.com/showcase/thomson-reuters-legal", "Contato via LinkedIn (TR Brasil)", "Contato via site TR", "https://careers.thomsonreuters.com/"],
        ["Jusbrasil", 7, "Benner (Juridico)", "Lawtech / ERP Juridico e Software de Gestao", "Blumenau/SC", "https://benner.com.br/juridico-sistema/", "https://br.linkedin.com/company/universobenner", "Contato via LinkedIn", "Contato via Gupy", "https://vemserbenner.gupy.io/"],
        ["Jusbrasil", 8, "Turivius", "Lawtech / Jurimetria e Pesquisa Jurisprudencial", "Sao Paulo/SP", "https://turivius.com", "https://br.linkedin.com/company/turivius", "Contato via LinkedIn", "Contato via LinkedIn", "N/D"],
        ["Jusbrasil", 9, "Looplex", "Lawtech / Automacao de Documentos Juridicos", "Sao Paulo/SP", "https://www.looplex.com.br", "https://www.linkedin.com/company/looplex", "Contato via LinkedIn", "Contato via LinkedIn", "N/D"],
        ["Jusbrasil", 10, "Digesto Tecnologia", "Lawtech / IA para Dados Juridicos", "Curitiba/PR", "https://www.digesto.com.br", "N/D", "Contato via LinkedIn", "Contato via site", "N/D"],
    ],

    "5. Inventivos - Aceleradoras": [
        ["Inventivos Tecnologia", 1, "ACE Ventures (ex-ACE Startups)", "Aceleradora de Startups / Inovacao Corporativa", "Sao Paulo/SP", "https://aceventures.com.br", "https://br.linkedin.com/company/aceventuresbr", "Contato via LinkedIn", "Contato via site", "https://aceventures.com.br/carreiras"],
        ["Inventivos Tecnologia", 2, "Distrito", "Hub de Inovacao / Aceleracao e Open Innovation", "Sao Paulo/SP", "https://www.distrito.me", "https://br.linkedin.com/company/distrito.me", "Contato via LinkedIn", "Contato via site", "https://www.distrito.me/carreiras"],
        ["Inventivos Tecnologia", 3, "Wayra Brasil (Telefonica/Vivo)", "Aceleradora Corporativa / Venture Capital", "Sao Paulo/SP", "https://br.wayra.com", "https://www.linkedin.com/company/wayrabrasil", "Contato via LinkedIn", "Contato via site", "N/D"],
        ["Inventivos Tecnologia", 4, "Cubo Itau", "Hub de Inovacao / Ecossistema de Startups", "Sao Paulo/SP", "https://cubo.network", "https://br.linkedin.com/company/cubo-network", "Contato via LinkedIn", "Contato via site", "https://cubo.network/carreiras"],
        ["Inventivos Tecnologia", 5, "Endeavor Brasil", "Aceleradora / Apoio a Scale-ups", "Sao Paulo/SP", "https://endeavor.org.br", "https://br.linkedin.com/company/endeavor-brasil", "Gerente de Gente & Gestao (vaga aberta recente)", "Contato via LinkedIn / Gupy", "https://endeavor.gupy.io/"],
        ["Inventivos Tecnologia", 6, "Darwin Startups", "Aceleradora de Startups / Venture Capital", "Florianopolis/SC", "https://www.darwinstartups.com", "https://br.linkedin.com/company/darwinstartups", "Contato via LinkedIn", "Contato via site", "N/D"],
        ["Inventivos Tecnologia", 7, "Troposlab", "Consultoria de Inovacao / Aceleracao de Startups", "Belo Horizonte/MG", "https://troposlab.com", "https://br.linkedin.com/company/troposlab", "Renata Horta (Diretora / Cofundadora)", "Contato via LinkedIn", "https://troposlab.com/carreiras"],
        ["Inventivos Tecnologia", 8, "Baita Aceleradora", "Aceleradora de Startups / Programas Corporativos", "Campinas/SP", "https://baita.ac", "https://br.linkedin.com/company/baita", "Contato via LinkedIn", "Contato via site", "N/D"],
        ["Inventivos Tecnologia", 9, "HOTMILK (PUCPR)", "Ecossistema de Inovacao / Aceleracao Universitaria", "Curitiba/PR", "https://hotmilk.pucpr.br", "https://br.linkedin.com/company/hotmilk-pucpr", "Fernando Bittencourt Luciano (Diretor)", "Contato via LinkedIn / PUCPR", "https://hotmilk.pucpr.br/"],
        ["Inventivos Tecnologia", 10, "InovAtiva Brasil (Sebrae/MDIC)", "Aceleradora Publica / Programa de Governo", "Brasilia/DF", "https://www.inovativa.online", "https://br.linkedin.com/company/inovativahub", "Contato via LinkedIn / Sebrae", "Contato via site", "https://www.inovativa.online/inovativa-brasil/"],
    ],

    "6. Nordeste Atacado - Dist. Agricola": [
        ["Nordeste Atacado", 1, "Lavoro Agro", "Distribuicao de Insumos Agricolas", "Sao Paulo/SP (HQ)", "https://www.lavoroagro.com", "https://www.linkedin.com/company/lavoroagro", "Contato via LinkedIn", "Contato via site", "https://www.lavoroagro.com/work-with-us/"],
        ["Nordeste Atacado", 2, "Belagricola", "Distribuicao de Insumos Agricolas e Graos", "Londrina/PR", "https://www.belagricola.com.br", "https://br.linkedin.com/company/belagricola", "Contato via LinkedIn", "Contato via site", "https://www.linkedin.com/company/belagricola/jobs"],
        ["Nordeste Atacado", 3, "Agro Amazonia (Sumitomo)", "Distribuicao de Insumos Agricolas e Agrotecnologia", "Cuiaba/MT", "https://agroamazonia.com", "https://br.linkedin.com/company/agroamazonia", "Contato via LinkedIn", "Contato via site", "https://carreiras.agroamazonia.com.br/"],
        ["Nordeste Atacado", 4, "Nutrien Solucoes Agricolas", "Distribuicao de Insumos Agricolas", "Sao Paulo/SP", "https://www.nutrien.com.br", "https://br.linkedin.com/company/nutrienbr", "Contato via LinkedIn", "Contato via site", "https://br.linkedin.com/company/nutrienbr/jobs"],
        ["Nordeste Atacado", 5, "Agrofel Graos e Insumos", "Distribuicao de Insumos Agricolas e Graos", "Porto Alegre/RS", "https://www.agrofel.com.br", "https://br.linkedin.com/company/agrofel", "Contato via LinkedIn", "Contato via Gupy", "https://agrofel.gupy.io/"],
        ["Nordeste Atacado", 6, "Coplacana - Cooperativa", "Cooperativa Agricola / Insumos e Varejo Rural", "Piracicaba/SP", "https://www.coplacana.com.br", "https://br.linkedin.com/company/coplacana-orgulho-do-agro", "Contato via LinkedIn", "Contato via site", "https://www.linkedin.com/company/coplacana-orgulho-do-agro/jobs"],
        ["Nordeste Atacado", 7, "Grupo Sinagro (Sinova)", "Distribuicao de Insumos Agricolas", "Primavera do Leste/MT", "https://www.sinagro.com.br", "https://www.linkedin.com/company/gruposinagro", "Contato via LinkedIn", "Contato via site", "N/D"],
        ["Nordeste Atacado", 8, "Synagro Comercial Agricola", "Distribuicao de Insumos Agricolas", "Luis Eduardo Magalhaes/BA", "N/D", "https://br.linkedin.com/company/synagro-comercial-agricola-sa", "Contato via LinkedIn", "Contato via LinkedIn", "N/D"],
        ["Nordeste Atacado", 9, "Agroquima", "Distribuicao de Insumos Agricolas", "Goiania/GO", "https://www.agroquima.com.br", "https://br.linkedin.com/company/agroquimaoficial", "Contato via LinkedIn", "Contato via site", "N/D"],
        ["Nordeste Atacado", 10, "Grupo Agrosema", "Distribuicao de Insumos Agricolas", "Brasil (multiestados)", "N/D", "https://br.linkedin.com/company/grupo-agrosema", "Contato via LinkedIn", "Contato via LinkedIn", "N/D"],
    ],

    "7. Grupo Mateus - Supermercados": [
        ["Grupo Mateus", 1, "Grupo Carrefour Brasil", "Varejo Alimentar / Hipermercados", "Barueri/SP", "https://www.grupocarrefourbrasil.com.br", "https://br.linkedin.com/company/grupocarrefourbrasil", "Catia Porto (VP de Recursos Humanos / CHRO)", "Contato via LinkedIn", "https://grupocarrefourbrasil.gupy.io/"],
        ["Grupo Mateus", 2, "Assai Atacadista", "Atacarejo", "Sao Paulo/SP", "https://www.assai.com.br", "https://br.linkedin.com/company/assaiatacadista", "Sandra Vicari (VP Gestao de Gente e Sustentabilidade)", "Contato via LinkedIn", "https://assai.gupy.io/"],
        ["Grupo Mateus", 3, "Supermercados BH", "Varejo Alimentar / Supermercados", "Belo Horizonte/MG", "https://www.supermercadosbh.com.br", "https://br.linkedin.com/company/supermercados-bh", "Elizene Martins (Gerente de RH)", "Contato via LinkedIn", "https://vagas.supermercadosbh.com.br/"],
        ["Grupo Mateus", 4, "GPA (Grupo Pao de Acucar)", "Varejo Alimentar / Supermercados", "Sao Paulo/SP", "https://www.gpabr.com", "https://br.linkedin.com/company/gpabr", "Mirella Basolli Gomiero (Dir. Exec. TI, RH e Sustentabilidade)", "Contato via LinkedIn", "https://www.gpabr.com/euescolhogpa/"],
        ["Grupo Mateus", 5, "Grupo Muffato", "Varejo Alimentar / Hipermercados", "Cascavel/PR", "https://www.supermuffato.com.br", "https://br.linkedin.com/company/grupomuffato", "Tais Ramlow da Silva Garcia (Diretora de RH)", "Contato via LinkedIn", "https://muffato.jobs.recrut.ai/"],
        ["Grupo Mateus", 6, "Grupo Pereira (Comper / Fort Atacadista)", "Varejo Alimentar / Atacarejo", "Campo Grande/MS", "https://www.grpereira.com.br", "https://br.linkedin.com/company/grupo-pereira", "Paulos Nogueira (Dir. de RH, Com. Interna e ESG)", "Contato via LinkedIn", "https://grupopereira.gupy.io/"],
        ["Grupo Mateus", 7, "Mart Minas & Dom Atacadista", "Atacado e Varejo", "Belo Horizonte/MG", "https://www.martminas.com.br", "https://br.linkedin.com/company/martminasatacadoevarejo", "Luciana Rodrigues (Gerente de Pessoas e Resultado)", "talentos203@martminas.com.br", "https://martminasedom.jobs.recrut.ai/"],
        ["Grupo Mateus", 8, "Cencosud Brasil (GBarbosa/Bretas/Prezunic)", "Varejo Alimentar / Supermercados", "Sao Paulo/SP", "https://www.cencosudbrasil.com.br", "https://br.linkedin.com/company/cencosud-brasil", "Nao identificado publicamente", "Contato via LinkedIn", "https://cencosudbrasil.gupy.io/"],
        ["Grupo Mateus", 9, "Grupo Koch", "Varejo Alimentar / Supermercados", "Gaspar/SC", "https://www.superkoch.com.br", "https://br.linkedin.com/company/koch-hipermercado-ltda", "Alisson Santos (Gerente de RH)", "Contato via LinkedIn", "https://grupokoch.jobs.recrut.ai/"],
        ["Grupo Mateus", 10, "Hiperideal", "Varejo Alimentar / Supermercados", "Salvador/BA", "https://www.hiperideal.com.br", "https://br.linkedin.com/company/hiperideal-supermercado", "Nao identificado publicamente", "Contato via LinkedIn", "https://hiperideal.jobs.recrut.ai/"],
    ],

    "8. Atakarejo - Atacarejo": [
        ["Atakarejo", 1, "Assai Atacadista", "Atacarejo", "Sao Paulo/SP", "https://www.assai.com.br", "https://br.linkedin.com/company/assaiatacadista", "Sandra Vicari (VP Gestao de Gente e Sustentabilidade)", "Contato via LinkedIn", "https://assai.gupy.io/"],
        ["Atakarejo", 2, "Atacadao (Grupo Carrefour)", "Atacarejo", "Barueri/SP", "https://www.atacadao.com.br", "https://br.linkedin.com/company/atacadao", "Catia Porto (VP de RH - Grupo Carrefour)", "Contato via LinkedIn", "https://grupocarrefourbrasil.gupy.io/"],
        ["Atakarejo", 3, "Tenda Atacado", "Atacarejo", "Sao Paulo/SP", "https://www.tendaatacado.com.br", "https://br.linkedin.com/company/tendaatacadooficial", "Cristina Caresia (Diretora de RH)", "Contato via LinkedIn", "https://tendaatacado.pandape.com.br/"],
        ["Atakarejo", 4, "Novo Atacarejo", "Atacarejo", "Recife/PE", "https://novoatacarejo.com", "https://br.linkedin.com/company/novoatacarejo", "Nao identificado publicamente", "Contato via LinkedIn", "https://novoatacarejo.jobs.recrut.ai/"],
        ["Atakarejo", 5, "Mart Minas & Dom Atacadista", "Atacado e Varejo", "Belo Horizonte/MG", "https://www.martminas.com.br", "https://br.linkedin.com/company/martminasatacadoevarejo", "Luciana Rodrigues (Gerente de Pessoas e Resultado)", "talentos203@martminas.com.br", "https://martminasedom.jobs.recrut.ai/"],
        ["Atakarejo", 6, "Fort Atacadista (Grupo Pereira)", "Atacarejo", "Campo Grande/MS", "https://www.fortatacadista.com.br", "https://br.linkedin.com/company/grupo-pereira", "Paulos Nogueira (Dir. RH - Grupo Pereira)", "Contato via LinkedIn", "https://grupopereira.gupy.io/"],
        ["Atakarejo", 7, "Atacadao Dia a Dia", "Atacarejo", "Brasilia/DF", "https://atacadaodiaadia.com.br", "https://br.linkedin.com/company/atacadaodiaadia", "Nao identificado publicamente", "rh.vagas@atacadaodiaadia.com.br", "https://atacadaodiaadia.com.br/trabalhe-conosco/"],
        ["Atakarejo", 8, "GBarbosa (Cencosud)", "Varejo / Atacarejo", "Aracaju/SE", "https://www.gbarbosa.com.br", "https://br.linkedin.com/company/cencosud-brasil", "Nao identificado publicamente (RH Cencosud)", "Contato via LinkedIn", "https://cencosudbrasil.gupy.io/"],
        ["Atakarejo", 9, "RedeMix", "Atacado e Varejo", "Salvador/BA", "https://www.redemix.com.br", "https://br.linkedin.com/company/gruporedemix", "Nao identificado publicamente", "Contato via LinkedIn", "N/D"],
        ["Atakarejo", 10, "Grupo Mateus (Mix Mateus)", "Atacarejo / Varejo Alimentar", "Sao Luis/MA", "https://www.grupomateus.com.br", "https://br.linkedin.com/company/grupo-mateus", "Nao identificado publicamente", "Contato via LinkedIn", "https://grupomateus.gupy.io/"],
    ],

    "9. RedeMix - Supermercados": [
        ["RedeMix", 1, "Grupo Mateus (Mix Mateus)", "Varejo Alimentar / Atacarejo", "Sao Luis/MA", "https://www.grupomateus.com.br", "https://br.linkedin.com/company/grupo-mateus", "Nao identificado publicamente", "Contato via LinkedIn", "https://grupomateus.gupy.io/"],
        ["RedeMix", 2, "Hiperideal", "Supermercados", "Salvador/BA", "https://www.hiperideal.com.br", "https://br.linkedin.com/company/hiperideal-supermercado", "Nao identificado publicamente", "Contato via LinkedIn", "https://hiperideal.jobs.recrut.ai/"],
        ["RedeMix", 3, "Atakarejo", "Atacarejo", "Salvador/BA", "https://atakarejo.com.br", "N/D", "Nao identificado publicamente", "Contato via LinkedIn", "N/D"],
        ["RedeMix", 4, "GBarbosa (Cencosud)", "Supermercados / Atacarejo", "Aracaju/SE", "https://www.gbarbosa.com.br", "https://br.linkedin.com/company/cencosud-brasil", "Nao identificado publicamente (RH Cencosud)", "Contato via LinkedIn", "https://cencosudbrasil.gupy.io/"],
        ["RedeMix", 5, "Assai Atacadista", "Atacarejo", "Sao Paulo/SP", "https://www.assai.com.br", "https://br.linkedin.com/company/assaiatacadista", "Sandra Vicari (VP Gestao de Gente e Sustentabilidade)", "Contato via LinkedIn", "https://assai.gupy.io/"],
        ["RedeMix", 6, "Atacadao (Grupo Carrefour)", "Atacarejo", "Barueri/SP", "https://www.atacadao.com.br", "https://br.linkedin.com/company/atacadao", "Catia Porto (VP de RH - Grupo Carrefour)", "Contato via LinkedIn", "https://grupocarrefourbrasil.gupy.io/"],
        ["RedeMix", 7, "Supermercados BH", "Supermercados", "Belo Horizonte/MG", "https://www.supermercadosbh.com.br", "https://br.linkedin.com/company/supermercados-bh", "Elizene Martins (Gerente de RH)", "Contato via LinkedIn", "https://vagas.supermercadosbh.com.br/"],
        ["RedeMix", 8, "GPA (Pao de Acucar)", "Supermercados", "Sao Paulo/SP", "https://www.gpabr.com", "https://br.linkedin.com/company/gpabr", "Mirella Basolli Gomiero (Dir. Exec. TI, RH e Sustent.)", "Contato via LinkedIn", "https://www.gpabr.com/euescolhogpa/"],
        ["RedeMix", 9, "Grupo Muffato", "Supermercados / Hipermercados", "Cascavel/PR", "https://www.supermuffato.com.br", "https://br.linkedin.com/company/grupomuffato", "Tais Ramlow da Silva Garcia (Diretora de RH)", "Contato via LinkedIn", "https://muffato.jobs.recrut.ai/"],
        ["RedeMix", 10, "Grupo Koch", "Supermercados", "Gaspar/SC", "https://www.superkoch.com.br", "https://br.linkedin.com/company/koch-hipermercado-ltda", "Alisson Santos (Gerente de RH)", "Contato via LinkedIn", "https://grupokoch.jobs.recrut.ai/"],
    ],

    "10. Lojas Guido - Moveis": [
        ["Lojas Guido", 1, "Magazine Luiza", "Varejo / Moveis e Eletro", "Franca/SP", "https://www.magazineluiza.com.br", "https://br.linkedin.com/company/magazine-luiza", "Patricia Pugas / Wiliam Miguel Santos (Dir. Gestao Pessoas)", "Contato via LinkedIn", "https://carreiras.magazineluiza.com.br/"],
        ["Lojas Guido", 2, "Grupo Casas Bahia (Via)", "Varejo / Moveis e Eletro", "Sao Caetano do Sul/SP", "https://www.casasbahia.com.br", "https://br.linkedin.com/company/grupocasasbahia", "Andreia Nunes (Dir. Exec. Gente, Gestao e Assuntos Corp.)", "Contato via LinkedIn", "https://ri.grupocasasbahia.com.br/"],
        ["Lojas Guido", 3, "MadeiraMadeira", "E-commerce / Moveis e Decoracao", "Curitiba/PR", "https://www.madeiramadeira.com.br", "https://br.linkedin.com/company/madeiramadeira", "Ana Gabriela Curcio (CHRO)", "Contato via LinkedIn", "https://madeiramadeira.gupy.io/"],
        ["Lojas Guido", 4, "Tok&Stok", "Varejo / Moveis e Decoracao", "Sao Paulo/SP", "https://www.tokstok.com.br", "https://br.linkedin.com/company/tok&stok", "Maria Carolina Brasil Borghesi (Diretora de RH)", "Contato via LinkedIn", "https://www.tokstok.com.br/trabalhe-conosco"],
        ["Lojas Guido", 5, "Havan", "Varejo / Departamento / Moveis", "Brusque/SC", "https://www.havan.com.br", "https://www.linkedin.com/company/havanoficial", "Nao identificado publicamente", "Contato via LinkedIn", "https://portalcliente.havan.com.br/trabalhe-conosco"],
        ["Lojas Guido", 6, "Marabraz", "Varejo / Moveis", "Sao Paulo/SP", "https://www.marabraz.com.br", "https://br.linkedin.com/company/lojas-marabraz", "Nao identificado publicamente", "Contato via LinkedIn", "https://www.marabraz.com.br/central-de-atendimento/trabalhe-conosco"],
        ["Lojas Guido", 7, "Novo Mundo Moveis", "Varejo / Moveis e Eletro", "Goiania/GO", "https://www.novomundo.com.br", "https://www.linkedin.com/company/novo-mundo-sa", "Nao identificado publicamente", "Contato via LinkedIn", "https://www.novomundo.com.br/institucional/trabalhe-conosco"],
        ["Lojas Guido", 8, "Zenir Moveis e Eletros", "Varejo / Moveis e Eletro", "Fortaleza/CE", "https://www.zenirmoveis.com.br", "https://br.linkedin.com/company/zenir-m%C3%B3veis-e-eletros", "Cledia Freitas (Gestora de RH)", "Contato via LinkedIn", "https://grupozenir.gupy.io/"],
        ["Lojas Guido", 9, "Macavi", "Varejo / Moveis e Eletro", "Oros/CE", "https://www.macavi.com.br", "https://br.linkedin.com/company/macavi", "Nao identificado publicamente", "Contato via LinkedIn", "N/D"],
        ["Lojas Guido", 10, "Moveis Simonetti", "Varejo / Moveis", "Pinheiros/ES", "https://simonetti.com.br", "https://br.linkedin.com/company/m%C3%B3veissimonetti", "Nao identificado publicamente", "Contato via LinkedIn", "https://www.moveissimonetti.com.br/?ft=trabalhe+conosco"],
    ],

    "11. Alpha Co - Moda Fitness": [
        ["Alpha Co", 1, "LIVE! Activewear", "Moda Fitness / Activewear", "Jaragua do Sul/SC", "https://www.liveoficial.com.br", "https://www.linkedin.com/company/liveoficial", "Nao identificado publicamente", "Contato via LinkedIn", "N/D"],
        ["Alpha Co", 2, "Track&Field", "Moda Fitness / Wellness", "Sao Paulo/SP", "https://www.tf.com.br", "https://br.linkedin.com/company/track-&-field", "Selda Pessoa Klein (Chief People & Management Officer)", "Contato via LinkedIn", "https://www.tfco.com.br/en/home-page-inst/contact/work-with-us/"],
        ["Alpha Co", 3, "Vestem", "Moda Fitness / Activewear", "Maringa/PR", "https://vestem.com", "N/D", "Nao identificado publicamente", "Contato via site", "N/D"],
        ["Alpha Co", 4, "Alto Giro", "Moda Fitness / Activewear", "Maringa/PR", "https://altogiro.com.br", "https://br.linkedin.com/company/alto-giro", "Nao identificado publicamente", "Contato via LinkedIn", "N/D"],
        ["Alpha Co", 5, "Insider Store", "Tecnologia Textil / Activewear", "Sao Paulo/SP", "https://insiderstore.com", "https://br.linkedin.com/company/insider-store", "Nao identificado publicamente", "Contato via LinkedIn", "N/D"],
        ["Alpha Co", 6, "Labellamafia", "Moda Fitness / Streetwear", "Jaragua do Sul/SC", "https://www.labellamafia.com.br", "N/D", "Nao identificado publicamente", "Contato via site", "N/D"],
        ["Alpha Co", 7, "Kaisan", "Moda Fitness", "Sao Paulo/SP", "https://www.kaisan.com.br", "N/D", "Nao identificado publicamente", "Contato via site", "N/D"],
        ["Alpha Co", 8, "Lupo Sport", "Moda Fitness / Activewear", "Araraquara/SP", "https://www.lupo.com.br", "https://br.linkedin.com/company/lupo", "Nao identificado publicamente", "Contato via LinkedIn", "https://www.lupo.com.br/trabalhe-conosco"],
        ["Alpha Co", 9, "Colcci Fitness (Grupo AMC Textil)", "Moda Fitness / Lifestyle", "Blumenau/SC", "https://www.colcci.com.br", "https://br.linkedin.com/company/colcci", "Nao identificado publicamente", "Contato via LinkedIn", "N/D"],
        ["Alpha Co", 10, "Olympikus (Vulcabras)", "Sportswear / Activewear", "Jundiai/SP", "https://www.olympikus.com.br", "https://br.linkedin.com/company/vulcabras", "Nao identificado publicamente", "Contato via LinkedIn", "https://vulcabras.gupy.io/"],
    ],

    "12. Caffeine Army - Suplementos": [
        ["Caffeine Army", 1, "Essential Nutrition (Essentia Group)", "Suplementos Funcionais / Nutricao", "Florianopolis/SC", "https://www.essentialnutrition.com.br", "https://br.linkedin.com/company/grupoessentia", "Nina Silveira (Gerente de Recursos Humanos)", "Contato via LinkedIn", "https://essentiagroup.gupy.io/"],
        ["Caffeine Army", 2, "Puravida (Nestle Health Science)", "Nutricao Funcional / Suplementos", "Sao Paulo/SP", "https://www.puravida.com.br", "https://br.linkedin.com/company/puravida", "Nao identificado publicamente (RH Nestle Brasil)", "Contato via LinkedIn", "N/D"],
        ["Caffeine Army", 3, "DUX Nutrition Lab", "Suplementos / Nutricao Esportiva", "Sao Paulo/SP", "https://www.duxnutrition.com.br", "https://br.linkedin.com/company/dux-nutrition", "Nao identificado publicamente", "Contato via LinkedIn", "N/D"],
        ["Caffeine Army", 4, "Integralmedica / Darkness (Grupo BRG)", "Suplementos Esportivos", "Sao Paulo/SP", "https://www.integralmedica.com.br", "https://br.linkedin.com/company/integralmedica", "Silvana Passos (Coordenadora de Gente e Gestao)", "Contato via LinkedIn", "https://brg-integralmedica.gupy.io/"],
        ["Caffeine Army", 5, "Growth Supplements", "Suplementos Esportivos", "Maringa/PR", "https://www.gsuplementos.com.br", "N/D", "Nao identificado publicamente", "Contato via site", "N/D"],
        ["Caffeine Army", 6, "Bold Snacks", "Barras Proteicas / Alimentos Funcionais", "Divinopolis/MG", "https://www.boldsnacks.com.br", "https://br.linkedin.com/company/bold-snacks", "Nao identificado publicamente", "Contato via LinkedIn", "N/D"],
        ["Caffeine Army", 7, "Max Titanium", "Suplementos Esportivos", "Matao/SP", "https://www.maxtitanium.com.br", "N/D", "Nao identificado publicamente", "Contato via site", "N/D"],
        ["Caffeine Army", 8, "3 Coracoes (Cafe Funcional)", "Cafe / Alimentos Funcionais", "Eusebio/CE", "https://www.3coracoes.com.br", "https://br.linkedin.com/company/grupo-3coracoes", "Nao identificado publicamente", "Contato via LinkedIn", "https://3coracoes.gupy.io/"],
        ["Caffeine Army", 9, "Sanavita", "Nutricao Funcional", "Sao Paulo/SP", "https://www.sanavita.com.br", "N/D", "Nao identificado publicamente", "Contato via site", "N/D"],
        ["Caffeine Army", 10, "Vitafor", "Suplementos / Nutricao Funcional", "Vinhedo/SP", "https://www.vitafor.com.br", "N/D", "Nao identificado publicamente", "Contato via site", "N/D"],
    ],

    "13. Acai NoKilo - Franquias Acai": [
        ["Acai NoKilo", 1, "The Best Acai (Grupo The Best)", "Franquia / Acaiteria Self-Service", "Londrina/PR", "https://thebestacai.com.br", "https://br.linkedin.com/company/grupothebest", "Nao identificado publicamente", "Contato via LinkedIn", "https://grupothebest.rhgestor.com.br/"],
        ["Acai NoKilo", 2, "Oakberry Acai Bowls", "Franquia / Acai Premium", "Sao Paulo/SP", "https://www.oakberry.com", "https://br.linkedin.com/company/oakberry", "Nao identificado publicamente", "Contato via LinkedIn", "N/D"],
        ["Acai NoKilo", 3, "Acai Concept", "Franquia / Acaiteria", "Recife/PE", "https://acaiconcept.com", "N/D", "Nao identificado publicamente", "Contato via site", "N/D"],
        ["Acai NoKilo", 4, "Acai da Barra", "Franquia / Acaiteria Self-Service", "Barra Bonita/SP", "https://acaidabarra.com.br", "https://br.linkedin.com/company/acai-da-barra", "Nao identificado publicamente", "Contato via LinkedIn", "N/D"],
        ["Acai NoKilo", 5, "JAH do Acai", "Franquia / Acaiteria Self-Service", "Conselheiro Lafaiete/MG", "https://jahdoacai.com.br", "N/D", "Nao identificado publicamente", "Contato via site", "https://jahdoacai.com.br/trabalhe-conosco/"],
        ["Acai NoKilo", 6, "Universo do Acai", "Franquia / Acaiteria", "N/D", "https://www.universodoacai.com.br", "N/D", "Nao identificado publicamente", "Contato via site", "N/D"],
        ["Acai NoKilo", 7, "Tropicool", "Franquia / Acai e Superfoods", "Sao Paulo/SP", "https://franquiatropicool.com.br", "N/D", "Nao identificado publicamente", "Contato via site", "N/D"],
        ["Acai NoKilo", 8, "Acai Natu", "Franquia / Acaiteria", "N/D", "https://acainatu.com.br", "N/D", "Nao identificado publicamente", "Contato via site", "N/D"],
        ["Acai NoKilo", 9, "Acai Artesanal", "Franquia / Acaiteria", "N/D", "https://acaiartesanal.com.br", "N/D", "Nao identificado publicamente", "Contato via site", "N/D"],
        ["Acai NoKilo", 10, "Acamix", "Franquia / Acaiteria", "N/D", "https://acamix.com.br", "N/D", "Nao identificado publicamente", "Contato via site", "N/D"],
    ],
}

# Create sheets
for sheet_idx, (sheet_name, rows) in enumerate(data.items()):
    if sheet_idx == 0:
        ws = wb.active
        ws.title = sheet_name[:31]
    else:
        ws = wb.create_sheet(title=sheet_name[:31])

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap
        cell.border = thin_border

    # Write data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap
            cell.border = thin_border
            if col_idx == 1:
                cell.fill = client_fill
                cell.font = client_font

    # Set column widths
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze panes
    ws.freeze_panes = "C2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

# Also create a consolidated sheet
ws_all = wb.create_sheet(title="CONSOLIDADO", index=0)
for col_idx, header in enumerate(headers, 1):
    cell = ws_all.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = thin_border

current_row = 2
for sheet_name, rows in data.items():
    for row_data in rows:
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_all.cell(row=current_row, column=col_idx, value=value)
            cell.alignment = wrap
            cell.border = thin_border
            if col_idx == 1:
                cell.fill = client_fill
                cell.font = client_font
        current_row += 1

for col_idx, width in enumerate(col_widths, 1):
    ws_all.column_dimensions[get_column_letter(col_idx)].width = width

ws_all.freeze_panes = "C2"
ws_all.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{current_row-1}"

output_path = "/home/headless/workspace/cih/clientes/pesquisacliente/concorrentes_clientes_CIH.xlsx"
wb.save(output_path)
print(f"Planilha salva em: {output_path}")
print(f"Total de concorrentes mapeados: {current_row - 2}")
