import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
category_font = Font(bold=True, color="FFFFFF", size=12)
category_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_alignment = Alignment(wrap_text=True, vertical='top')

headers = [
    "Cliente CIH (Referencia)",
    "N.",
    "Empresa Concorrente",
    "Segmento",
    "Cidade/UF",
    "Site Oficial",
    "LinkedIn (Empresa)",
    "Gestor de RH / Diretor de Pessoas",
    "Cargo do Gestor RH",
    "LinkedIn do Gestor RH",
    "Contato RH (email/telefone)",
    "Portal Carreiras / Trabalhe Conosco",
    "Observacoes"
]

col_widths = [28, 5, 35, 28, 22, 40, 50, 35, 40, 55, 35, 50, 40]

# ============================================================
# DATA
# ============================================================

data = []

# ----- 1. BAHIANA DE MEDICINA (EBMSP) - Concorrentes em Educacao Medica -----
bahiana_competitors = [
    ["Bahiana de Medicina (EBMSP)", "1", "PUCPR - Pontificia Universidade Catolica do Parana", "Educacao Superior / Saude", "Curitiba/PR",
     "https://www.pucpr.br", "https://br.linkedin.com/school/pucpr/",
     "A confirmar (Diretoria de Gestao de Pessoas)", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://www.pucpr.br/trabalhe-conosco/", "1a colocada faculdade medicina privada RUF 2025. Reitor: Ir. Rogerio Renato Mateucci"],

    ["Bahiana de Medicina (EBMSP)", "2", "PUCRS - Pontificia Universidade Catolica do RS", "Educacao Superior / Saude", "Porto Alegre/RS",
     "https://www.pucrs.br", "https://br.linkedin.com/school/pucrs/",
     "A confirmar (Diretoria de Gestao de Pessoas - PROAF)", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://portal.pucrs.br/sou-tecnico-administrativo/", "2a colocada faculdade medicina privada RUF 2025"],

    ["Bahiana de Medicina (EBMSP)", "3", "Faculdade Israelita de Ciencias da Saude Albert Einstein", "Educacao Superior / Saude", "Sao Paulo/SP",
     "https://ensino.einstein.br", "https://br.linkedin.com/school/ensinoeinstein/",
     "A confirmar", "Diretor(a) de RH / Gente e Gestao", "", "",
     "https://atendimento.einstein.br/categoria/outros/trabalhe-conosco/", "Vinculada ao Hospital Albert Einstein. Ensino integrado a pratica hospitalar"],

    ["Bahiana de Medicina (EBMSP)", "4", "FCMSCSP - Fac. Ciencias Medicas Santa Casa de Sao Paulo", "Educacao Superior / Saude", "Sao Paulo/SP",
     "https://fcmsantacasasp.edu.br", "https://br.linkedin.com/school/fcmscsp/",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://fcmsantacasasp.edu.br/vemprasanta/", "Mais de 60 anos de historia em formacao medica classica e humanista"],

    ["Bahiana de Medicina (EBMSP)", "5", "UNIFOR - Universidade de Fortaleza", "Educacao Superior / Saude", "Fortaleza/CE",
     "https://unifor.br", "https://br.linkedin.com/school/uniforoficial/",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://unifor.br/trabalhe-conosco", "Fundacao Edson Queiroz. Nota maxima MEC em Medicina"],

    ["Bahiana de Medicina (EBMSP)", "6", "Afya Educacional", "Educacao Superior / Saude (multi-campus)", "Nova Lima/MG (sede)",
     "https://www.afya.com.br", "https://br.linkedin.com/company/afyamedicina",
     "A confirmar", "Diretor(a) de Gente e Gestao", "", "",
     "https://afya.gupy.io/", "Maior grupo de educacao medica do Brasil. ~10 mil colaboradores"],

    ["Bahiana de Medicina (EBMSP)", "7", "Anima Educacao (Anhembi Morumbi / Inspirali)", "Educacao Superior / Saude", "Sao Paulo/SP (sede)",
     "https://www.animaeducacao.com.br", "https://br.linkedin.com/company/animaeducacao",
     "A confirmar", "Diretor(a) de Gente e Gestao", "", "",
     "https://trabalheconosco.vagas.com.br/animaeducacao", "Um dos maiores grupos educacionais do Brasil. Anhembi Morumbi e Inspirali (medicina)"],

    ["Bahiana de Medicina (EBMSP)", "8", "UniCEUB - Centro Universitario de Brasilia", "Educacao Superior / Saude", "Brasilia/DF",
     "https://www.uniceub.br", "https://br.linkedin.com/school/uniceub-centro-universitario-de-brasilia/",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://www.uniceub.br/trabalhe-conosco", "Chanceler: Getulio Americo Moreira Lopes"],

    ["Bahiana de Medicina (EBMSP)", "9", "UniFTC / Faculdade Zarns Salvador", "Educacao Superior / Saude", "Salvador/BA",
     "https://www.uniftc.edu.br", "https://br.linkedin.com/company/uniftc/",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://www.uniftc.edu.br/trabalhe-conosco", "Clariens Educacao. CEO: Thiago Sayao. Reitor Zarns: Carlos Alberto Ramos"],

    ["Bahiana de Medicina (EBMSP)", "10", "Ser Educacional (UNINASSAU)", "Educacao Superior / Saude (multi-campus)", "Recife/PE (sede)",
     "https://www.sereducacional.com", "https://br.linkedin.com/company/sereducacional",
     "A confirmar", "Diretor(a) de Gente e Gestao", "", "",
     "https://sereducacional.gupy.io/", "Um dos maiores grupos educacionais. +200 mil alunos. Cursos de medicina em expansao"],
]

# ----- 2. SABIN DIAGNOSTICO (GRUPO SABIN) - Concorrentes Diagnostico -----
sabin_competitors = [
    ["Sabin Diagnostico (Grupo Sabin)", "1", "DASA - Diagnosticos da America", "Diagnostico Medico / Saude Integrada", "Sao Paulo/SP",
     "https://www.dasa.com.br", "https://br.linkedin.com/company/dasa",
     "Fabio Lacerda", "VP de Gente, Gestao e Cultura (Unidade Diagnosticos)", "https://www.linkedin.com/in/fabioslacerda/", "",
     "https://dasa.gupy.io/", "Maior empresa de diagnostico da America Latina. Marcas: Delboni, Lavoisier, Sergio Franco, Bronstein, Alvaro"],

    ["Sabin Diagnostico (Grupo Sabin)", "2", "Grupo Fleury", "Diagnostico Medico", "Sao Paulo/SP",
     "https://www.grupofleury.com.br", "https://br.linkedin.com/company/grupo-fleury",
     "A confirmar", "Diretor(a) Executivo(a) de Pessoas", "", "",
     "https://carreiras.grupofleury.com.br/", "CEO: Jeane Tsutsui. +90 anos de operacao. Referencia em medicina diagnostica premium"],

    ["Sabin Diagnostico (Grupo Sabin)", "3", "Hermes Pardini (Grupo Pardini)", "Diagnostico Medico / Analises Clinicas", "Belo Horizonte/MG",
     "https://www.hermespardini.com.br", "https://br.linkedin.com/company/laborat%C3%B3rio-hermes-pardini",
     "A confirmar", "Diretor(a) de Gente e Gestao", "", "",
     "https://trabalheconosco.vagas.com.br/grupopardini", "Fundado ha mais de 60 anos. Agora parte do Grupo Fleury"],

    ["Sabin Diagnostico (Grupo Sabin)", "4", "DB Diagnosticos (Diagnosticos do Brasil)", "Diagnostico Medico (Apoio Laboratorial)", "Curitiba/PR",
     "https://www.diagnosticosdobrasil.com.br", "https://br.linkedin.com/company/dbdiagnosticos",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://diagnosticosdobrasil.com.br/trabalhe-conosco", "1o laboratorio 100% de apoio no Brasil. +35 milhoes exames/mes"],

    ["Sabin Diagnostico (Grupo Sabin)", "5", "Alta Diagnosticos", "Diagnostico Medico / Imagem", "Sao Paulo/SP",
     "https://altadiagnosticos.com.br", "https://br.linkedin.com/company/altadiagnosticos",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://altadiagnosticos.com.br/trabalhe-conosco", "Laboratorio premium com unidades em SP e RJ"],

    ["Sabin Diagnostico (Grupo Sabin)", "6", "Cura Imagem e Diagnostico", "Diagnostico Medico / Imagem", "Sao Paulo/SP",
     "https://www.cura.com.br", "https://br.linkedin.com/company/curamedicinadiagnostica",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "", "40 anos de historia. +30 unidades em SP. 1.700 colaboradores"],

    ["Sabin Diagnostico (Grupo Sabin)", "7", "Laboratorio Richet", "Diagnostico Medico / Analises Clinicas", "Rio de Janeiro/RJ",
     "https://www.richet.com.br", "https://br.linkedin.com/company/richet-medicina-diagnostica",
     "A confirmar", "Diretor(a) de RH", "", "",
     "", "Fundado em 1947. 6 unidades no RJ. Exames exclusivos e personalizados"],

    ["Sabin Diagnostico (Grupo Sabin)", "8", "Laboratorio Sergio Franco (Grupo DASA)", "Diagnostico Medico / Analises Clinicas", "Rio de Janeiro/RJ",
     "https://www.sergiofranco.com.br", "https://br.linkedin.com/company/laboratorio-sergio-franco",
     "Vinculado a DASA - Fabio Lacerda (VP Gente)", "VP de Gente, Gestao e Cultura (DASA)", "https://www.linkedin.com/in/fabioslacerda/", "",
     "https://dasa.gupy.io/", "Parte do Grupo DASA. Forte presenca no RJ"],

    ["Sabin Diagnostico (Grupo Sabin)", "9", "Delboni Auriemo (Grupo DASA)", "Diagnostico Medico / Analises Clinicas", "Sao Paulo/SP",
     "https://www.dfrfranco.com.br", "https://br.linkedin.com/company/delboniauriemo",
     "Vinculado a DASA - Fabio Lacerda (VP Gente)", "VP de Gente, Gestao e Cultura (DASA)", "https://www.linkedin.com/in/fabioslacerda/", "",
     "https://dasa.gupy.io/", "Parte do Grupo DASA. +50 anos em medicina diagnostica em SP"],

    ["Sabin Diagnostico (Grupo Sabin)", "10", "Laboratorio Bronstein (Grupo DASA)", "Diagnostico Medico / Analises Clinicas", "Rio de Janeiro/RJ",
     "https://bronstein.com.br", "https://br.linkedin.com/company/bronstein-medicina-diagnostica",
     "Vinculado a DASA - Fabio Lacerda (VP Gente)", "VP de Gente, Gestao e Cultura (DASA)", "https://www.linkedin.com/in/fabioslacerda/", "",
     "https://dasa.gupy.io/", "Parte do Grupo DASA. Presenca no estado do RJ"],
]

# ----- 3. HOSPITAL PORTUGUES DA BAHIA - Concorrentes Hospitais -----
hp_competitors = [
    ["Hospital Portugues da Bahia", "1", "Hospital Israelita Albert Einstein", "Saude / Hospitalar", "Sao Paulo/SP",
     "https://www.einstein.br", "https://br.linkedin.com/company/hospitalalberteinstein",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://atendimento.einstein.br/categoria/outros/trabalhe-conosco/", "Melhor hospital do Brasil (Newsweek). Diretor Geral: Henrique Neves"],

    ["Hospital Portugues da Bahia", "2", "Hospital Sirio-Libanes", "Saude / Hospitalar", "Sao Paulo/SP",
     "https://hospitalsiriolibanes.org.br", "https://br.linkedin.com/company/hospitalsiriolibanes",
     "Renato Bezerra da Silva (Gerente Sr. RH)", "Gerente Senior de RH - Business Partner", "https://www.linkedin.com/in/renatobezerra-rh/", "",
     "https://vagas.hsl.org.br/", "Top hospital Brasil. Reconhecido internacionalmente"],

    ["Hospital Portugues da Bahia", "3", "Rede D'Or Sao Luiz", "Saude / Hospitalar (rede)", "Rio de Janeiro/RJ (sede)",
     "https://www.rededorsaoluiz.com.br", "https://www.linkedin.com/company/rededor",
     "Rodrigo Pacca", "Diretor Executivo de Recursos Humanos", "https://www.linkedin.com/in/rodrigopacca/", "",
     "https://trabalheconosco.vagas.com.br/rededorsaoluiz", "Maior rede hospitalar privada do Brasil"],

    ["Hospital Portugues da Bahia", "4", "Hospital Alemao Oswaldo Cruz", "Saude / Hospitalar", "Sao Paulo/SP",
     "https://www.hospitaloswaldocruz.org.br", "https://br.linkedin.com/company/hospitalalemaooc",
     "Maria Carolina Lourenco Gomes", "Diretora Executiva de Pessoas, Sustentabilidade e Resp. Social", "https://www.linkedin.com/in/maria-carolina-gomes-carol-gomes-8416a81b/", "",
     "https://hospitaloswaldocruz.gupy.io/", "Hospital filantropico de referencia. Selo Igualdade Racial"],

    ["Hospital Portugues da Bahia", "5", "BP - Beneficencia Portuguesa de Sao Paulo", "Saude / Hospitalar", "Sao Paulo/SP",
     "https://www.bp.org.br", "https://br.linkedin.com/company/bporgbr",
     "A confirmar (Diretoria de Pessoas e Experiencia do Cliente)", "Diretor(a) de Pessoas, Experiencia, Marketing, ESG", "", "",
     "https://vemserbp.gupy.io/", "Hub de saude. 7.500 colaboradores, 4.500 medicos. 70% mulheres na equipe"],

    ["Hospital Portugues da Bahia", "6", "Hospital Moinhos de Vento", "Saude / Hospitalar", "Porto Alegre/RS",
     "https://www.hospitalmoinhos.org.br", "https://br.linkedin.com/company/hospital-moinhos-de-vento",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://hospitalmoinhos.gupy.io/", "Referencia no Sul. CEO: Mohamed Parrini Mutlaq"],

    ["Hospital Portugues da Bahia", "7", "Rede Mater Dei de Saude", "Saude / Hospitalar (rede)", "Belo Horizonte/MG",
     "https://www.materdei.com.br", "https://br.linkedin.com/company/redematerdei",
     "Grazielly Souza (a confirmar cargo exato)", "Profissional de RH Senior / Gestao de Pessoas", "https://www.linkedin.com/in/graziellysouza/", "",
     "https://app.jobconvo.com/pt-br/careers/hospital-mater-dei/", "Fundada em 1980. Gustavo Paiva: Diretor Geral"],

    ["Hospital Portugues da Bahia", "8", "A.C.Camargo Cancer Center", "Saude / Hospitalar / Oncologia", "Sao Paulo/SP",
     "https://accamargo.org.br", "https://br.linkedin.com/company/accamargo",
     "A confirmar", "Diretor(a) de Gente e Gestao", "", "",
     "https://accamargo.org.br/form/contato", "Referencia em oncologia. CEO: Victor Piana"],

    ["Hospital Portugues da Bahia", "9", "HCor - Hospital do Coracao", "Saude / Hospitalar / Cardiologia", "Sao Paulo/SP",
     "https://www.hcor.com.br", "https://br.linkedin.com/company/hcor",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://www.hcor.com.br/institucional/trabalhe-conosco/", "Referencia em cardiologia. Ranking Newsweek 2026"],

    ["Hospital Portugues da Bahia", "10", "Hospital Nove de Julho", "Saude / Hospitalar", "Sao Paulo/SP",
     "https://www.h9j.com.br", "https://br.linkedin.com/company/hospitalnovejulho",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://www.h9j.com.br/trabalhe-conosco", "Parte do grupo Rede D'Or. Ranking Newsweek 2026"],
]

# ----- 4. SANTA CASA DA BAHIA - Concorrentes Hospitais (overlap com HP) -----
sc_competitors = [
    ["Santa Casa da Bahia", "1", "Rede D'Or Sao Luiz", "Saude / Hospitalar (rede)", "Rio de Janeiro/RJ (sede)",
     "https://www.rededorsaoluiz.com.br", "https://www.linkedin.com/company/rededor",
     "Rodrigo Pacca", "Diretor Executivo de Recursos Humanos", "https://www.linkedin.com/in/rodrigopacca/", "",
     "https://trabalheconosco.vagas.com.br/rededorsaoluiz", "Tambem concorrente Hospital Portugues. Maior rede hospitalar privada"],

    ["Santa Casa da Bahia", "2", "Hospital Sirio-Libanes", "Saude / Hospitalar", "Sao Paulo/SP",
     "https://hospitalsiriolibanes.org.br", "https://br.linkedin.com/company/hospitalsiriolibanes",
     "Renato Bezerra da Silva (Gerente Sr. RH)", "Gerente Senior de RH - Business Partner", "https://www.linkedin.com/in/renatobezerra-rh/", "",
     "https://vagas.hsl.org.br/", "Tambem concorrente Hospital Portugues"],

    ["Santa Casa da Bahia", "3", "Hospital Alemao Oswaldo Cruz", "Saude / Hospitalar", "Sao Paulo/SP",
     "https://www.hospitaloswaldocruz.org.br", "https://br.linkedin.com/company/hospitalalemaooc",
     "Maria Carolina Lourenco Gomes", "Diretora Executiva de Pessoas, Sustentabilidade e Resp. Social", "https://www.linkedin.com/in/maria-carolina-gomes-carol-gomes-8416a81b/", "",
     "https://hospitaloswaldocruz.gupy.io/", "Tambem concorrente Hospital Portugues. Hospital filantropico"],

    ["Santa Casa da Bahia", "4", "BP - Beneficencia Portuguesa de Sao Paulo", "Saude / Hospitalar (filantropico)", "Sao Paulo/SP",
     "https://www.bp.org.br", "https://br.linkedin.com/company/bporgbr",
     "A confirmar (Diretoria de Pessoas e Experiencia)", "Diretor(a) de Pessoas, Experiencia, Marketing, ESG", "", "",
     "https://vemserbp.gupy.io/", "Tambem concorrente Hospital Portugues"],

    ["Santa Casa da Bahia", "5", "Santa Casa de Misericordia de Sao Paulo (ISCMSP)", "Saude / Hospitalar (filantropico)", "Sao Paulo/SP",
     "https://www.santacasasp.org.br", "https://br.linkedin.com/company/abordesantacasa",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://www.santacasasp.org.br/portal/trabalhe-conosco", "Maior hospital filantropico da America Latina"],

    ["Santa Casa da Bahia", "6", "Hospital Moinhos de Vento", "Saude / Hospitalar", "Porto Alegre/RS",
     "https://www.hospitalmoinhos.org.br", "https://br.linkedin.com/company/hospital-moinhos-de-vento",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://hospitalmoinhos.gupy.io/", "Tambem concorrente Hospital Portugues"],

    ["Santa Casa da Bahia", "7", "Hospital Israelita Albert Einstein", "Saude / Hospitalar", "Sao Paulo/SP",
     "https://www.einstein.br", "https://br.linkedin.com/company/hospitalalberteinstein",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://atendimento.einstein.br/categoria/outros/trabalhe-conosco/", "Tambem concorrente Hospital Portugues"],

    ["Santa Casa da Bahia", "8", "Prevent Senior", "Saude / Hospitalar / Operadora", "Sao Paulo/SP",
     "https://www.preventsenior.com.br", "https://br.linkedin.com/company/preventsenior",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://www.preventsenior.com.br/trabalhe-conosco", "Operadora + rede propria (Sancta Maggiore). ~9.000 colaboradores"],

    ["Santa Casa da Bahia", "9", "Rede Mater Dei de Saude", "Saude / Hospitalar (rede)", "Belo Horizonte/MG",
     "https://www.materdei.com.br", "https://br.linkedin.com/company/redematerdei",
     "Grazielly Souza (a confirmar cargo exato)", "Profissional de RH Senior / Gestao de Pessoas", "https://www.linkedin.com/in/graziellysouza/", "",
     "https://app.jobconvo.com/pt-br/careers/hospital-mater-dei/", "Tambem concorrente Hospital Portugues"],

    ["Santa Casa da Bahia", "10", "Hospital Pequeno Principe", "Saude / Hospitalar (filantropico/pediatrico)", "Curitiba/PR",
     "https://pequenoprincipe.org.br", "https://br.linkedin.com/company/hospital-pequeno-principe",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://pequenoprincipe.org.br/trabalhe-conosco", "Maior hospital pediatrico do pais. Ranking Newsweek 2026"],
]

# ----- 5. UNIMED SALVADOR/BAHIA - Concorrentes Planos de Saude -----
unimed_competitors = [
    ["Unimed Salvador/Bahia", "1", "Hapvida NotreDame Intermedica", "Saude / Operadora + Rede Propria", "Fortaleza/CE (sede)",
     "https://www.hapvida.com.br", "https://www.linkedin.com/company/hapvidasaude",
     "Majo Campos (ate 2022) / Verificar atual", "VP de Gente, Gestao de Pessoas e Diversidade", "", "",
     "https://hapvidandi.gupy.io/", "Lider do setor: 8,8 milhoes de usuarios. Rede propria de hospitais e clinicas"],

    ["Unimed Salvador/Bahia", "2", "Bradesco Saude (Grupo Bradesco Seguros)", "Saude / Seguro Saude", "Sao Paulo/SP",
     "https://www.bradescoseguros.com.br", "https://br.linkedin.com/company/bradesco-seguros",
     "Valdirene Soares Secato", "Diretora de RH, Universeg, Sustentabilidade e Ouvidoria", "https://www.linkedin.com/in/valdirene-soares-secato/", "",
     "https://www.estagiobradescoseguros.com/", "3,8 milhoes de usuarios. Grupo Bradesco Seguros"],

    ["Unimed Salvador/Bahia", "3", "Amil (UnitedHealth Group Brasil)", "Saude / Operadora", "Sao Paulo/SP",
     "https://www.amil.com.br", "https://br.linkedin.com/company/amil",
     "Filipe Abitan", "Diretor de RH", "https://br.linkedin.com/in/filipe-abitan-0312887", "",
     "https://amil.gupy.io/", "3,1 milhoes de usuarios. Parte do UnitedHealth Group"],

    ["Unimed Salvador/Bahia", "4", "SulAmerica Saude", "Saude / Seguro Saude", "Rio de Janeiro/RJ",
     "https://portal.sulamericaseguros.com.br", "https://br.linkedin.com/company/sulamerica",
     "Renato Luzzi", "Diretor de Pessoas (CHRO), Mobilidade e Logistica", "", "",
     "https://sulamerica.gupy.io/", "2,7 milhoes de usuarios. Adquirida pela Rede D'Or"],

    ["Unimed Salvador/Bahia", "5", "NotreDame Intermedica (GNDI)", "Saude / Operadora + Rede Propria", "Sao Paulo/SP",
     "https://www.gndi.com.br", "https://br.linkedin.com/company/gndi",
     "Eliana Vieira", "Diretora Executiva de Recursos Humanos", "https://www.linkedin.com/in/eliana-vieira-55b60728/", "",
     "https://www2.gndi.com.br/grupo/trabalhe-conosco", "Parte do grupo Hapvida. Rede propria ampla"],

    ["Unimed Salvador/Bahia", "6", "Porto Seguro Saude", "Saude / Seguro Saude", "Sao Paulo/SP",
     "https://www.portoseguro.com.br/seguro-saude", "https://br.linkedin.com/company/porto",
     "A confirmar (Diretoria de Pessoas e Cultura)", "Diretor(a) de Pessoas e Cultura", "", "",
     "https://porto.gupy.io/", "Parte do grupo Porto. Superintendente Gente: Anderson Bertolotti Valadares"],

    ["Unimed Salvador/Bahia", "7", "Prevent Senior", "Saude / Operadora + Rede Propria", "Sao Paulo/SP",
     "https://www.preventsenior.com.br", "https://br.linkedin.com/company/preventsenior",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://www.preventsenior.com.br/trabalhe-conosco", "Especializada em publico 60+. Rede propria de hospitais"],

    ["Unimed Salvador/Bahia", "8", "DASA (SulAmerica integrada)", "Saude / Seguro Saude + Diagnostico", "Sao Paulo/SP",
     "https://www.dasa.com.br", "https://br.linkedin.com/company/dasa",
     "Fabio Lacerda", "VP de Gente, Gestao e Cultura", "https://www.linkedin.com/in/fabioslacerda/", "",
     "https://dasa.gupy.io/", "Rede integrada de saude. Integracao com SulAmerica Saude"],

    ["Unimed Salvador/Bahia", "9", "Cassi (Caixa de Assistencia dos Func. Banco do Brasil)", "Saude / Autogestao", "Brasilia/DF",
     "https://www.cassi.com.br", "https://br.linkedin.com/company/cassi",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://www.cassi.com.br/trabalhe-conosco", "Uma das maiores autogestoes do Brasil"],

    ["Unimed Salvador/Bahia", "10", "Care Plus (Grupo Bupa)", "Saude / Seguro Saude (premium)", "Sao Paulo/SP",
     "https://www.careplus.com.br", "https://br.linkedin.com/company/careplus",
     "A confirmar", "Diretor(a) de Gestao de Pessoas", "", "",
     "https://careplus.gupy.io/", "Plano premium. Parte do grupo internacional Bupa"],
]

all_data = bahiana_competitors + sabin_competitors + hp_competitors + sc_competitors + unimed_competitors

# Create worksheet
ws = wb.active
ws.title = "Concorrentes - Pesquisa Completa"

# Write headers
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    cell.border = thin_border

# Set column widths
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Write data
current_client = ""
row_idx = 2
for row_data in all_data:
    # Add category separator when client changes
    if row_data[0] != current_client:
        if current_client != "":
            row_idx += 1  # blank row between sections
        current_client = row_data[0]

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap_alignment
        cell.border = thin_border

    # Alternate row coloring within each section
    if int(row_data[1]) % 2 == 0:
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = alt_fill

    row_idx += 1

# Freeze panes
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx-1}"

# Set row height for header
ws.row_dimensions[1].height = 40

# Save
output_path = "/home/headless/workspace/cih/clientes/pesquisacliente/concorrentes_5empresas_pesquisa.xlsx"
wb.save(output_path)
print(f"Arquivo salvo em: {output_path}")
