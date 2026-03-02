#!/usr/bin/env python3
"""
Reconstrói o PDI do Elias Fernandes com melhorias baseadas nos aprendizados do PDI da Ayla:
1. Acompanhamento Semanal expandido de 13 para 26 semanas (ciclo completo de 6 meses)
2. Resultado e Evolução com 6 meses em vez de 3
3. Checkpoints para todos os 6 meses (era só 3)
4. Semáforo com 6 meses
5. Nova aba Guia 1-1 adaptada para contexto Elias ↔ Caio
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Cores CIH ──
TEAL = '0097A7'
DARK_TEAL = '00796B'
WHITE = 'FFFFFF'
LIGHT_TEAL = 'E0F2F1'
LIGHT_GRAY = 'F5F5F5'
LIGHT_GREEN = 'E8F5E9'
LIGHT_YELLOW = 'FFFDE7'
LIGHT_RED = 'FFEBEE'
LIGHT_ORANGE = 'FFF3E0'
LIGHT_BLUE = 'E3F2FD'
PALE_YELLOW = 'FFF8E1'
LIGHT_PURPLE = 'F3E5F5'
GRAY_TEXT = '757575'
GREEN_BG = '43A047'
YELLOW_BG = 'FFC107'
RED_BG = 'E53935'
DARK_GRAY = '424242'
ORANGE_TEXT = 'F57C00'
RED_TEXT = 'E53935'

# ── Estilos ──
def hfill(c):
    return PatternFill(start_color=c, end_color=c, fill_type='solid')

header_fill = hfill(TEAL)
header_font = Font(bold=True, size=11, color=WHITE)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
title_font = Font(bold=True, size=14, color=TEAL)
title_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
subtitle_font = Font(bold=True, size=12, color=DARK_TEAL)
label_fill = hfill(LIGHT_TEAL)
label_font = Font(bold=True, size=10)
data_font = Font(size=10)
data_align_l = Alignment(horizontal='left', vertical='center', wrap_text=True)
data_align_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
ctx_fill = hfill(LIGHT_GRAY)
ctx_font = Font(bold=True, size=10)
ctx_data_align = Alignment(vertical='top', wrap_text=True)
src_fill = hfill(LIGHT_GRAY)
bold_font = Font(bold=True, size=10)
section_fill = hfill(LIGHT_TEAL)
section_font = Font(bold=True, size=11, color=DARK_TEAL)
tip_font = Font(size=10, italic=True, color=GRAY_TEXT)
example_font = Font(size=10, color='37474F')

thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

def set_header_row(ws, row, cols):
    for ci, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = thin_border

def set_data_cell(ws, row, col, val, font=None, fill=None, align=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font or data_font
    if fill:
        c.fill = fill
    c.alignment = align or data_align_l
    c.border = thin_border
    return c

def cell(ws, row, col, val, font=None, fill=None, align=None, merge_end=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font or data_font
    if fill:
        c.fill = fill
    c.alignment = align or data_align_l
    c.border = thin_border
    if merge_end:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end)
    return c

def level_fill(level):
    if level <= 1: return hfill(LIGHT_RED)
    if level == 2: return hfill(LIGHT_YELLOW)
    return hfill(LIGHT_GREEN)

def type_fill(tipo):
    if '70%' in tipo: return hfill(LIGHT_BLUE)
    if '20%' in tipo: return hfill(PALE_YELLOW)
    return hfill(LIGHT_PURPLE)

def section_title(ws, row, text, end_col=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = section_fill
    c.font = section_font
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = thin_border
    ws.row_dimensions[row].height = 28
    return row + 1

# ═══════════════════════════════════════════════
# DADOS DO ELIAS
# ═══════════════════════════════════════════════
COMPETENCIAS = [
    (1, 'Escuta Ativa e Comunicação Adaptada ao Perfil da Equipe', 2, 3, '30%'),
    (2, 'Transição de Executor para Líder (Soltar a Execução)', 1, 3, '30%'),
    (3, 'Gestão Técnica e Formação da Equipe', 1, 3, '25%'),
    (4, 'Inteligência Emocional e Gestão de Energia', 2, 3, '15%'),
]

ACOES = [
    (1,'Escuta Ativa e Comunicação','70% Experiência','Reuniões individuais mensais',
     '20-30 min com cada um dos 7 membros. Roteiro: Como você está? O que está difícil? O que posso fazer? Mais escutar, menos falar.',
     'Mar/2026','Ago/2026','Mensal','7 reuniões/mês com anotações','Elias','DISC + Devolutiva'),
    (2,'Escuta Ativa e Comunicação','70% Experiência','Conduzir reunião diária de produção',
     'Assumir a condução das reuniões de produção com o time mecânico. Descrever ordens de serviço em aberto, definir prioridades, dar direcionamento de tarefas. Caio observa e depois dá feedback.',
     'Mar/2026','Ago/2026','Diário','Reunião conduzida por Elias 5x/semana','Elias','Kick-off (Caio)'),
    (3,'Escuta Ativa e Comunicação','70% Experiência','Mapa de perfis do time',
     'Com apoio de Eleine, identificar perfil DISC de cada membro. Criar cola com estilo de comunicação ideal por pessoa.',
     'Mar/2026','Abr/2026','Única','Mapa completo dos 7 perfis','Elias + Eleine','DISC'),
    (4,'Escuta Ativa e Comunicação','70% Experiência','Regra dos 10 segundos',
     "Após fazer pergunta em reunião, contar até 10 mentalmente antes de responder. Dar espaço para o time. Especialmente importante porque Elias é 'de poucas palavras' e tende a encurtar diálogos.",
     'Mar/2026','Ago/2026','Diário','Auto-avaliação semanal','Elias','DISC + Kick-off'),
    (5,'Escuta Ativa e Comunicação','20% Social','Mentoria semanal com Caio',
     ' 15-20 min para discutir situações com a equipe: como eu lidaria vs. como devo lidar. Caio compartilha como ele próprio aprendeu a se comunicar.',
     'Mar/2026','Ago/2026','Semanal','4+ conversas/mês registradas','Elias + Caio','DISC + Kick-off'),
    (6,'Escuta Ativa e Comunicação','10% Formal','Estudo dos perfis DISC',
     'Ler material completo do resultado DISC (Caio imprime e entrega). Entender diferenças entre Comunicador, Executor, Analista e Planejador.',
     'Mar/2026','Mar/2026','Única','Leitura concluída + 3 insights','Elias','DISC + Devolutiva'),
    (7,'Transição Executor → Líder','70% Experiência','Sabatina semanal com o time',
     'Apresentar demanda/problema e pedir que o time traga soluções. Elias observa, questiona e orienta — NÃO executa. Foco: fazer o time pensar e resolver.',
     'Mar/2026','Ago/2026','Semanal','1 sabatina/semana com registro','Elias','Devolutiva'),
    (8,'Transição Executor → Líder','70% Experiência','Diário "Quase Meti a Mão"',
     'Registrar toda vez que sentir o impulso de pegar na execução. Anotar: situação, sentimento, o que fez em vez de executar. Foco em DIAGNÓSTICO, não execução (conforme definido por Caio no kick-off).',
     'Mar/2026','Mai/2026','Diário','Redução progressiva dos episódios','Elias','DISC + Kick-off'),
    (9,'Transição Executor → Líder','70% Experiência','Delegação com ensino',
     'Ao delegar: 1) Explicar O QUÊ; 2) Perguntar COMO faria; 3) Complementar; 4) Acompanhar sem fazer. Se resultado não for perfeito, usar como oportunidade de ensino — não refazer.',
     'Mar/2026','Ago/2026','Contínuo','3 delegações/semana neste formato','Elias','Devolutiva'),
    (10,'Transição Executor → Líder','70% Experiência','Foco no diagnóstico técnico',
     'Assumir o papel de ponto focal técnico: receber as demandas da recepção, diagnosticar, e DIRECIONAR o mecânico adequado para a execução. Encurtar a distância recepção ↔ área técnica, sem executar o serviço.',
     'Mar/2026','Ago/2026','Contínuo','Elias é consultado em 80%+ dos diagnósticos','Elias','Kick-off (Caio)'),
    (11,'Transição Executor → Líder','20% Social','Conversa com Caio sobre delegação',
     "Discutir situações onde é difícil soltar a execução. Buscar feedback de como Caio aprendeu a delegar. Caio compartilha quais tarefas já pode 'soltar' para Elias decidir.",
     'Mar/2026','Ago/2026','Quinzenal','2-3 tarefas delegadas por mês','Elias + Caio','Kick-off'),
    (12,'Transição Executor → Líder','10% Formal','Sessão Líder Multiplicador',
     "Eleine apresenta em sessão de 30min: líder não faz melhor que o time — faz o time fazer melhor. Referência: 'O trabalho do líder é trabalhar o trabalho' (Liz Wiseman).",
     'Mar/2026','Mar/2026','Única','Sessão realizada + 3 princípios','Eleine','Devolutiva'),
    (13,'Gestão Técnica e Formação','70% Experiência','Revisão e assinatura de relatórios diários',
     'Revisar e assinar os relatórios diários de cada mecânico. Garantir que está a par de todos os serviços. Usar como instrumento de acompanhamento e gestão, não apenas burocracia.',
     'Mar/2026','Ago/2026','Diário','100% dos relatórios revisados e assinados','Elias','Kick-off (Caio)'),
    (14,'Gestão Técnica e Formação','70% Experiência','Avaliação técnica e comportamental da equipe',
     "Avaliar cada mecânico em 2 dimensões: competência técnica e postura/atitude. Criar ranking para alocar recursos de forma eficaz. Usar termo 'avaliação' (não 'julgamento', conforme orientação de Eleine).",
     'Abr/2026','Mai/2026','Mensal','Ficha de avaliação preenchida para os 7','Elias + Eleine','Kick-off (Caio + Eleine)'),
    (15,'Gestão Técnica e Formação','70% Experiência','Identificar e desenvolver um potencial sucessor',
     'Elias identifica pelo menos 1 mecânico com potencial para cobrir suas funções (inicialmente para férias). Começa a envolvê-lo em diagnósticos e decisões. Formar alguém reduz o risco de dependência técnica que preocupa Caio.',
     'Abr/2026','Ago/2026','Contínuo','1 pessoa sendo desenvolvida ativamente','Elias + Caio','Kick-off (Eleine)'),
    (16,'Gestão Técnica e Formação','70% Experiência','Gestão da planilha de serviços',
     'Assumir a atualização e análise da planilha de gerenciamento de serviços externos. Usar para tomada de decisão sobre alocação de mão de obra, priorização e prazos. Reportar para Caio com análise, não apenas dados.',
     'Mar/2026','Ago/2026','Diário','Planilha atualizada diariamente com análise semanal','Elias','Kick-off (Caio)'),
    (17,'Gestão Técnica e Formação','20% Social','Observar Caio fazendo gestão',
     "Acompanhar Caio em momentos de avaliação de desempenho, tomada de decisão sobre equipe, e feedback. Aprender o 'olhar de gestão' observando quem já faz. Eleine orienta: 'A melhor forma de aprender é observando Caio fazer com ele.'",
     'Mar/2026','Mai/2026','Quinzenal','4+ situações observadas com reflexão','Elias + Caio','Kick-off (Eleine)'),
    (18,'Inteligência Emocional','70% Experiência','Regra Pause & Reframe',
     "Em frustração/conflito, pausar e perguntar: 'Como quero que essa pessoa saia se sentindo dessa conversa?' Reformular se necessário.",
     'Mar/2026','Ago/2026','Diário','3+ situações/semana registradas','Elias','DISC'),
    (19,'Inteligência Emocional','70% Experiência','Mapa de energia semanal',
     'Registrar atividades que dão energia (+) e drenam (-). Identificar padrões em 4 semanas. Ajustar rotina para proteger energia (EN=6,9 — baixa).',
     'Mar/2026','Abr/2026','Semanal','Mapa com 3+ padrões identificados','Elias','DISC'),
    (20,'Inteligência Emocional','20% Social','Feedback anônimo do time',
     'Eleine coleta anonimamente a cada 2 meses: O que melhorou? O que ainda incomoda? Perguntas sobre comunicação, escuta, clima e liderança.',
     'Abr/2026','Ago/2026','Bimestral','Evolução entre 1ª e 2ª coleta','Eleine','DISC + Devolutiva'),
]

DISC_INDICADORES = [
    ('Energia (EN)', 6.9, 'Baixa', 'Capacidade reduzida de absorver mudanças', 'Mapa de energia + rituais de recarga'),
    ('Moral (MRL)', 1.2, 'Normal Baixo', 'Sente que mudanças devem ocorrer', 'Alinhar expectativas com Caio'),
    ('Flexibilidade (IF)', -4.76, 'Normal Baixo', 'Dificuldade para adaptar comportamento', 'Exposição gradual a novas abordagens'),
    ('Energia do Perfil (ENP)', 72.22, 'Extremamente Alta', 'Potencial enorme represado', 'Canalizar em desafios de liderança'),
    ('Exigência do Meio (IEM)', 66.76, 'Normal Alto', 'Ambiente cobra adaptações', 'Reduzir gap com desenvolvimento'),
    ('Positividade (IP)', 25, 'Extremamente Alto', 'Autoestima forte — recurso positivo', 'Usar como alavanca para desenvolvimento'),
]

# ═══════════════════════════════════════════════
output_path = '/home/headless/workspace/cih/operacional/execução/polidiesel/elias-fernandes/PDI_Elias_Fernandes_Polidiesel.xlsx'
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════
# ABA 1 — RESUMO PDI (mantido igual)
# ═══════════════════════════════════════════════
ws = wb.active
ws.title = 'Resumo PDI'
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 45
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15

ws.merge_cells('A1:H1')
c = ws.cell(row=1, column=1, value='PDI — ELIAS FERNANDES DOS SANTOS | POLIDIESEL')
c.font = title_font; c.alignment = title_align

ws.merge_cells('A2:H2')
c = ws.cell(row=2, column=1, value='Plano de Desenvolvimento Individual — Ciclo Março a Agosto/2026 (v3 — revisado com aprendizados)')
c.font = Font(bold=True, size=12, color=DARK_TEAL); c.alignment = title_align

ficha = [
    ('Colaborador', 'Elias Fernandes dos Santos'),
    ('Cargo', 'Responsável Técnico / Supervisor de Mecânica (em transição)'),
    ('Empresa', 'Polidiesel'),
    ('Tempo de Casa', '~20 anos'),
    ('Perfil DISC', 'Executor (D) — Autocrata'),
    ('Liderança Direta', 'Caio Freitas (sócio — responsável pelo acompanhamento do PDI)'),
    ('Consultora CIH', 'Eleine Passos'),
    ('Equipe', '7 colaboradores (área mecânica)'),
    ('Ciclo PDI', 'Março/2026 a Agosto/2026 (6 meses)'),
    ('Confiança de Caio', '10/10 em confiança | 9/10 em expectativa de evolução'),
]
for i, (lab, val) in enumerate(ficha, 4):
    cl = ws.cell(row=i, column=1, value=lab)
    cl.fill = label_fill; cl.font = label_font; cl.border = thin_border
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    cv = ws.cell(row=i, column=2, value=val)
    cv.font = data_font; cv.border = thin_border

r = 15
ws.merge_cells(f'A{r}:H{r}')
ws.cell(row=r, column=1, value='CONTEXTO DA TRANSIÇÃO (KICK-OFF COM CAIO)').font = subtitle_font

contexto = [
    ('Papel definido por Caio', 'Responsável técnico direcionado ao administrativo de mecânica. Foco em DIAGNÓSTICO, não execução. Intermediar recepção e área técnica.'),
    ('Expectativa de Caio', 'Elias conduzir reuniões de produção, ser ponto focal técnico, revisar/assinar relatórios diários, dar direcionamento de tarefas ao time.'),
    ('Gap identificado', 'Área mecânica ficou sem liderança após redução de pessoal. Diagnósticos errados e informações perdidas sem um responsável.'),
    ('Ferramentas em uso', 'Planilha de gerenciamento para serviços externos (em teste há 3 meses). Plano de TV para gestão à vista. Sugestão de tablet para campo.'),
    ('Risco de dependência', 'Caio reconhece risco: Elias é insubstituível tecnicamente. Há 2 anos quase saiu. Precisa formar pelo menos 1 backup.'),
    ('Responsabilidade 50/50', 'Desenvolvimento é 50% do colaborador e 50% da empresa (Eleine). Caio fornecerá suporte (acesso, tempo, recursos).'),
]
for j, (lab, val) in enumerate(contexto, r+1):
    cl = ws.cell(row=j, column=1, value=lab)
    cl.fill = ctx_fill; cl.font = ctx_font; cl.border = thin_border
    ws.merge_cells(start_row=j, start_column=2, end_row=j, end_column=8)
    cv = ws.cell(row=j, column=2, value=val)
    cv.font = data_font; cv.alignment = ctx_data_align; cv.border = thin_border
    ws.row_dimensions[j].height = 35

r = 23
ws.merge_cells(f'A{r}:H{r}')
ws.cell(row=r, column=1, value='COMPETÊNCIAS PRIORIZADAS (4 COMPETÊNCIAS)').font = subtitle_font

r = 24
set_header_row(ws, r, ['#','Competência','Nível Atual','Meta','Peso','Status Mês 1','Status Mês 3','Status Mês 6'])

for ci, (num, nome, atual, meta, peso) in enumerate(COMPETENCIAS, 0):
    row = 25 + ci
    set_data_cell(ws, row, 1, num, align=data_align_c)
    set_data_cell(ws, row, 2, nome, align=data_align_l)
    set_data_cell(ws, row, 3, atual, fill=level_fill(atual), align=data_align_c)
    set_data_cell(ws, row, 4, meta, fill=hfill(LIGHT_GREEN), align=data_align_c)
    set_data_cell(ws, row, 5, peso, align=data_align_c)

r = 30
ws.merge_cells(f'A{r}:H{r}')
ws.cell(row=r, column=1, value='INDICADORES DISC — PONTOS DE ATENÇÃO').font = subtitle_font

r = 31
set_header_row(ws, r, ['Indicador','Valor','Classificação','Impacto','Ação Recomendada','','',''])

for j, (ind, val, classif, impacto, acao) in enumerate(DISC_INDICADORES, r+1):
    set_data_cell(ws, j, 1, ind, align=data_align_l)
    set_data_cell(ws, j, 2, val, align=data_align_c)
    classif_fill = hfill(LIGHT_ORANGE) if classif in ('Baixa','Normal Baixo','Atenção') else hfill(LIGHT_GREEN)
    set_data_cell(ws, j, 3, classif, fill=classif_fill, align=data_align_c)
    set_data_cell(ws, j, 4, impacto, align=data_align_l)
    set_data_cell(ws, j, 5, acao, align=data_align_l)

# ═══════════════════════════════════════════════
# ABA 2 — PLANO DE AÇÕES (mantido igual)
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet('Plano de Ações')
ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 35
ws2.column_dimensions['E'].width = 60
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 12
ws2.column_dimensions['H'].width = 12
ws2.column_dimensions['I'].width = 38
ws2.column_dimensions['J'].width = 15
ws2.column_dimensions['K'].width = 18
ws2.column_dimensions['L'].width = 12

ws2.merge_cells('A1:L1')
ws2.cell(row=1, column=1, value='PLANO DE AÇÕES — PDI ELIAS FERNANDES (v3 — 20 ações, 4 competências)').font = title_font
ws2['A1'].alignment = title_align

set_header_row(ws2, 3, ['#','Competência','Tipo 70-20-10','Ação','Descrição Detalhada',
                         'Prazo Início','Prazo Fim','Frequência','Indicador de Sucesso',
                         'Responsável','Fonte','Status'])

for i, a in enumerate(ACOES):
    row = 4 + i
    num, comp, tipo, acao, desc, ini, fim, freq, indic, resp, fonte = a
    set_data_cell(ws2, row, 1, num, align=data_align_c)
    set_data_cell(ws2, row, 2, comp, align=data_align_l)
    set_data_cell(ws2, row, 3, tipo, fill=type_fill(tipo), align=data_align_l)
    set_data_cell(ws2, row, 4, acao, align=data_align_l)
    set_data_cell(ws2, row, 5, desc, align=data_align_l)
    set_data_cell(ws2, row, 6, ini, align=data_align_c)
    set_data_cell(ws2, row, 7, fim, align=data_align_c)
    set_data_cell(ws2, row, 8, freq, align=data_align_c)
    set_data_cell(ws2, row, 9, indic, align=data_align_l)
    set_data_cell(ws2, row, 10, resp, align=data_align_c)
    set_data_cell(ws2, row, 11, fonte, fill=src_fill, align=data_align_l)
    set_data_cell(ws2, row, 12, 'Pendente', align=data_align_c)

# ═══════════════════════════════════════════════
# ABA 3 — ACOMPANHAMENTO SEMANAL (EXPANDIDO: 26 semanas)
# ═══════════════════════════════════════════════
ws3 = wb.create_sheet('Acompanhamento Semanal')
ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 55
ws3.column_dimensions['D'].width = 14
ws3.column_dimensions['E'].width = 14
ws3.column_dimensions['F'].width = 14
ws3.column_dimensions['G'].width = 14
ws3.column_dimensions['H'].width = 35
ws3.column_dimensions['I'].width = 35
ws3.column_dimensions['J'].width = 30
ws3.column_dimensions['K'].width = 30

ws3.merge_cells('A1:K1')
ws3.cell(row=1, column=1, value='ACOMPANHAMENTO SEMANAL — MARÇO A AGOSTO/2026 (6 MESES — 26 SEMANAS)').font = title_font
ws3['A1'].alignment = title_align

set_header_row(ws3, 3, ['Semana','Período','Foco Principal',
    'Escuta Ativa\n(Nota 1-5)','Transição Líder\n(Nota 1-5)','Gestão Técnica\n(Nota 1-5)',
    'Intelig. Emocional\n(Nota 1-5)',
    'Ações Realizadas','Dificuldades / Bloqueios','Próximos Passos','Observações do Consultor'])

semanas = [
    # Mês 1 — Março
    ('Semana 1','02/03 a 06/03','Leitura DISC + Conduzir 1ª reunião de produção + Início reuniões individuais + Mapa de energia',False),
    ('Semana 2','09/03 a 13/03','1ª sabatina + Regra dos 10s + Delegação com ensino + Revisão de relatórios diários',False),
    ('Semana 3','16/03 a 20/03','Sessão Líder Multiplicador + Foco em diagnóstico (não execução) + Gestão da planilha',False),
    ('Semana 4','23/03 a 27/03','CHECKPOINT MÊS 1 — Avaliar primeiras semanas com Eleine + Caio avalia condução de reuniões',True),
    # Mês 2 — Abril
    ('Semana 5','30/03 a 03/04','Consolidar sabatinas + Aprofundar reuniões individuais + Iniciar avaliação técnica do time',False),
    ('Semana 6','06/04 a 10/04','Mapa de perfis completo + Identificar potencial sucessor + 1ª mentoria com Caio',False),
    ('Semana 7','13/04 a 17/04','1ª coleta feedback anônimo do time + Observar Caio fazendo gestão',False),
    ('Semana 8','20/04 a 24/04','CHECKPOINT MÊS 2 — Avaliar evolução com Eleine + Caio + Avaliar planilha e relatórios',True),
    # Mês 3 — Maio
    ('Semana 9','27/04 a 01/05','Avaliar padrões do mapa de energia + Ficha de avaliação dos 7 mecânicos',False),
    ('Semana 10','04/05 a 08/05','Ajustar ações que drenam energia + Aprofundar formação do sucessor',False),
    ('Semana 11','11/05 a 15/05','Aprofundar delegação — identificar tarefas operacionais para soltar definitivamente',False),
    ('Semana 12','18/05 a 22/05','Consolidar gestão técnica + Preparar evidências de evolução',False),
    ('Semana 13','25/05 a 29/05','CHECKPOINT MÊS 3 — Revisão geral + Ajuste do PDI + Feedback de Caio sobre condução de reuniões',True),
    # Mês 4 — Junho
    ('Semana 14','01/06 a 05/06','Consolidar hábitos de escuta + Aprofundar sabatinas (desafios mais complexos para o time)',False),
    ('Semana 15','08/06 a 12/06','Avaliar sucessor: cobriu alguma ausência? Envolver em diagnósticos mais complexos',False),
    ('Semana 16','15/06 a 19/06','Avaliar delegação total — quais tarefas Elias já soltou definitivamente?',False),
    ('Semana 17','22/06 a 26/06','CHECKPOINT MÊS 4 — 2ª coleta feedback anônimo + Comparar com 1ª coleta + Caio avalia alívio operacional',True),
    # Mês 5 — Julho
    ('Semana 18','29/06 a 03/07','Testar autonomia: Elias conduz 1 semana sem suporte direto de Caio',False),
    ('Semana 19','06/07 a 10/07','Análise da semana autônoma + Ajustes na gestão técnica + Planilha com análise semanal',False),
    ('Semana 20','13/07 a 17/07','Avaliar comunicação adaptada: time percebe diferença? Reuniões individuais com profundidade?',False),
    ('Semana 21','20/07 a 24/07','CHECKPOINT MÊS 5 — Avaliar consolidação de competências + Preparar para avaliação final',True),
    # Mês 6 — Agosto
    ('Semana 22','27/07 a 31/07','Consolidação: Pause & Reframe automático? Delegação natural? Sucessor atuando?',False),
    ('Semana 23','03/08 a 07/08','Preparar evidências de evolução + Atualizar avaliação técnica da equipe',False),
    ('Semana 24','10/08 a 14/08','3ª coleta feedback anônimo do time + Comparar evolução completa (3 coletas)',False),
    ('Semana 25','17/08 a 21/08','Ajustes finais + Relatório de evolução + Parecer de prontidão',False),
    ('Semana 26','24/08 a 28/08','CHECKPOINT FINAL — Avaliação final + Parecer de consolidação da transição + Plano de continuidade',True),
]

for i, (sem, per, foco, is_cp) in enumerate(semanas, 4):
    fill_row = hfill(LIGHT_ORANGE) if is_cp else None
    set_data_cell(ws3, i, 1, sem, font=Font(bold=True, size=10), fill=fill_row, align=data_align_c)
    set_data_cell(ws3, i, 2, per, fill=fill_row, align=data_align_c)
    set_data_cell(ws3, i, 3, foco, fill=fill_row, align=Alignment(vertical='top', wrap_text=True))
    ws3.row_dimensions[i].height = 45

# ═══════════════════════════════════════════════
# ABA 4 — RESULTADO E EVOLUÇÃO (EXPANDIDO: 6 meses)
# ═══════════════════════════════════════════════
ws4 = wb.create_sheet('Resultado e Evolução')
ws4.column_dimensions['A'].width = 48
for col in 'BCDEFGHIJK':
    ws4.column_dimensions[col].width = 14

ws4.merge_cells('A1:K1')
ws4.cell(row=1, column=1, value='RESULTADO DO PDI — EVOLUÇÃO DAS COMPETÊNCIAS (6 MESES)').font = title_font
ws4['A1'].alignment = title_align

set_header_row(ws4, 3, ['Competência','Nível Inicial','Meta','Mês 1','Mês 2','Mês 3','Mês 4','Mês 5','Mês 6 (Final)','Nível Final','Atingiu?'])

comp_names_full = [
    'Escuta Ativa e Comunicação Adaptada',
    'Transição Executor → Líder',
    'Gestão Técnica e Formação da Equipe',
    'Inteligência Emocional e Gestão de Energia',
]
comp_levels = [(2,3),(1,3),(1,3),(2,3)]

for i, (nome, (ini, meta)) in enumerate(zip(comp_names_full, comp_levels), 4):
    set_data_cell(ws4, i, 1, nome, align=data_align_l)
    set_data_cell(ws4, i, 2, ini, fill=level_fill(ini), align=data_align_c)
    set_data_cell(ws4, i, 3, meta, fill=hfill(LIGHT_GREEN), align=data_align_c)

# Indicadores de acompanhamento
r = 9
ws4.merge_cells(f'A{r}:K{r}')
ws4.cell(row=r, column=1, value='INDICADORES DE ACOMPANHAMENTO').font = subtitle_font

r = 10
set_header_row(ws4, r, ['Indicador','Baseline','Meta Mês 1','Real Mês 1','Meta Mês 3','Real Mês 3','Meta Mês 6','Real Mês 6','','',''])

indicadores = [
    ('Reuniões individuais realizadas (/7)', 0, 7, '', 7, '', 7, ''),
    ('Reuniões de produção conduzidas por Elias', 0, '20/mês', '', '20/mês', '', '20/mês', ''),
    ('Sabatinas semanais realizadas', 0, 4, '', 4, '', 4, ''),
    ('Delegações com ensino por semana', 0, 3, '', 3, '', 3, ''),
    ("Episódios 'Quase Meti a Mão' (↓)", '—', 'Registrar', '', 'Reduzir 30%', '', 'Reduzir 70%', ''),
    ('Relatórios diários revisados e assinados', 0, '100%', '', '100%', '', '100%', ''),
    ('Planilha de serviços atualizada', '—', 'Diário', '', 'Diário + análise', '', 'Diário + análise + decisão', ''),
    ('Diagnósticos feitos por Elias (sem executar)', '—', 'Registrar', '', '80%+ dos casos', '', '95%+ dos casos', ''),
    ('Trocas com Caio por mês', 0, 4, '', 4, '', 4, ''),
    ('Potencial sucessor em desenvolvimento', 0, 'Identificar', '', 'Envolver', '', 'Cobrir ausência', ''),
    ('Avaliação técnica da equipe', '—', '—', '', '1ª ficha', '', '3ª atualização', ''),
    ('Nota feedback anônimo do time (1-5)', '—', '—', '', 'Coletar 1ª', '', 'Comparar 3 coletas', ''),
    ('Situações Pause & Reframe por semana', 0, 3, '', 3, '', 3, ''),
    ('Ações do PDI concluídas (acumulado)', '0/20', '7/20', '', '14/20', '', '20/20', ''),
]

for i, ind in enumerate(indicadores, r+1):
    for j, val in enumerate(ind):
        set_data_cell(ws4, i, j+1, val, align=data_align_c if j > 0 else data_align_l)

# Feedback anônimo do time
r2 = r + 1 + len(indicadores) + 1
ws4.merge_cells(f'A{r2}:K{r2}')
ws4.cell(row=r2, column=1, value='FEEDBACK ANÔNIMO DO TIME (3 COLETAS)').font = subtitle_font

r2 += 1
set_header_row(ws4, r2, ['Pergunta','1ª Coleta (Mês 2)\nNota 1-5','Comentários',
                          '2ª Coleta (Mês 4)\nNota 1-5','Comentários',
                          '3ª Coleta (Mês 6)\nNota 1-5','Comentários','Evolução','','',''])

perguntas_fb = [
    'Como você avalia a comunicação do Elias com você?',
    'Você sente que o Elias ouve sua opinião?',
    'O Elias tem dado espaço para você executar e aprender?',
    'Como você se sente nas interações com o Elias?',
    'O Elias tem ajudado no seu desenvolvimento profissional?',
    'Elias conduz bem as reuniões de produção?',
    'Elias dá direcionamento técnico claro?',
    'Você sabe a quem recorrer para questões técnicas?',
    'Nota geral para a liderança do Elias (1 a 5)',
]

for i, p in enumerate(perguntas_fb, r2+1):
    set_data_cell(ws4, i, 1, p, align=Alignment(vertical='top', wrap_text=True))

# Resumo executivo
r3 = r2 + 1 + len(perguntas_fb) + 1
ws4.merge_cells(f'A{r3}:K{r3}')
ws4.cell(row=r3, column=1, value='RESUMO EXECUTIVO — RESULTADO DO CICLO (6 MESES)').font = subtitle_font

resumo_items = [
    'Principais conquistas:',
    'Competências com maior evolução:',
    'Competências que precisam de mais tempo:',
    'Impacto percebido na equipe:',
    'Impacto na operação (diagnósticos, planilha, reuniões):',
    'Percepção de Caio sobre autonomia de Elias:',
    'Evolução do potencial sucessor:',
    'Evolução do feedback anônimo do time (3 coletas):',
    'Alívio operacional para Caio — consegue focar em comercial/gestão?',
    'Recomendações para o próximo ciclo:',
    'Nota geral do PDI (consultor):',
]

for i, item in enumerate(resumo_items, r3+1):
    ws4.cell(row=i, column=1, value=item).font = Font(bold=True, size=10)
    ws4.merge_cells(start_row=i, start_column=2, end_row=i, end_column=11)
    ws4.row_dimensions[i].height = 30

# ═══════════════════════════════════════════════
# ABA 5 — CHECKPOINTS (EXPANDIDO: 6 meses)
# ═══════════════════════════════════════════════
ws5 = wb.create_sheet('Checkpoints')
ws5.column_dimensions['A'].width = 8
ws5.column_dimensions['B'].width = 55
ws5.column_dimensions['C'].width = 25
for col in 'DEFG':
    ws5.column_dimensions[col].width = 15

ws5.merge_cells('A1:G1')
ws5.cell(row=1, column=1, value='CHECKPOINTS MENSAIS — REGISTRO DE SESSÕES (6 MESES)').font = title_font
ws5['A1'].alignment = title_align

def checkpoint_header(ws, r, titulo):
    ws.merge_cells(f'A{r}:G{r}')
    c = ws.cell(row=r, column=1, value=titulo)
    c.fill = hfill(LIGHT_TEAL); c.font = Font(bold=True, size=12, color=DARK_TEAL)
    ws.cell(row=r+1, column=1, value='Participantes:').font = Font(bold=True, size=10)
    ws.cell(row=r+1, column=2, value='Eleine + Elias + Caio').font = data_font
    ws.cell(row=r+2, column=1, value='Data realizada:').font = Font(bold=True, size=10)
    ws.cell(row=r+2, column=2, value='___/___/______').font = data_font
    r += 3
    set_header_row(ws, r, ['#','Pergunta / Item','Resposta / Evidência','','','',''])
    return r + 1

def checkpoint_perguntas(ws, r, perguntas):
    for i, p in enumerate(perguntas):
        row = r + i
        set_data_cell(ws, row, 1, i+1, align=data_align_c)
        set_data_cell(ws, row, 2, p, align=Alignment(vertical='top', wrap_text=True))
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        ws.row_dimensions[row].height = 30
    return r + len(perguntas) + 1

# Checkpoint Mês 1
r = 3
r = checkpoint_header(ws5, r, 'CHECKPOINT MÊS 1 — Fim de Março/2026')
r = checkpoint_perguntas(ws5, r, [
    'Elias leu o material DISC completo?',
    'Quantas reuniões individuais realizou?',
    'Elias está conduzindo as reuniões diárias de produção? Como está a qualidade?',
    'Sabatinas: quantas fez? Como foi a participação do time?',
    "Diário 'Quase Meti a Mão': quantos episódios registrados?",
    'Regra dos 10 segundos: conseguiu praticar?',
    'Elias está revisando e assinando os relatórios diários?',
    'Planilha de serviços: está atualizando? Usa para tomada de decisão?',
    'Trocas com Caio: quantas? Foram úteis?',
    'Mapa de energia: iniciou o registro?',
    'Sessão Líder Multiplicador realizada?',
    'Foco no diagnóstico: Elias está direcionando sem executar?',
    'Nível de energia e motivação de Elias (auto-avaliação 1-5):',
    'Percepção de Caio: Elias está assumindo o papel? O que falta?',
    'Ajustes necessários no PDI:',
])

# Checkpoint Mês 2
r = checkpoint_header(ws5, r, 'CHECKPOINT MÊS 2 — Fim de Abril/2026')
r = checkpoint_perguntas(ws5, r, [
    'Mapa de perfis do time está completo?',
    'Elias está adaptando a comunicação por perfil?',
    'Condução de reuniões: evolução da qualidade e do engajamento do time?',
    'Sabatinas: evolução da participação do time?',
    'Delegação com ensino: 3/semana sendo feitas?',
    "Diário: houve redução dos episódios de 'meter a mão'?",
    'Avaliação técnica e comportamental: Elias preencheu a ficha dos 7?',
    'Potencial sucessor identificado? Quem? Já envolveu em diagnósticos?',
    '1ª coleta de feedback anônimo do time realizada?',
    'Resultado da 1ª coleta (resumo):',
    'Planilha: Elias está fazendo análise ou apenas preenchendo?',
    'Percepção de Caio: Elias está aliviando a carga dele? Em que medida?',
    'Ajustes necessários no PDI:',
])

# Checkpoint Mês 3
r = checkpoint_header(ws5, r, 'CHECKPOINT MÊS 3 — Fim de Maio/2026')
r = checkpoint_perguntas(ws5, r, [
    'Evolução geral nas 4 competências (nota 1-5 cada):',
    'Mapa de energia: padrões identificados? Ajustes feitos?',
    'Sabatinas: o time está mais engajado? Traz soluções sozinho?',
    'Delegação: Elias conseguiu soltar tarefas operacionais?',
    'Gestão técnica: relatórios, planilha, diagnósticos — está consolidado?',
    'Sucessor: pessoa está evoluindo? Cobriu alguma ausência de Elias?',
    'Pause & Reframe: está usando? Exemplos?',
    'Trocas com Caio: estão acontecendo semanalmente?',
    'Percepção geral da equipe (informal):',
    'Caio sente que está com mais liberdade para focar em comercial/gestão?',
    'O que funcionou melhor até agora?',
    'O que precisa mudar para os próximos 3 meses?',
    'Ajustes necessários no PDI para a 2ª metade do ciclo:',
])

# Checkpoint Mês 4 — NOVO
r = checkpoint_header(ws5, r, 'CHECKPOINT MÊS 4 — Fim de Junho/2026')
r = checkpoint_perguntas(ws5, r, [
    '2ª coleta de feedback anônimo realizada? Comparação com a 1ª:',
    'Escuta ativa: o time percebe que Elias escuta mais? Evidências?',
    'Reuniões individuais: estão gerando mudanças reais? Exemplos?',
    'Transição: Elias está mais no papel de líder do que executor?',
    'Diagnósticos: qual % Elias faz sem executar?',
    'Sabatinas: o time traz soluções com mais autonomia?',
    'Sucessor: está preparado para cobrir férias de Elias?',
    'Planilha: Elias toma decisões baseadas nos dados?',
    'Inteligência emocional: houve conflitos? Como Elias lidou?',
    'Caio nota alívio operacional real? Em que áreas?',
    'Elias expressa satisfação com o novo papel?',
    'O que consolidou e o que ainda precisa de atenção?',
    'Prioridades para os 2 meses finais:',
])

# Checkpoint Mês 5 — NOVO
r = checkpoint_header(ws5, r, 'CHECKPOINT MÊS 5 — Fim de Julho/2026')
r = checkpoint_perguntas(ws5, r, [
    'Semana autônoma: como foi? Elias conduziu sem Caio por perto?',
    'Comunicação adaptada por perfil: virou hábito natural?',
    'Delegação: Elias aceita resultados "bom o suficiente" sem refazer?',
    'Time reconhece Elias como líder (não mais como par)?',
    'Gestão técnica: relatórios, planilha e diagnósticos automáticos?',
    'Sucessor: evoluiu o suficiente para dar segurança à operação?',
    'Pause & Reframe: está automático ou precisa de lembrança?',
    'Energia de Elias: melhorou com ajustes do mapa?',
    'O que Elias faz hoje que não fazia em março?',
    'Caio confia em Elias para decisões autônomas?',
    'Preparar evidências para avaliação final:',
    'Pontos que precisam de atenção no mês final:',
])

# Checkpoint Mês 6 — NOVO (FINAL)
r = checkpoint_header(ws5, r, 'CHECKPOINT FINAL — Fim de Agosto/2026')
r = checkpoint_perguntas(ws5, r, [
    'Nota final nas 4 competências (1-5 cada):',
    'Escuta ativa: consolidou? O time confirma?',
    'Transição: Elias é líder ou ainda executa?',
    'Gestão técnica: consolidou relatórios, planilha, avaliações?',
    'Inteligência emocional: conflitos reduzidos? Clima melhorou?',
    '3ª coleta feedback anônimo: evolução geral comparada à 1ª?',
    'Sucessor: pronto para cobrir Elias? Elias confia nele?',
    'Caio se sente aliviado operacionalmente? Quanto (%)?',
    'O que funcionou melhor em todo o PDI?',
    'O que não funcionou e por quê?',
    'Recomendações de desenvolvimento para o próximo ciclo:',
    'Parecer final: Elias consolidou a transição?',
    'Próximos passos (plano de continuidade):',
])

# ═══════════════════════════════════════════════
# ABA 6 — SEMÁFORO (EXPANDIDO: 6 meses)
# ═══════════════════════════════════════════════
ws6 = wb.create_sheet('Semáforo')
ws6.column_dimensions['A'].width = 22
ws6.column_dimensions['B'].width = 40
ws6.column_dimensions['C'].width = 18
for col in 'DEFGHI':
    ws6.column_dimensions[col].width = 14

ws6.merge_cells('A1:I1')
ws6.cell(row=1, column=1, value='SEMÁFORO DE PROGRESSO — VISÃO EXECUTIVA (6 MESES)').font = title_font
ws6['A1'].alignment = title_align

ws6.merge_cells('A2:I2')
c = ws6.cell(row=2, column=1, value='Verde = No caminho | Amarelo = Atenção | Vermelho = Intervenção necessária')
c.font = Font(size=10, color=GRAY_TEXT); c.alignment = data_align_c

set_header_row(ws6, 4, ['Dimensão','Indicador-Chave','Meta','Mês 1','Mês 2','Mês 3','Mês 4','Mês 5','Mês 6'])

semaforo_items = [
    ('Escuta Ativa', 'Reuniões individuais realizadas', '7/mês'),
    ('Escuta Ativa', 'Condução de reuniões de produção', 'Diário'),
    ('Escuta Ativa', 'Time se sente ouvido (feedback)', '≥ 3,5/5'),
    ('Transição Líder', 'Sabatinas realizadas', '4/mês'),
    ('Transição Líder', 'Tarefas operacionais delegadas', '≥ 2 novas/mês'),
    ('Transição Líder', 'Diagnósticos feitos sem executar', '80%+ → 95%'),
    ('Transição Líder', "Episódios 'mão na massa' (↓)", 'Reduzir 70%'),
    ('Gestão Técnica', 'Relatórios diários revisados/assinados', '100%'),
    ('Gestão Técnica', 'Planilha de serviços atualizada', 'Diário + análise'),
    ('Gestão Técnica', 'Avaliação técnica da equipe preenchida', '3 atualizações'),
    ('Gestão Técnica', 'Sucessor em desenvolvimento', 'Cobrir ausência'),
    ('Intelig. Emocional', 'Situações Pause & Reframe', '3/semana'),
    ('Intelig. Emocional', 'Nota do time sobre clima', '≥ 3,5/5'),
    ('Intelig. Emocional', 'Evolução no feedback anônimo', '3 coletas com melhora'),
    ('Geral', 'Ações do PDI concluídas', '100%'),
    ('Geral', 'Trocas com Caio realizadas', '4/mês'),
    ('Geral', 'Energia/motivação Elias (1-5)', '≥ 4'),
    ('Geral', 'Caio sente alívio operacional?', 'Sim'),
    ('Geral', 'Time reconhece Elias como líder?', 'Sim'),
]

for i, (dim, ind, meta) in enumerate(semaforo_items, 5):
    set_data_cell(ws6, i, 1, dim, align=data_align_l)
    set_data_cell(ws6, i, 2, ind, align=data_align_l)
    set_data_cell(ws6, i, 3, meta, align=data_align_c)

r = 5 + len(semaforo_items) + 1
ws6.cell(row=r, column=1, value='LEGENDA:').font = Font(bold=True, size=10)

r += 1
c = ws6.cell(row=r, column=1, value='VERDE')
c.fill = hfill(GREEN_BG); c.font = Font(bold=True, size=10, color=WHITE); c.alignment = data_align_c
ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
ws6.cell(row=r, column=2, value='70%+ das ações concluídas, indicadores no caminho, time reporta evolução, Caio percebe alívio operacional').font = data_font

r += 1
c = ws6.cell(row=r, column=1, value='AMARELO')
c.fill = hfill(YELLOW_BG); c.font = Font(bold=True, size=10, color=DARK_GRAY); c.alignment = data_align_c
ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
ws6.cell(row=r, column=2, value='40-69% das ações, alguns indicadores atrasados, 1 checkpoint com pendências, Elias ainda executa mais que diagnostica').font = data_font

r += 1
c = ws6.cell(row=r, column=1, value='VERMELHO')
c.fill = hfill(RED_BG); c.font = Font(bold=True, size=10, color=WHITE); c.alignment = data_align_c
ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
ws6.cell(row=r, column=2, value='<40% das ações, indicadores estagnados, time não percebe mudança, energia caindo, sem avanço na formação de sucessor').font = data_font

# ═══════════════════════════════════════════════
# ABA 7 — COMPETÊNCIAS REFERÊNCIA (mantido igual)
# ═══════════════════════════════════════════════
ws7 = wb.create_sheet('Competências Referência')
ws7.column_dimensions['A'].width = 22
ws7.column_dimensions['B'].width = 32
ws7.column_dimensions['C'].width = 55
ws7.column_dimensions['D'].width = 22
ws7.column_dimensions['E'].width = 20

ws7.merge_cells('A1:E1')
ws7.cell(row=1, column=1, value='MAPEAMENTO DE COMPETÊNCIAS — BASE PARA AVALIAÇÃO').font = title_font
ws7['A1'].alignment = title_align

ws7.merge_cells('A2:E2')
c = ws7.cell(row=2, column=1, value='Fonte: Planilha 190 Exemplos de Competências para Avaliação de Desempenho + Contexto Polidiesel')
c.font = Font(size=10, color=GRAY_TEXT); c.alignment = data_align_c

set_header_row(ws7, 4, ['Categoria','Competência','Relevância para Elias','Nível Atual Estimado (1-5)','Prioridade no PDI'])

comps_ref = [
    ('Gestão de Pessoas','Coaching e Mentoria','CRÍTICA — precisa formar sucessor e desenvolver o time',1,'FFEBEE','E53935','Competência 3'),
    ('Gestão de Pessoas','Gerenciamento de Desempenho','ALTA — avaliar mecânicos tecnicamente e por postura',1,'FFEBEE','E53935','Competência 3'),
    ('Gestão de Pessoas','Trabalho em Equipe','ALTA — transitar de executor para facilitador do time',2,'FFFDE7','F57C00','Competência 2'),
    ('Gestão de Pessoas','Treinamento e Desenvolvimento','MÉDIA — compartilhar conhecimento técnico com a equipe',2,'FFFDE7',None,'Competência 3'),
    ('Gestão de Pessoas','Adaptabilidade e Flexibilidade','ALTA — IF=-4,76, transição de papel exige flexibilidade',2,'FFFDE7','F57C00','Competências 2 e 4'),
    ('Comunicação','Habilidades de Escuta',"CRÍTICA — DISC mostra 'Capacidade de Ouvir' muito baixa",2,'FFEBEE','E53935','Competência 1'),
    ('Comunicação','Habilidades de Fala',"MÉDIA — Elias é 'de poucas palavras' (Caio no kick-off)",2,'FFFDE7',None,'Competência 1'),
    ('Comunicação','Técnicas de Negociação','MÉDIA — intermediar recepção ↔ área técnica com clientes',2,'FFFDE7',None,'Competência 1'),
    ('Liderança','Gestão Estratégica','ALTA — visão do quadro geral, não apenas execução',1,'FFEBEE','E53935','Competência 2'),
    ('Liderança','Responsabilidade e Confiabilidade','ALTA — já possui (Caio nota 10/10), manter e formalizar',4,'E8F5E9',None,'Ativo existente'),
    ('Liderança','Persuasão e Influência','MÉDIA — time ainda o vê como par, precisa construir autoridade',2,'FFFDE7',None,'Competência 2'),
    ('Desenvolvimento Pessoal','Potencial de Crescimento','ALTA — Caio vê 9/10 de potencial',3,'FFF3E0',None,'Ativo existente'),
    ('Desenvolvimento Pessoal','Mudança de Liderança','CRÍTICA — lidar com impactos da transição no time',1,'FFEBEE','F57C00','Competência 2'),
    ('Desenvolvimento Pessoal','Foco nos Resultados','ALTA — já possui (perfil Executor), direcionar para gestão',4,'E8F5E9',None,'Ativo existente'),
    ('Raciocínio Lógico','Tomada de Decisões','ALTA — papel de diagnóstico e direcionamento técnico',3,'FFF3E0',None,'Competência 3'),
    ('Raciocínio Lógico','Resolução de Problemas','ALTA — ponto focal técnico, resolver sem executar',3,'FFF3E0',None,'Competência 2'),
    ('Raciocínio Lógico','Abordagem Metódica','MÉDIA — uso de planilha, relatórios, gestão organizada',2,'FFFDE7',None,'Competência 3'),
    ('Comportamental','Empatia','CRÍTICA — DISC mostra empatia muito baixa',1,'FFEBEE','E53935','Competência 4'),
    ('Comportamental','Habilidades Interpessoais','ALTA — construir relações de liderança, não só de execução',2,'FFFDE7',None,'Competências 1 e 4'),
    ('Comportamental','Gestão do Tempo','MÉDIA — equilibrar novo papel administrativo com operacional residual',2,'FFFDE7',None,'Competência 2'),
    ('Competências Técnicas','Conhecimento de Informática','MÉDIA — planilha, possível uso de tablet no futuro',2,'FFFDE7',None,'Competência 3'),
    ('Transferível','Planejamento e Organização','ALTA — gestão diária, alocação de recursos, priorização',2,'FFFDE7',None,'Competência 3'),
]

for i, (cat, comp, rel, nivel, nivel_color, rel_color, prio) in enumerate(comps_ref, 5):
    set_data_cell(ws7, i, 1, cat, align=data_align_l)
    set_data_cell(ws7, i, 2, comp, align=data_align_l)
    rel_font = Font(bold=True, size=10, color=rel_color) if rel_color else data_font
    set_data_cell(ws7, i, 3, rel, font=rel_font, align=data_align_l)
    set_data_cell(ws7, i, 4, nivel, fill=hfill(nivel_color), align=data_align_c)
    set_data_cell(ws7, i, 5, prio, align=data_align_c)

# ═══════════════════════════════════════════════
# ABA 8 — GUIA 1-1 (NOVA — adaptada para Elias ↔ Caio)
# ═══════════════════════════════════════════════
ws8 = wb.create_sheet('Guia 1-1')

ws8.column_dimensions['A'].width = 6
ws8.column_dimensions['B'].width = 32
ws8.column_dimensions['C'].width = 50
ws8.column_dimensions['D'].width = 50
ws8.column_dimensions['E'].width = 22
ws8.column_dimensions['F'].width = 22
ws8.column_dimensions['G'].width = 22
ws8.column_dimensions['H'].width = 22

# TÍTULO
ws8.merge_cells('A1:H1')
c = ws8.cell(row=1, column=1, value='GUIA DE REUNIÕES 1:1 — CAIO ↔ ELIAS')
c.font = title_font; c.alignment = title_align

ws8.merge_cells('A2:H2')
c = ws8.cell(row=2, column=1, value='Instruções, modelos e exemplos práticos para acompanhamento do PDI via reuniões one-on-one')
c.font = Font(size=10, color=GRAY_TEXT); c.alignment = data_align_c
ws8.row_dimensions[1].height = 32
ws8.row_dimensions[2].height = 22

# SEÇÃO 1 — O QUE É
r = 4
r = section_title(ws8, r, '1. O QUE É A REUNIÃO 1:1 E POR QUE FAZER', 8)

items_oq = [
    ('Definição',
     'Encontro individual e recorrente entre Caio (sócio) e Elias (líder em transição), com foco no desenvolvimento da liderança, alinhamento e acompanhamento do PDI.'),
    ('Quem participa',
     'Caio Freitas (sócio) + Elias Fernandes (supervisor). Eleine participa nos checkpoints mensais.'),
    ('Frequência recomendada',
     'Semanal nas primeiras 8 semanas (15-20 min). Depois pode ser quinzenal (20-30 min). Nunca cancelar sem reagendar.'),
    ('Onde realizar',
     'Preferencialmente no escritório, sem interrupções da oficina. Se não for possível, separar um momento na área técnica com privacidade.'),
    ('Regra de ouro',
     'A reunião é 70% do Elias e 30% do Caio. Caio provoca com perguntas, Elias traz os desafios. Lembrar: Elias é de poucas palavras — dar tempo e espaço.'),
    ('Por que é importante',
     'Elias está em transição de executor para líder após 20 anos. Sem acompanhamento frequente, a tendência é voltar ao padrão antigo (meter a mão na execução).'),
    ('Benefícios comprovados',
     'Acelera a transição de cargo, reduz recaídas ao modo executor, fortalece a confiança mútua e previne a perda do colaborador (Elias quase saiu há 2 anos).'),
]

for lab, desc in items_oq:
    cell(ws8, r, 1, '', fill=hfill(LIGHT_TEAL))
    cell(ws8, r, 2, lab, font=bold_font, fill=hfill(LIGHT_TEAL))
    ws8.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    cell(ws8, r, 3, desc, align=data_align_l)
    ws8.row_dimensions[r].height = 40
    r += 1

# SEÇÃO 2 — BOAS PRÁTICAS
r += 1
r = section_title(ws8, r, '2. BOAS PRÁTICAS PARA A 1:1 COM ELIAS', 8)

set_header_row(ws8, r, ['', 'Fazer', 'Evitar', '', '', '', '', ''])
r += 1

boas_praticas = [
    ('Agendar horário fixo — Elias precisa de rotina e previsibilidade (perfil Executor)',
     'Marcar de improviso ou "quando der" — isso sinaliza que não é prioridade'),
    ('Começar reconhecendo algo que Elias fez bem — ele responde a reconhecimento',
     'Começar corrigindo ou listando falhas — Elias pode se fechar'),
    ('Fazer perguntas abertas e ESPERAR a resposta (Elias é de poucas palavras)',
     'Responder pela pessoa ou apressar — contar até 10 mentalmente'),
    ('Conectar a conversa com as competências do PDI — "como está a delegação?"',
     'Falar só de tarefas operacionais sem foco em desenvolvimento'),
    ('Usar exemplos concretos — "na terça vi que você diagnosticou sem executar"',
     'Feedback genérico — "você precisa melhorar a comunicação"'),
    ('Anotar compromissos ao final — 2-3 ações específicas até a próxima 1:1',
     'Sair sem combinados claros — Elias funciona melhor com metas concretas'),
    ('Celebrar progressos — Elias tem Positividade extremamente alta, usar a favor',
     'Só falar de problemas — energia (EN) já é baixa, não drenar mais'),
    ('Compartilhar como Caio aprendeu — "quando eu estava aprendendo a delegar..."',
     'Dar ordens sem contexto — Elias precisa entender o porquê'),
]

for fazer, evitar in boas_praticas:
    cell(ws8, r, 1, '✓', font=Font(bold=True, size=10, color='43A047'), align=data_align_c)
    cell(ws8, r, 2, fazer, align=data_align_l, merge_end=4)
    cell(ws8, r, 5, '✗', font=Font(bold=True, size=10, color='E53935'), align=data_align_c)
    ws8.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    cell(ws8, r, 6, evitar, align=data_align_l)
    ws8.row_dimensions[r].height = 40
    r += 1

# SEÇÃO 3 — MODELO DE PAUTA
r += 1
r = section_title(ws8, r, '3. MODELO DE PAUTA — ADAPTADO PARA O PDI DO ELIAS', 8)

ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cell(ws8, r, 1, 'Modelo "8 Áreas" adaptado para transição de liderança — conecta cada bloco às 4 competências do PDI.',
     font=tip_font, align=data_align_l)
r += 1

set_header_row(ws8, r, ['#', 'Bloco da Pauta', 'O que abordar', 'Exemplo de Pergunta para Elias', 'Tempo', 'Competência PDI', '', ''])
r += 1

blocos_pauta = [
    (1, 'Check-in e Energia',
     'Como Elias está se sentindo? Nível de energia, frustrações, motivação. Atenção: EN é baixa (6,9).',
     '"Elias, como está sua energia essa semana? O que te drenou e o que te deu gás?"',
     '3 min', 'Comp. 4'),
    (2, 'Conquistas da Semana',
     'O que deu certo? Focar em momentos onde Elias exerceu liderança sem executar.',
     '"Me conta uma situação onde você direcionou o time sem precisar pegar na ferramenta."',
     '3 min', 'Comp. 2'),
    (3, 'Escuta e Comunicação',
     'Como estão as reuniões individuais? Usou a regra dos 10 segundos? Adaptou comunicação por perfil?',
     '"Na reunião de produção de hoje, você percebeu se deu espaço pro time falar?"',
     '3 min', 'Comp. 1'),
    (4, 'Transição e Delegação',
     'Quantas vezes sentiu impulso de executar? Delegou com ensino? Sabatina funcionou?',
     '"Quantas vezes essa semana você quase meteu a mão? O que fez em vez disso?"',
     '4 min', 'Comp. 2'),
    (5, 'Gestão Técnica',
     'Relatórios diários revisados? Planilha atualizada? Diagnósticos feitos?',
     '"Algum diagnóstico essa semana te fez querer executar o serviço? Como lidou?"',
     '3 min', 'Comp. 3'),
    (6, 'Formação do Time',
     'Evolução do sucessor? Ensinou algo novo ao time? Avaliação técnica atualizada?',
     '"Como está o [nome do sucessor]? Você o envolveu em algum diagnóstico difícil?"',
     '3 min', 'Comp. 3'),
    (7, 'Gestão Emocional',
     'Usou Pause & Reframe? Conflitos com o time? Como se sentiu em situações difíceis?',
     '"Teve alguma situação com o time que te tirou do sério? O que fez?"',
     '3 min', 'Comp. 4'),
    (8, 'Fechamento e Compromissos',
     'Resumir pontos, definir 2-3 ações específicas, alinhar expectativas da próxima semana.',
     '"Para essa semana, nossos combinados são: 1) ___, 2) ___, 3) ___. Fechado?"',
     '3 min', 'Todas'),
]

for num, bloco, abordagem, pergunta, tempo, comp in blocos_pauta:
    cell(ws8, r, 1, num, font=Font(bold=True, size=11, color=TEAL), align=data_align_c)
    cell(ws8, r, 2, bloco, font=bold_font, align=data_align_l)
    cell(ws8, r, 3, abordagem, align=data_align_l)
    cell(ws8, r, 4, pergunta, font=example_font, align=data_align_l)
    cell(ws8, r, 5, tempo, align=data_align_c)
    ws8.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    cell(ws8, r, 6, comp, align=data_align_c)
    ws8.row_dimensions[r].height = 55
    r += 1

ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cell(ws8, r, 1, 'Tempo total: ~25 min | Não precisa usar todos os blocos — escolha 4-5 por reunião e alterne.',
     font=tip_font, align=data_align_l)
r += 1

# SEÇÃO 4 — PERGUNTAS POR COMPETÊNCIA
r += 1
r = section_title(ws8, r, '4. BANCO DE PERGUNTAS POR COMPETÊNCIA DO PDI', 8)

ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cell(ws8, r, 1, 'Caio pode escolher 2-3 perguntas por reunião. Lembrar: Elias é de poucas palavras — dar tempo para responder.',
     font=tip_font, align=data_align_l)
r += 1

perguntas_por_comp = [
    ('Escuta Ativa e Comunicação Adaptada', hfill(LIGHT_BLUE), [
        'Como estão as reuniões individuais com o time? Alguém te surpreendeu?',
        'Você percebeu diferença quando adapta a comunicação ao perfil da pessoa?',
        'Na reunião de produção, você deu espaço para o time falar ou foi mais direto?',
        'Usou a regra dos 10 segundos? Foi difícil esperar?',
        'Algum membro do time veio te procurar espontaneamente? Isso mudou desde que começou as individuais?',
    ]),
    ('Transição de Executor para Líder', hfill(LIGHT_YELLOW), [
        'Quantas vezes essa semana você sentiu vontade de "meter a mão"? O que fez em vez disso?',
        'Conseguiu delegar algo usando o método (O QUÊ → COMO faria → Complementar → Acompanhar)?',
        'Na sabatina, o time trouxe soluções ou ficou esperando você resolver?',
        'Tem alguma tarefa operacional que você ainda não conseguiu soltar? O que te trava?',
        'Você se vê mais como líder ou como o melhor mecânico da equipe? O que mudou?',
    ]),
    ('Gestão Técnica e Formação da Equipe', hfill(LIGHT_GREEN), [
        'Está conseguindo revisar os relatórios diários? Encontrou algo que precisou de intervenção?',
        'A planilha de serviços: está só preenchendo ou está tirando conclusões dos dados?',
        'Como está o [sucessor]? Ele já fez algum diagnóstico sozinho?',
        'Você confiaria no time se precisasse se ausentar por 3 dias? O que falta?',
        'Na avaliação técnica, alguém te surpreendeu positiva ou negativamente?',
    ]),
    ('Inteligência Emocional e Gestão de Energia', hfill(LIGHT_PURPLE), [
        'Sua energia essa semana foi mais "+" ou mais "–"? O que influenciou?',
        'Teve algum momento de frustração com o time? Como lidou?',
        'Usou o Pause & Reframe? Me dá um exemplo de como reformulou a situação.',
        'O que te dá mais energia: diagnosticar ou liderar o time?',
        'Você sente que o time te respeita mais como líder agora do que no mês passado?',
    ]),
]

for comp_name, comp_fill, perguntas in perguntas_por_comp:
    ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    cell(ws8, r, 1, comp_name, font=Font(bold=True, size=10, color=DARK_TEAL), fill=comp_fill,
         align=Alignment(horizontal='left', vertical='center'))
    ws8.row_dimensions[r].height = 24
    r += 1
    for p in perguntas:
        cell(ws8, r, 1, '•', align=data_align_c)
        ws8.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        cell(ws8, r, 2, p, align=data_align_l)
        ws8.row_dimensions[r].height = 22
        r += 1
    r += 1

# SEÇÃO 5 — EXEMPLOS DE FEEDBACK
r = section_title(ws8, r, '5. EXEMPLOS DE FEEDBACK — BASEADOS NO CONTEXTO DO ELIAS', 8)

ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cell(ws8, r, 1, 'Modelos de feedback que Caio pode adaptar. Lembrar: Elias tem Positividade alta — responde bem a reconhecimento. Energia baixa — não sobrecarregar com críticas.',
     font=tip_font, align=data_align_l)
r += 1

set_header_row(ws8, r, ['', 'Situação', 'Tipo', 'Exemplo de Feedback', '', '', '', ''])
r += 1

feedbacks = [
    ('Elias diagnosticou problema e direcionou mecânico sem executar',
     'Positivo',
     '"Elias, aquele diagnóstico do motor do caminhão — você identificou o problema, direcionou o João e ele resolveu. Isso é exatamente o papel que a gente definiu. Quando você faz isso, o time cresce e a operação flui."'),
    ('Elias executou um serviço que deveria ter delegado',
     'Construtivo',
     '"Elias, vi que você fez o ajuste da bomba ontem. Entendo que é mais rápido quando você faz, mas lembra do nosso combinado? Quem poderia ter feito? O que te impediu de delegar?"'),
    ('Reunião de produção foi bem conduzida — time participou',
     'Positivo',
     '"Elias, a reunião de produção hoje foi muito boa. Você deu espaço pro Marcos falar e ele trouxe uma ideia que nenhum de nós tinha pensado. Isso acontece quando o líder escuta."'),
    ('Elias foi impaciente com mecânico que errou',
     'Construtivo',
     '"Elias, percebi que você ficou frustrado com o Pedro. Entendo — o erro foi evitável. Mas lembra do Pause & Reframe? Como você queria que o Pedro saísse se sentindo daquela conversa? Da próxima, o que poderia fazer diferente?"'),
    ('Sucessor fez primeiro diagnóstico sozinho com sucesso',
     'Positivo',
     '"Elias, o [nome] fez o diagnóstico do motor sozinho e acertou. Isso é resultado do seu trabalho de formação. Há 3 meses ele não saberia por onde começar. Você está formando alguém que pode te dar suporte."'),
    ('Elias não fez reunião individual no mês',
     'Construtivo',
     '"Elias, esse mês não tivemos as reuniões individuais. Sei que a oficina está cheia, mas essas conversas são a base da sua liderança. Que dia da semana funciona melhor para encaixar? Precisa de apenas 20 minutos por pessoa."'),
    ('Planilha está sendo atualizada mas sem análise',
     'Construtivo',
     '"Elias, a planilha está sempre em dia — ótimo. Mas ela pode ser mais que um preenchimento. O que esses números dizem sobre a alocação da equipe? Tenta me trazer uma análise semana que vem — tipo: quais serviços atrasaram e por quê?"'),
    ('Elias usou Pause & Reframe em situação tensa',
     'Positivo',
     '"Elias, aquela situação com o cliente que reclamou do prazo — você manteve a calma, explicou o que aconteceu e propôs solução. Antes, isso poderia ter virado discussão. Isso é inteligência emocional na prática."'),
    ('Sabatina gerou solução vinda do time',
     'Positivo',
     '"Elias, na sabatina de ontem o time trouxe a solução do problema da peça. Você fez a pergunta certa e esperou. Isso é o líder multiplicador: fazer o time pensar, não dar a resposta."'),
    ('Elias voltou a executar serviços no final do mês (recaída)',
     'Construtivo + Incentivo',
     '"Elias, sei que final de mês a pressão aumenta e a tentação de pegar na ferramenta é grande. Mas olha quanto você evoluiu — no mês 1 eram 10 episódios, agora foram 3. Vamos pensar juntos como blindar o final do mês?"'),
]

for sit, tipo, fb in feedbacks:
    tipo_fill = hfill(LIGHT_GREEN) if 'Positivo' in tipo else hfill(LIGHT_YELLOW)
    cell(ws8, r, 1, '', fill=tipo_fill)
    cell(ws8, r, 2, sit, font=bold_font, align=data_align_l)
    cell(ws8, r, 3, tipo, fill=tipo_fill, font=Font(bold=True, size=10), align=data_align_c)
    ws8.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
    cell(ws8, r, 4, fb, font=example_font, align=data_align_l)
    ws8.row_dimensions[r].height = 65
    r += 1

# SEÇÃO 6 — REGISTRO DAS REUNIÕES
r += 1
r = section_title(ws8, r, '6. REGISTRO DAS REUNIÕES 1:1 (PREENCHER A CADA ENCONTRO)', 8)

set_header_row(ws8, r, ['#', 'Data', 'Competência em foco', 'Principais pontos discutidos',
                    'Compromissos assumidos', 'Prazo', 'Status', 'Observações'])
r += 1

for i in range(1, 13):
    cell(ws8, r, 1, i, font=Font(bold=True, size=10, color=TEAL), align=data_align_c)
    cell(ws8, r, 2, '___/___/___', align=data_align_c)
    for col in range(3, 9):
        cell(ws8, r, col, '', align=data_align_l)
    ws8.row_dimensions[r].height = 50
    r += 1

# SEÇÃO 7 — CALENDÁRIO SUGERIDO
r += 1
r = section_title(ws8, r, '7. CALENDÁRIO SUGERIDO DE 1:1 — MARÇO A AGOSTO/2026', 8)

set_header_row(ws8, r, ['#', 'Data sugerida', 'Foco principal', 'Competência PDI',
                    'Perguntas sugeridas', 'Tipo de feedback', '', ''])
r += 1

calendario = [
    (1, '2ª semana Mar', 'Onboarding da 1:1 + como Elias se sente com a transição',
     'Todas', 'Check-in geral, expectativas, medos e energia', 'Incentivo e acolhimento'),
    (2, '4ª semana Mar', 'CP Mês 1: Reuniões de produção + primeiras delegações + DISC',
     'Comp. 1 e 2', 'Condução de reuniões, regra 10s, diário', 'Construtivo + positivo'),
    (3, '2ª semana Abr', 'Mapa de perfis + feedback anônimo + avaliação técnica',
     'Comp. 1 e 3', 'Perfis do time, sucessor, planilha', 'Positivo sobre evolução'),
    (4, '4ª semana Abr', 'CP Mês 2: Evolução geral + alívio para Caio + 1ª coleta',
     'Todas', 'Balanço geral, delegação, gestão técnica', 'Balanço construtivo'),
    (5, '2ª semana Mai', 'Energia + sabatinas + consolidação de delegação',
     'Comp. 2 e 4', 'Mapa de energia, sabatinas, Pause & Reframe', 'Celebrar hábitos'),
    (6, '4ª semana Mai', 'CP Mês 3: Revisão de metade do ciclo + ajustes para 2ª metade',
     'Todas', 'Visão de Caio, equipe, diagnósticos', 'Feedback formal de progresso'),
    (7, '2ª semana Jun', 'Autonomia do time + sucessor em ação + comunicação adaptada',
     'Comp. 1 e 3', 'Sucessor, comunicação, gestão técnica', 'Positivo ou plano de ação'),
    (8, '4ª semana Jun', 'CP Mês 4: 2ª coleta feedback + alívio operacional para Caio',
     'Todas', 'Feedback anônimo, evolução, delegação total', 'Comparativo 1ª vs 2ª coleta'),
    (9, '2ª semana Jul', 'Semana autônoma: Elias conduz sem suporte direto de Caio',
     'Comp. 2 e 3', 'Autonomia total, decisões, diagnósticos', 'Avaliar prontidão'),
    (10, '4ª semana Jul', 'CP Mês 5: Consolidação de competências + evidências',
     'Todas', 'Evolução geral, energia, time, sucessor', 'Preparar avaliação final'),
    (11, '2ª semana Ago', '3ª coleta feedback + preparar relatório final',
     'Todas', 'Evolução completa (3 coletas), evidências', 'Reconhecimento do progresso'),
    (12, '4ª semana Ago', 'CP FINAL: Avaliação final + parecer de consolidação + plano futuro',
     'Todas', 'Nota final, parecer, próximos passos', 'Relatório final'),
]

for num, data, foco, comp, pergs, fb_tipo in calendario:
    is_cp = 'CP' in foco or 'FINAL' in foco
    row_fill = hfill(LIGHT_ORANGE) if is_cp else None
    cell(ws8, r, 1, num, font=Font(bold=True, size=10, color=TEAL), fill=row_fill, align=data_align_c)
    cell(ws8, r, 2, data, font=bold_font, fill=row_fill, align=data_align_c)
    cell(ws8, r, 3, foco, fill=row_fill, align=data_align_l)
    cell(ws8, r, 4, comp, fill=row_fill, align=data_align_c)
    cell(ws8, r, 5, pergs, fill=row_fill, align=data_align_l)
    ws8.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    cell(ws8, r, 6, fb_tipo, fill=row_fill, align=data_align_l)
    ws8.row_dimensions[r].height = 45
    r += 1

# SEÇÃO 8 — DICAS PARA CAIO
r += 1
r = section_title(ws8, r, '8. DICAS RÁPIDAS PARA CAIO', 8)

dicas = [
    'Elias é Executor — responde melhor a desafios concretos ("tenta fazer X essa semana") do que a teoria ("você deveria ser mais assim").',
    'Energia BAIXA (6,9) — não sobrecarregar a 1:1 com cobranças. Equilibrar: para cada ponto de melhoria, reconhecer uma conquista.',
    'Positividade EXTREMAMENTE ALTA — Elias tem autoconfiança forte. Use a favor: "Confio em você pra isso" funciona melhor que "Você precisa melhorar".',
    'Elias é de poucas palavras — faça a pergunta e ESPERE. Conte até 10. O silêncio é produtivo.',
    'Compartilhe sua própria experiência — "quando eu comecei a delegar, também foi difícil..." cria conexão e mostra que você entende.',
    'Use o diário "Quase Meti a Mão" como ponto de partida — não para julgar, mas para entender o impulso.',
    'Registre tudo nesta aba — os registros alimentam os checkpoints e o relatório final de evolução.',
    'Consistência > perfeição: 15 min semanais de 1:1 valem mais que 1 hora mensal. Nunca pule sem reagendar.',
]

for dica in dicas:
    cell(ws8, r, 1, '→', font=Font(bold=True, size=11, color=TEAL), align=data_align_c)
    ws8.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    cell(ws8, r, 2, dica, align=data_align_l)
    ws8.row_dimensions[r].height = 32
    r += 1

# Rodapé
r += 1
ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cell(ws8, r, 1, 'Guia elaborado pela CIH Consultoria em Inteligência Humana | Consultora: Eleine Passos | Fontes: Convenia, Feedz, Qulture.Rocks',
     font=Font(size=9, italic=True, color=GRAY_TEXT), align=data_align_c)

# ═══════════════════════════════════════════════
# SALVAR
# ═══════════════════════════════════════════════
wb.save(output_path)
print(f'PDI Elias v3 salvo com sucesso em: {output_path}')
print('Melhorias aplicadas:')
print('  1. Acompanhamento Semanal: 13 → 26 semanas (ciclo completo de 6 meses)')
print('  2. Resultado e Evolução: 3 → 6 meses de acompanhamento')
print('  3. Checkpoints: 3 → 6 meses (adicionados meses 4, 5 e 6/Final)')
print('  4. Semáforo: 3 → 6 colunas de mês + 2 novos indicadores')
print('  5. Nova aba: Guia 1-1 adaptado para Caio ↔ Elias')
print('  6. Feedback anônimo: 2 → 3 coletas (bimestral em 6 meses)')
print('  7. Versão atualizada para v3')
