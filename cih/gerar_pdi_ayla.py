#!/usr/bin/env python3
"""Gera o PDI da Ayla Lima — Vale Assessoria, no modelo do PDI Elias Fernandes."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Cores ──
TEAL = '0097A7'
DARK_TEAL = '00796B'
WHITE = 'FFFFFF'
LIGHT_TEAL = 'E0F2F1'
LIGHT_GRAY = 'F5F5F5'
LIGHT_GREEN = 'E8F5E9'
LIGHT_YELLOW = 'FFFDE7'
LIGHT_RED = 'FFEBEE'
LIGHT_ORANGE = 'FFF3E0'
LIGHT_BLUE = 'E3F2FD'
PALE_YELLOW = 'FFF8E1'
LIGHT_PURPLE = 'F3E5F5'
GRAY_TEXT = '757575'
GREEN_BG = '43A047'
YELLOW_BG = 'FFC107'
RED_BG = 'E53935'
DARK_GRAY = '424242'
ORANGE_TEXT = 'F57C00'
RED_TEXT = 'E53935'

# ── Estilos reutilizáveis ──
def hfill(c):
    return PatternFill(start_color=c, end_color=c, fill_type='solid')

header_fill = hfill(TEAL)
header_font = Font(bold=True, size=11, color=WHITE)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
title_font = Font(bold=True, size=14, color=TEAL)
title_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
subtitle_font = Font(bold=True, size=12, color=DARK_TEAL)
label_fill = hfill(LIGHT_TEAL)
label_font = Font(bold=True, size=10)
data_font = Font(size=10)
data_align_l = Alignment(horizontal='left', vertical='center', wrap_text=True)
data_align_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
ctx_fill = hfill(LIGHT_GRAY)
ctx_font = Font(bold=True, size=10)
ctx_data_align = Alignment(vertical='top', wrap_text=True)
src_fill = hfill(LIGHT_GRAY)

thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

def set_header_row(ws, row, cols, merge_end=None):
    """Aplica estilo de cabeçalho teal a uma linha."""
    for ci, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = thin_border

def set_data_cell(ws, row, col, val, font=None, fill=None, align=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font or data_font
    if fill:
        c.fill = fill
    c.alignment = align or data_align_l
    c.border = thin_border
    return c

def level_fill(level):
    if level <= 1: return hfill(LIGHT_RED)
    if level == 2: return hfill(LIGHT_YELLOW)
    return hfill(LIGHT_GREEN)

def type_fill(tipo):
    if '70%' in tipo: return hfill(LIGHT_BLUE)
    if '20%' in tipo: return hfill(PALE_YELLOW)
    return hfill(LIGHT_PURPLE)

# ═══════════════════════════════════════════════
# DADOS DA AYLA
# ═══════════════════════════════════════════════
COMPETENCIAS = [
    (1, 'Organização e Gestão de Rotina', 1, 3, '25%'),
    (2, 'Atenção aos Detalhes e Saída do Automático', 1, 3, '25%'),
    (3, 'Comunicação Profissional e Formalidade', 2, 3, '20%'),
    (4, 'Pensamento Analítico e Gestão de Clientes', 1, 3, '20%'),
    (5, 'Autonomia e Senso de Dono', 2, 3, '10%'),
]

ACOES = [
    # (num, competência_curta, tipo, ação, desc, início, fim, freq, indicador, resp, fonte)
    (1,'Organização e Gestão','70% Experiência','Implementar planner/agenda física',
     'Comprar e utilizar um planner ou caderno para organizar a semana. Dividir atividades por turno (manhã/tarde). Listar tarefas pendentes, prazos de aprovação e entregas do dia. Substituir o grupo de WhatsApp individual.',
     'Mar/2026','Jun/2026','Diário','Planner preenchido 5x/semana','Ayla','Devolutiva'),
    (2,'Organização e Gestão','70% Experiência','Checklist diário por turno',
     'Criar checklist estruturado com tarefas obrigatórias de cada turno: verificar aprovações pendentes, enviar conteúdos programados, conferir Trello, enviar pendências ao final do dia.',
     'Mar/2026','Jun/2026','Diário','100% checklists preenchidos','Ayla','Briefing liderança'),
    (3,'Organização e Gestão','70% Experiência','Verificação do Trello 2x por turno',
     'Abrir o Trello obrigatoriamente no início e fim de cada turno. Verificar: cards vencidos, exclusões recentes, alterações de status, conteúdos aguardando aprovação. Anotar no planner.',
     'Mar/2026','Jun/2026','Diário','4 verificações/dia registradas','Ayla','Devolutiva + Briefing'),
    (4,'Organização e Gestão','70% Experiência','Rotina de encerramento do dia',
     'Antes de finalizar o expediente: enviar lista de pendências para o grupo, verificar aprovações atrasadas, conferir se todos os posts foram enviados. Criar hábito de "fechar o dia".',
     'Mar/2026','Jun/2026','Diário','Pendências enviadas 5x/semana','Ayla','Devolutiva'),
    (5,'Organização e Gestão','20% Social','Acompanhamento quinzenal com Larissa',
     '15 min quinzenais para revisar juntas: uso do planner, checklist, organização do Trello. Larissa dá feedback sobre evolução da organização e pontualidade nas entregas.',
     'Mar/2026','Jun/2026','Quinzenal','2 conversas/mês registradas','Ayla + Larissa','Briefing'),
    (6,'Organização e Gestão','10% Formal','Estudo sobre gestão de tempo',
     'Assistir conteúdo sobre técnicas de organização pessoal (Pomodoro, GTD ou similar). Escolher 1 técnica e implementar no dia a dia. Discutir com Eleine na sessão.',
     'Mar/2026','Abr/2026','Única','1 técnica escolhida e implementada','Ayla','Devolutiva'),
    (7,'Atenção aos Detalhes','70% Experiência','Protocolo "PARE, CONFIRA, ENVIE"',
     'Antes de qualquer envio (post, aprovação, mensagem para cliente): PARAR, CONFERIR (grupo correto? conteúdo correto? aprovação dada?), e só então ENVIAR. Fixar regra no monitor como lembrete visual.',
     'Mar/2026','Jun/2026','Contínuo','Zero erros de envio por semana','Ayla','Devolutiva'),
    (8,'Atenção aos Detalhes','70% Experiência','Diário de erros e aprendizados',
     'Registrar cada erro cometido: o que aconteceu, por que aconteceu, como evitar. Revisar semanalmente para identificar padrões. Foco: sair do automático.',
     'Mar/2026','Mai/2026','Contínuo','Redução progressiva de erros','Ayla','DISC + Devolutiva'),
    (9,'Atenção aos Detalhes','70% Experiência','Revisão dupla de postagens',
     'Antes de publicar qualquer conteúdo, verificar: 1) Está no card correto do Trello? 2) Foi aprovado? 3) É o grupo/rede certa? 4) Texto e imagem estão corretos? Criar checklist mental de 4 pontos.',
     'Mar/2026','Jun/2026','Contínuo','0 postagens erradas/semana','Ayla','Briefing'),
    (10,'Atenção aos Detalhes','20% Social','Feedback semanal com Larissa sobre erros',
     'Larissa reporta erros detectados na semana. Ayla analisa se eram evitáveis com o protocolo. Ajustar processo conforme necessário.',
     'Mar/2026','Jun/2026','Semanal','Redução semanal de erros','Ayla + Larissa','Briefing'),
    (11,'Comunicação Profissional','70% Experiência','Criar banco de mensagens padrão',
     'Criar templates para: cobranças de aprovação, comunicados de feriado, onboarding de novos clientes, solicitação de materiais, follow-up. Usar linguagem formal e detalhista.',
     'Mar/2026','Abr/2026','Única','5+ templates criados e aprovados','Ayla','Devolutiva'),
    (12,'Comunicação Profissional','70% Experiência','Adaptar linguagem por perfil de cliente',
     'Identificar quais clientes preferem comunicação direta (executor) vs detalhada (analista). Ajustar tom, nível de detalhe e formalidade conforme cada cliente.',
     'Mar/2026','Jun/2026','Contínuo','0 reclamações de tom/linguagem','Ayla','DISC'),
    (13,'Comunicação Profissional','20% Social','Revisão de comunicações com Larissa',
     'Submeter comunicações importantes para revisão antes do envio. Larissa indica ajustes de tom, formalidade e detalhamento. Reduzir gradualmente a necessidade de revisão.',
     'Mar/2026','Mai/2026','Semanal','Aprovação em 1ª tentativa ≥80%','Ayla + Larissa','Devolutiva'),
    (14,'Comunicação Profissional','10% Formal','Estudo do resultado DISC completo',
     'Ler o material DISC completo (38 páginas). Entender perfis Comunicador, Executor, Analista e Planejador. Aplicar conceitos de comunicação adaptada no dia a dia.',
     'Mar/2026','Mar/2026','Única','Leitura concluída + 3 insights','Ayla','DISC'),
    (15,'Pensamento Analítico','70% Experiência','Análise semanal de clientes no Trello',
     'Dedicar 30 min por semana para analisar cada cliente: cards atrasados, aprovações pendentes, gargalos. Mapear situação e antecipar problemas.',
     'Mar/2026','Jun/2026','Semanal','Análise de 100% dos clientes/semana','Ayla','Briefing'),
    (16,'Pensamento Analítico','70% Experiência','Relatório "Top 3 Clientes" quinzenal',
     'Trazer para Larissa os 3 clientes mais críticos: qual o problema, possível causa, sugestão de ação. Desenvolver pensamento analítico aplicado.',
     'Mar/2026','Jun/2026','Quinzenal','2 relatórios/mês com análise','Ayla','Briefing'),
    (17,'Pensamento Analítico','70% Experiência','Registro de soluções no playbook pessoal',
     'Ao resolver um problema, registrar: situação, solução aplicada, resultado. Usar como base para situações futuras similares. Construir playbook pessoal.',
     'Mar/2026','Jun/2026','Contínuo','3+ soluções registradas/mês','Ayla','Briefing + Devolutiva'),
    (18,'Pensamento Analítico','20% Social','Discussão com Larissa sobre leitura de cenários',
     'Apresentar análise dos clientes para Larissa e discutir: o que Ayla viu vs o que não percebeu. Desenvolver visão holística e antecipação de problemas.',
     'Mar/2026','Jun/2026','Quinzenal','2 discussões/mês com insights','Ayla + Larissa','Devolutiva'),
    (19,'Autonomia e Senso de Dono','70% Experiência','Cobrar proativamente a equipe',
     'Não esperar que design atrase para cobrar. Criar rotina de follow-up: verificar status das demandas, cobrar com assertividade e profissionalismo. Escalar para Larissa se necessário.',
     'Mar/2026','Jun/2026','Contínuo','0 atrasos por falta de cobrança','Ayla','Devolutiva'),
    (20,'Autonomia e Senso de Dono','70% Experiência','Pedir ajuda cedo em vez de travar',
     'Quando travar em uma tarefa por mais de 15 min, pedir ajuda. Registrar: o que travou, quanto tempo ficou travada. Meta: reduzir tempo travada de 30min+ para <15min.',
     'Mar/2026','Jun/2026','Contínuo','Tempo médio até pedir ajuda <15min','Ayla','DISC + Devolutiva'),
]

DISC_INDICADORES = [
    ('Perfil Analista (A)', '19,48%', 'Baixo', 'Gap principal — cargo exige detalhismo', 'Foco do PDI em organização e atenção'),
    ('Autoestima (IAE)', 'Baixa', 'Atenção', 'Autocrítica elevada, risco de desânimo', 'Feedback positivo e celebração de conquistas'),
    ('Aproveitamento (IA)', 'Normal Baixo', 'Atenção', 'Sente habilidades subutilizadas', 'Dar mais responsabilidade e desafios'),
    ('Autoconfiança (IAC)', 'Normal Baixa', 'Atenção', 'Sente necessidade de mudar muito', 'Evolução gradual com metas curtas'),
    ('Energia (EN)', 'Alta', 'Positivo', 'Disposição e motivação presentes', 'Canalizar em organização e análise'),
    ('Área de Talento', 'Diplomata', 'Extremamente Alto', 'Excelente em relações interpessoais', 'Usar como alavanca na comunicação'),
]

# ═══════════════════════════════════════════════
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════
# ABA 1 — RESUMO PDI
# ═══════════════════════════════════════════════
ws = wb.active
ws.title = 'Resumo PDI'
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 45
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15

# Título
ws.merge_cells('A1:H1')
c = ws.cell(row=1, column=1, value='PDI — AYLA JESUS LIMA | VALE ASSESSORIA')
c.font = title_font; c.alignment = title_align

ws.merge_cells('A2:H2')
c = ws.cell(row=2, column=1, value='Plano de Desenvolvimento Individual — Ciclo Março a Junho/2026')
c.font = Font(bold=True, size=12, color=DARK_TEAL); c.alignment = title_align

# Ficha do colaborador
ficha = [
    ('Colaboradora', 'Ayla Jesus Lima'),
    ('Cargo', 'Estagiária → Analista de Operações e Atendimento de Contas (em transição)'),
    ('Empresa', 'Vale Assessoria'),
    ('Tempo de Casa', '~1 ano'),
    ('Perfil DISC', 'Comunicador Executor (CE) — Diplomata'),
    ('Liderança Direta', 'Larissa Carinhanha (gestora — responsável pelo acompanhamento do PDI)'),
    ('Consultora CIH', 'Eleine Passos'),
    ('E-mail', 'aylalima93@gmail.com'),
    ('Ciclo PDI', 'Março/2026 a Junho/2026 (4 meses)'),
    ('Objetivo', 'Reduzir riscos na promoção e acelerar maturidade comportamental e profissional'),
]
for i, (lab, val) in enumerate(ficha, 4):
    cl = ws.cell(row=i, column=1, value=lab)
    cl.fill = label_fill; cl.font = label_font; cl.border = thin_border
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    cv = ws.cell(row=i, column=2, value=val)
    cv.font = data_font; cv.border = thin_border

# Seção de contexto
r = 15
ws.merge_cells(f'A{r}:H{r}')
ws.cell(row=r, column=1, value='CONTEXTO DA PROMOÇÃO (BRIEFING COM LIDERANÇA)').font = subtitle_font

contexto = [
    ('Papel definido', 'Analista de Operações e Atendimento de Contas (PJ). Responsável por gestão de materiais, comunicação com clientes, organização interna, onboarding e apoio a projetos.'),
    ('Expectativa da liderança', 'Sair do modo automático, reduzir erros de envio/postagem, ser mais organizada, formal na comunicação com clientes e analítica. Cobrar equipe com assertividade.'),
    ('Gap identificado', 'Perfil Analista mais baixo (19,48%) vs alta exigência do cargo. Modo automático gera erros frequentes. Informalidade excessiva na comunicação profissional.'),
    ('Ferramentas em uso', 'Trello (gestão de tarefas), WhatsApp (grupos de clientes), Canva (criação emergencial), Google Drive. Sugestão: planner físico + checklist diário.'),
    ('Risco na promoção', 'Sem desenvolvimento, erros operacionais podem aumentar com nova responsabilidade. Falta de organização e detalhismo compromete a entrega ao cliente.'),
    ('Compromisso 50/50', 'Desenvolvimento é 50% da colaboradora e 50% da empresa. Liderança fornecerá feedback constante e tempo para implementação das mudanças.'),
]
for j, (lab, val) in enumerate(contexto, r+1):
    cl = ws.cell(row=j, column=1, value=lab)
    cl.fill = ctx_fill; cl.font = ctx_font; cl.border = thin_border
    ws.merge_cells(start_row=j, start_column=2, end_row=j, end_column=8)
    cv = ws.cell(row=j, column=2, value=val)
    cv.font = data_font; cv.alignment = ctx_data_align; cv.border = thin_border
    ws.row_dimensions[j].height = 35

# Competências priorizadas
r = 23
ws.merge_cells(f'A{r}:H{r}')
ws.cell(row=r, column=1, value='COMPETÊNCIAS PRIORIZADAS (5 COMPETÊNCIAS)').font = subtitle_font

r = 24
set_header_row(ws, r, ['#','Competência','Nível Atual','Meta','Peso','Status Mês 1','Status Mês 2','Status Mês 3'])

for ci, (num, nome, atual, meta, peso) in enumerate(COMPETENCIAS, 0):
    row = 25 + ci
    set_data_cell(ws, row, 1, num, align=data_align_l)
    set_data_cell(ws, row, 2, nome, align=data_align_l)
    set_data_cell(ws, row, 3, atual, fill=level_fill(atual), align=data_align_c)
    set_data_cell(ws, row, 4, meta, fill=hfill(LIGHT_GREEN), align=data_align_c)
    set_data_cell(ws, row, 5, peso, align=data_align_c)

# Indicadores DISC
r = 31
ws.merge_cells(f'A{r}:H{r}')
ws.cell(row=r, column=1, value='INDICADORES DISC — PONTOS DE ATENÇÃO').font = subtitle_font

r = 32
set_header_row(ws, r, ['Indicador','Valor','Classificação','Impacto','Ação Recomendada','','',''])

for j, (ind, val, classif, impacto, acao) in enumerate(DISC_INDICADORES, r+1):
    set_data_cell(ws, j, 1, ind, align=data_align_l)
    set_data_cell(ws, j, 2, val, align=data_align_l)
    classif_fill = hfill(LIGHT_ORANGE) if classif in ('Baixo','Atenção') else hfill(LIGHT_GREEN)
    set_data_cell(ws, j, 3, classif, fill=classif_fill, align=data_align_l)
    set_data_cell(ws, j, 4, impacto, align=data_align_l)
    set_data_cell(ws, j, 5, acao, align=data_align_l)

# ═══════════════════════════════════════════════
# ABA 2 — PLANO DE AÇÕES
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet('Plano de Ações')
ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 35
ws2.column_dimensions['E'].width = 60
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 12
ws2.column_dimensions['H'].width = 12
ws2.column_dimensions['I'].width = 38
ws2.column_dimensions['J'].width = 15
ws2.column_dimensions['K'].width = 18
ws2.column_dimensions['L'].width = 12

ws2.merge_cells('A1:L1')
ws2.cell(row=1, column=1, value='PLANO DE AÇÕES — PDI AYLA LIMA (20 ações, 5 competências)').font = title_font
ws2['A1'].alignment = title_align

set_header_row(ws2, 3, ['#','Competência','Tipo 70-20-10','Ação','Descrição Detalhada',
                         'Prazo Início','Prazo Fim','Frequência','Indicador de Sucesso',
                         'Responsável','Fonte','Status'])

for i, a in enumerate(ACOES):
    row = 4 + i
    num, comp, tipo, acao, desc, ini, fim, freq, indic, resp, fonte = a
    set_data_cell(ws2, row, 1, num, align=data_align_l)
    set_data_cell(ws2, row, 2, comp, align=data_align_l)
    set_data_cell(ws2, row, 3, tipo, fill=type_fill(tipo), align=data_align_l)
    set_data_cell(ws2, row, 4, acao, align=data_align_l)
    set_data_cell(ws2, row, 5, desc, align=data_align_l)
    set_data_cell(ws2, row, 6, ini, align=data_align_l)
    set_data_cell(ws2, row, 7, fim, align=data_align_l)
    set_data_cell(ws2, row, 8, freq, align=data_align_l)
    set_data_cell(ws2, row, 9, indic, align=data_align_l)
    set_data_cell(ws2, row, 10, resp, align=data_align_l)
    set_data_cell(ws2, row, 11, fonte, fill=src_fill, align=data_align_l)
    set_data_cell(ws2, row, 12, 'Pendente', align=data_align_c)

# ═══════════════════════════════════════════════
# ABA 3 — ACOMPANHAMENTO SEMANAL
# ═══════════════════════════════════════════════
ws3 = wb.create_sheet('Acompanhamento Semanal')
ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 55
ws3.column_dimensions['D'].width = 14
ws3.column_dimensions['E'].width = 14
ws3.column_dimensions['F'].width = 14
ws3.column_dimensions['G'].width = 14
ws3.column_dimensions['H'].width = 14
ws3.column_dimensions['I'].width = 35
ws3.column_dimensions['J'].width = 35
ws3.column_dimensions['K'].width = 30
ws3.column_dimensions['L'].width = 30

ws3.merge_cells('A1:L1')
ws3.cell(row=1, column=1, value='ACOMPANHAMENTO SEMANAL — MARÇO A JUNHO/2026').font = title_font
ws3['A1'].alignment = title_align

set_header_row(ws3, 3, ['Semana','Período','Foco Principal',
    'Organização\n(Nota 1-5)','Atenção\n(Nota 1-5)','Comunicação\n(Nota 1-5)',
    'Pens. Analítico\n(Nota 1-5)','Autonomia\n(Nota 1-5)',
    'Ações Realizadas','Dificuldades / Bloqueios','Próximos Passos','Observações do Consultor'])

semanas = [
    ('Semana 1','03/03 a 07/03','Implementar planner + 1° checklist + Protocolo PARE-CONFIRA-ENVIE + Verificação Trello',False),
    ('Semana 2','10/03 a 14/03','Início banco de mensagens + Análise clientes Trello + Diário de erros + Leitura DISC',False),
    ('Semana 3','17/03 a 21/03','1° Top 3 Clientes + Feedback Larissa + Adaptar linguagem + Cobranças assertivas',False),
    ('Semana 4','24/03 a 28/03','CHECKPOINT MÊS 1 — Avaliar primeiras semanas + Estudo gestão de tempo',True),
    ('Semana 5','31/03 a 04/04','Consolidar organização + Aprofundar análise de clientes + Templates de mensagens',False),
    ('Semana 6','07/04 a 11/04','2° Top 3 Clientes + Revisão comunicações + Playbook pessoal',False),
    ('Semana 7','14/04 a 18/04','Ajustar protocolo de revisão + Feedback Larissa + Verificar erros recorrentes',False),
    ('Semana 8','21/04 a 25/04','CHECKPOINT MÊS 2 — Avaliar evolução + Larissa avalia organização + Revisão PDI',True),
    ('Semana 9','28/04 a 02/05','Consolidar comunicação formal + Aprofundar análise de cenários + Cobrança equipe',False),
    ('Semana 10','05/05 a 09/05','3° Top 3 Clientes + Avaliar playbook + Autonomia em decisões',False),
    ('Semana 11','12/05 a 16/05','Verificar hábitos consolidados + Planner e checklist automáticos?',False),
    ('Semana 12','19/05 a 23/05','Preparar evidências de evolução + Consolidar competências',False),
    ('Semana 13','26/05 a 30/05','CHECKPOINT MÊS 3 — Revisão geral + Avaliar prontidão para promoção',True),
    ('Semana 14','02/06 a 06/06','Consolidação final + Testar autonomia total',False),
    ('Semana 15','09/06 a 13/06','Ajustes finais + Preparar parecer de evolução',False),
    ('Semana 16','16/06 a 20/06','CHECKPOINT FINAL — Avaliação final + Parecer de prontidão para promoção',True),
]

for i, (sem, per, foco, is_cp) in enumerate(semanas, 4):
    fill_row = hfill(LIGHT_ORANGE) if is_cp else None
    c1 = set_data_cell(ws3, i, 1, sem, font=Font(bold=True, size=10), fill=fill_row, align=data_align_c)
    c2 = set_data_cell(ws3, i, 2, per, fill=fill_row, align=data_align_c)
    c3 = set_data_cell(ws3, i, 3, foco, fill=fill_row, align=Alignment(vertical='top', wrap_text=True))
    ws3.row_dimensions[i].height = 45

# ═══════════════════════════════════════════════
# ABA 4 — RESULTADO E EVOLUÇÃO
# ═══════════════════════════════════════════════
ws4 = wb.create_sheet('Resultado e Evolução')
ws4.column_dimensions['A'].width = 48
for col in 'BCDEFGH':
    ws4.column_dimensions[col].width = 15

ws4.merge_cells('A1:H1')
ws4.cell(row=1, column=1, value='RESULTADO DO PDI — EVOLUÇÃO DAS COMPETÊNCIAS').font = title_font
ws4['A1'].alignment = title_align

set_header_row(ws4, 3, ['Competência','Nível Inicial','Meta','Mês 1','Mês 2','Mês 3','Nível Final','Atingiu?'])

comp_names_full = [
    'Organização e Gestão de Rotina',
    'Atenção aos Detalhes e Saída do Automático',
    'Comunicação Profissional e Formalidade',
    'Pensamento Analítico e Gestão de Clientes',
    'Autonomia e Senso de Dono',
]
comp_levels = [(1,3),(1,3),(2,3),(1,3),(2,3)]

for i, (nome, (ini, meta)) in enumerate(zip(comp_names_full, comp_levels), 4):
    set_data_cell(ws4, i, 1, nome, align=data_align_l)
    set_data_cell(ws4, i, 2, ini, fill=level_fill(ini), align=data_align_c)
    set_data_cell(ws4, i, 3, meta, fill=hfill(LIGHT_GREEN), align=data_align_c)

# Indicadores de acompanhamento
r = 10
ws4.merge_cells(f'A{r}:H{r}')
ws4.cell(row=r, column=1, value='INDICADORES DE ACOMPANHAMENTO').font = subtitle_font

r = 11
set_header_row(ws4, r, ['Indicador','Baseline','Meta Mês 1','Real Mês 1','Meta Mês 2','Real Mês 2','Meta Mês 3','Real Mês 3'])

indicadores = [
    ('Checklist diário preenchido (%)', 0, '80%', '', '90%', '', '100%', ''),
    ('Verificações Trello por dia', 0, 4, '', 4, '', 4, ''),
    ('Erros de envio por semana (↓)', '—', 'Registrar', '', 'Reduzir 30%', '', 'Reduzir 60%', ''),
    ('Templates de mensagens criados', 0, 5, '', 'Usar', '', 'Ajustar', ''),
    ('Análises semanais de clientes', 0, 4, '', 4, '', 4, ''),
    ('Relatórios Top 3 Clientes', 0, 2, '', 2, '', 2, ''),
    ('Cobranças assertivas sem atraso', '—', 'Registrar', '', '80%+', '', '90%+', ''),
    ('Protocolo PARE-CONFIRA-ENVIE aplicado', '—', 'Sempre', '', 'Automático', '', 'Consolidado', ''),
    ('Reuniões feedback com Larissa', 0, 2, '', 2, '', 2, ''),
    ('Soluções registradas no playbook', 0, 3, '', 3, '', 3, ''),
    ('Planner utilizado diariamente', 'Não', '80%', '', '90%', '', '100%', ''),
    ('Ações do PDI concluídas (acumulado)', '0/20', '7/20', '', '14/20', '', '18/20', ''),
]

for i, ind in enumerate(indicadores, r+1):
    for j, val in enumerate(ind):
        set_data_cell(ws4, i, j+1, val, align=data_align_c if j > 0 else data_align_l)

# Feedback da liderança
r2 = r + 1 + len(indicadores) + 1
ws4.merge_cells(f'A{r2}:H{r2}')
ws4.cell(row=r2, column=1, value='FEEDBACK DA LIDERANÇA (COLETAS COM LARISSA)').font = subtitle_font

r2 += 1
set_header_row(ws4, r2, ['Pergunta','1ª Coleta (Mês 1)\nNota 1-5','Comentários',
                          '2ª Coleta (Mês 2)\nNota 1-5','Comentários','Evolução','',''])

perguntas_fb = [
    'Como você avalia a organização da Ayla nas entregas?',
    'Ayla tem reduzido erros de envio e postagem?',
    'A comunicação com clientes está mais profissional e formal?',
    'Ayla está usando o Trello de forma consistente?',
    'Ayla pede ajuda quando precisa ou trava sozinha?',
    'As cobranças à equipe estão mais assertivas e profissionais?',
    'Ayla demonstra pensamento analítico sobre os clientes?',
    'Ayla demonstra mais autonomia e senso de responsabilidade?',
    'Nota geral para a evolução da Ayla (1 a 5)',
]

for i, p in enumerate(perguntas_fb, r2+1):
    set_data_cell(ws4, i, 1, p, align=Alignment(vertical='top', wrap_text=True))

# Resumo executivo
r3 = r2 + 1 + len(perguntas_fb) + 1
ws4.merge_cells(f'A{r3}:H{r3}')
ws4.cell(row=r3, column=1, value='RESUMO EXECUTIVO — RESULTADO DO CICLO').font = subtitle_font

resumo_items = [
    'Principais conquistas:',
    'Competências com maior evolução:',
    'Competências que precisam de mais tempo:',
    'Impacto percebido nos clientes:',
    'Impacto na operação (organização, comunicação, entregas):',
    'Percepção de Larissa sobre maturidade de Ayla:',
    'Evolução na redução de erros:',
    'Recomendações para o próximo ciclo:',
    'Nota geral do PDI (consultor):',
]

for i, item in enumerate(resumo_items, r3+1):
    ws4.cell(row=i, column=1, value=item).font = Font(bold=True, size=10)
    ws4.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
    ws4.row_dimensions[i].height = 30

# ═══════════════════════════════════════════════
# ABA 5 — CHECKPOINTS
# ═══════════════════════════════════════════════
ws5 = wb.create_sheet('Checkpoints')
ws5.column_dimensions['A'].width = 8
ws5.column_dimensions['B'].width = 55
ws5.column_dimensions['C'].width = 25
for col in 'DEFG':
    ws5.column_dimensions[col].width = 15

ws5.merge_cells('A1:G1')
ws5.cell(row=1, column=1, value='CHECKPOINTS MENSAIS — REGISTRO DE SESSÕES').font = title_font
ws5['A1'].alignment = title_align

# Checkpoint Mês 1
r = 3
ws5.merge_cells(f'A{r}:G{r}')
c = ws5.cell(row=r, column=1, value='CHECKPOINT MÊS 1 — Fim de Março/2026')
c.fill = hfill(LIGHT_TEAL); c.font = Font(bold=True, size=12, color=DARK_TEAL)

ws5.cell(row=r+1, column=1, value='Participantes:').font = Font(bold=True, size=10)
ws5.cell(row=r+1, column=2, value='Eleine + Ayla + Larissa').font = data_font
ws5.cell(row=r+2, column=1, value='Data realizada:').font = Font(bold=True, size=10)
ws5.cell(row=r+2, column=2, value='___/___/______').font = data_font

r += 3
set_header_row(ws5, r, ['#','Pergunta / Item','Resposta / Evidência','','','',''])

perguntas_m1 = [
    'Ayla está usando o planner diariamente?',
    'Checklist por turno: está preenchendo e seguindo?',
    'Verificação do Trello: está fazendo 2x por turno?',
    'Protocolo PARE-CONFIRA-ENVIE: está aplicando antes de cada envio?',
    'Diário de erros: quantos erros registrou? Quais padrões?',
    'Banco de mensagens padrão: quantos templates criou?',
    'Comunicação com clientes: está mais formal e detalhista?',
    'Análise de clientes: fez análise semanal no Trello?',
    'Top 3 Clientes: entregou o 1° relatório?',
    'Cobranças à equipe: está cobrando proativamente?',
    'Pedir ajuda: tempo médio até pedir ajuda diminuiu?',
    'Leitura DISC: concluiu? Quais insights?',
    'Nível de motivação de Ayla (autoavaliação 1-5):',
    'Percepção de Larissa: Ayla está evoluindo? O que falta?',
    'Ajustes necessários no PDI:',
]

for i, p in enumerate(perguntas_m1):
    row = r + 1 + i
    set_data_cell(ws5, row, 1, i+1, align=data_align_c)
    set_data_cell(ws5, row, 2, p, align=Alignment(vertical='top', wrap_text=True))
    ws5.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    ws5.row_dimensions[row].height = 30

# Checkpoint Mês 2
r = r + 1 + len(perguntas_m1) + 1
ws5.merge_cells(f'A{r}:G{r}')
c = ws5.cell(row=r, column=1, value='CHECKPOINT MÊS 2 — Fim de Abril/2026')
c.fill = hfill(LIGHT_TEAL); c.font = Font(bold=True, size=12, color=DARK_TEAL)

ws5.cell(row=r+1, column=1, value='Participantes:').font = Font(bold=True, size=10)
ws5.cell(row=r+1, column=2, value='Eleine + Ayla + Larissa').font = data_font
ws5.cell(row=r+2, column=1, value='Data realizada:').font = Font(bold=True, size=10)
ws5.cell(row=r+2, column=2, value='___/___/______').font = data_font

r += 3
set_header_row(ws5, r, ['#','Pergunta / Item','Resposta / Evidência','','','',''])

perguntas_m2 = [
    'Organização: planner e checklist viraram hábito?',
    'Erros de envio: houve redução significativa?',
    'Protocolo de revisão: está automático ou ainda precisa lembrar?',
    'Templates de mensagem: estão sendo usados? São eficazes?',
    'Comunicação: Larissa percebe mais formalidade e profissionalismo?',
    'Análise de clientes: Ayla antecipa problemas ou apenas reage?',
    'Top 3 Clientes: qualidade da análise está melhorando?',
    'Autonomia: Ayla resolve mais sozinha ou ainda depende de passo a passo?',
    'Cobranças: equipe responde melhor?',
    'Playbook de soluções: está construindo e usando?',
    'Percepção de Larissa: confiança na promoção aumentou?',
    'Feedback de clientes (informal): perceberam mudança?',
    'Ajustes necessários no PDI:',
]

for i, p in enumerate(perguntas_m2):
    row = r + 1 + i
    set_data_cell(ws5, row, 1, i+1, align=data_align_c)
    set_data_cell(ws5, row, 2, p, align=Alignment(vertical='top', wrap_text=True))
    ws5.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    ws5.row_dimensions[row].height = 30

# Checkpoint Mês 3
r = r + 1 + len(perguntas_m2) + 1
ws5.merge_cells(f'A{r}:G{r}')
c = ws5.cell(row=r, column=1, value='CHECKPOINT MÊS 3 — Fim de Maio/2026')
c.fill = hfill(LIGHT_TEAL); c.font = Font(bold=True, size=12, color=DARK_TEAL)

ws5.cell(row=r+1, column=1, value='Participantes:').font = Font(bold=True, size=10)
ws5.cell(row=r+1, column=2, value='Eleine + Ayla + Larissa').font = data_font
ws5.cell(row=r+2, column=1, value='Data realizada:').font = Font(bold=True, size=10)
ws5.cell(row=r+2, column=2, value='___/___/______').font = data_font

r += 3
set_header_row(ws5, r, ['#','Pergunta / Item','Resposta / Evidência','','','',''])

perguntas_m3 = [
    'Evolução geral nas 5 competências (nota 1-5 cada):',
    'Organização: é consistente sem supervisão?',
    'Atenção: erros recorrentes foram eliminados?',
    'Comunicação: padrão profissional consolidado?',
    'Pensamento analítico: traz análise sem ser solicitada?',
    'Autonomia: resolve problemas e cobra equipe com naturalidade?',
    'Planner e checklist: rotina automatizada?',
    'Feedback de Larissa: pronta para a promoção?',
    'Feedback de clientes: percebem evolução?',
    'O que funcionou melhor no PDI?',
    'O que precisa mudar para consolidar?',
    'Parecer de prontidão para promoção:',
    'Recomendações de desenvolvimento futuro:',
]

for i, p in enumerate(perguntas_m3):
    row = r + 1 + i
    set_data_cell(ws5, row, 1, i+1, align=data_align_c)
    set_data_cell(ws5, row, 2, p, align=Alignment(vertical='top', wrap_text=True))
    ws5.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    ws5.row_dimensions[row].height = 30

# ═══════════════════════════════════════════════
# ABA 6 — SEMÁFORO
# ═══════════════════════════════════════════════
ws6 = wb.create_sheet('Semáforo')
ws6.column_dimensions['A'].width = 22
ws6.column_dimensions['B'].width = 40
ws6.column_dimensions['C'].width = 18
ws6.column_dimensions['D'].width = 15
ws6.column_dimensions['E'].width = 15
ws6.column_dimensions['F'].width = 15

ws6.merge_cells('A1:F1')
ws6.cell(row=1, column=1, value='SEMÁFORO DE PROGRESSO — VISÃO EXECUTIVA').font = title_font
ws6['A1'].alignment = title_align

ws6.merge_cells('A2:F2')
c = ws6.cell(row=2, column=1, value='Verde = No caminho | Amarelo = Atenção | Vermelho = Intervenção necessária')
c.font = Font(size=10, color=GRAY_TEXT); c.alignment = data_align_c

set_header_row(ws6, 4, ['Dimensão','Indicador-Chave','Meta','Mês 1','Mês 2','Mês 3'])

semaforo_items = [
    ('Organização', 'Planner utilizado diariamente', 'Diário'),
    ('Organização', 'Checklist preenchido por turno', 'Diário'),
    ('Organização', 'Trello verificado 2x/turno', 'Diário'),
    ('Atenção', 'Erros de envio por semana (↓)', '0/semana'),
    ('Atenção', 'Protocolo PARE-CONFIRA-ENVIE aplicado', '100%'),
    ('Atenção', 'Pendências enviadas ao final do dia', 'Diário'),
    ('Comunicação', 'Templates criados e em uso', '5+'),
    ('Comunicação', 'Comunicação formal com clientes', '0 reclamações'),
    ('Comunicação', 'Adaptação por perfil de cliente', 'Contínuo'),
    ('Pens. Analítico', 'Análise semanal de clientes', 'Semanal'),
    ('Pens. Analítico', 'Relatório Top 3 Clientes', 'Quinzenal'),
    ('Pens. Analítico', 'Soluções registradas no playbook', '3+/mês'),
    ('Autonomia', 'Cobranças assertivas realizadas', '0 atrasos'),
    ('Autonomia', 'Tempo até pedir ajuda', '<15 min'),
    ('Autonomia', 'Iniciativa sem passo a passo', 'Crescente'),
    ('Geral', 'Ações do PDI concluídas', '90%+'),
    ('Geral', 'Reuniões com Larissa realizadas', '2/mês'),
    ('Geral', 'Motivação Ayla (1-5)', '≥ 4'),
    ('Geral', 'Larissa confia na promoção?', 'Sim'),
]

for i, (dim, ind, meta) in enumerate(semaforo_items, 5):
    set_data_cell(ws6, i, 1, dim, align=data_align_l)
    set_data_cell(ws6, i, 2, ind, align=data_align_l)
    set_data_cell(ws6, i, 3, meta, align=data_align_l)

# Legenda
r = 5 + len(semaforo_items) + 1
ws6.cell(row=r, column=1, value='LEGENDA:').font = Font(bold=True, size=10)

r += 1
c = ws6.cell(row=r, column=1, value='VERDE')
c.fill = hfill(GREEN_BG); c.font = Font(bold=True, size=10, color=WHITE); c.alignment = data_align_c
ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
ws6.cell(row=r, column=2, value='70%+ das ações concluídas, indicadores no caminho, Larissa percebe evolução, clientes satisfeitos').font = data_font

r += 1
c = ws6.cell(row=r, column=1, value='AMARELO')
c.fill = hfill(YELLOW_BG); c.font = Font(bold=True, size=10, color=DARK_GRAY); c.alignment = data_align_c
ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
ws6.cell(row=r, column=2, value='40-69% das ações, alguns indicadores atrasados, erros ainda frequentes, organização inconsistente').font = data_font

r += 1
c = ws6.cell(row=r, column=1, value='VERMELHO')
c.fill = hfill(RED_BG); c.font = Font(bold=True, size=10, color=WHITE); c.alignment = data_align_c
ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
ws6.cell(row=r, column=2, value='<40% das ações, erros não diminuíram, modo automático persiste, Larissa não percebe evolução, promoção em risco').font = data_font

# ═══════════════════════════════════════════════
# ABA 7 — COMPETÊNCIAS REFERÊNCIA
# ═══════════════════════════════════════════════
ws7 = wb.create_sheet('Competências Referência')
ws7.column_dimensions['A'].width = 22
ws7.column_dimensions['B'].width = 32
ws7.column_dimensions['C'].width = 55
ws7.column_dimensions['D'].width = 22
ws7.column_dimensions['E'].width = 20

ws7.merge_cells('A1:E1')
ws7.cell(row=1, column=1, value='MAPEAMENTO DE COMPETÊNCIAS — BASE PARA AVALIAÇÃO').font = title_font
ws7['A1'].alignment = title_align

ws7.merge_cells('A2:E2')
c = ws7.cell(row=2, column=1, value='Fonte: Planilha 190 Exemplos de Competências para Avaliação de Desempenho + Contexto Vale Assessoria')
c.font = Font(size=10, color=GRAY_TEXT); c.alignment = data_align_c

set_header_row(ws7, 4, ['Categoria','Competência','Relevância para Ayla','Nível Atual Estimado (1-5)','Prioridade no PDI'])

comps_ref = [
    ('Gestão de Operações','Organização e Planejamento','CRÍTICA — esquece tarefas, não usa ferramentas adequadas',1,'FFEBEE','E53935','Competência 1'),
    ('Gestão de Operações','Gestão do Tempo','CRÍTICA — não prioriza, não fecha o dia com revisão',1,'FFEBEE','E53935','Competência 1'),
    ('Gestão de Operações','Atenção ao Detalhe','CRÍTICA — modo automático, perfil Analista mais baixo (19,48%)',1,'FFEBEE','E53935','Competência 2'),
    ('Gestão de Operações','Controle de Qualidade','ALTA — erros de envio, postagens em grupos errados',2,'FFFDE7','F57C00','Competência 2'),
    ('Gestão de Operações','Gestão de Processos','ALTA — não segue fluxos de aprovação consistentemente',2,'FFFDE7','F57C00','Competência 1'),
    ('Comunicação','Comunicação Escrita','ALTA — precisa ser mais formal e detalhista',2,'FFFDE7','F57C00','Competência 3'),
    ('Comunicação','Comunicação Verbal','MÉDIA — boa comunicação oral, mas falta formalidade',3,'FFF3E0',None,'Competência 3'),
    ('Comunicação','Habilidades de Escuta','MÉDIA — no automático, perde informações importantes',2,'FFFDE7',None,'Competência 2'),
    ('Comunicação','Técnicas de Negociação','MÉDIA — precisa cobrar equipe e clientes assertivamente',2,'FFFDE7',None,'Competência 5'),
    ('Raciocínio Lógico','Pensamento Analítico','CRÍTICA — complica tarefas simples, falta raciocínio sequencial',1,'FFEBEE','E53935','Competência 4'),
    ('Raciocínio Lógico','Resolução de Problemas','ALTA — trava sozinha em vez de pedir ajuda',2,'FFFDE7','F57C00','Competência 5'),
    ('Raciocínio Lógico','Abordagem Metódica','ALTA — precisa desenvolver pensamento estruturado',1,'FFEBEE','F57C00','Competência 4'),
    ('Comportamental','Proatividade','ALTA — já possui, ponto forte a manter',4,'E8F5E9',None,'Ativo existente'),
    ('Comportamental','Autonomia','ALTA — precisa reduzir dependência de passo a passo',2,'FFFDE7','F57C00','Competência 5'),
    ('Comportamental','Assertividade','ALTA — dificuldade em cobrar equipe profissionalmente',2,'FFFDE7','F57C00','Competência 5'),
    ('Comportamental','Adaptabilidade','MÉDIA — demora para incorporar novos processos',2,'FFFDE7',None,'Competência 2'),
    ('Atendimento','Foco no Cliente','ALTA — já tem feeling, precisa mais análise',3,'FFF3E0',None,'Competência 4'),
    ('Atendimento','Relacionamento Interpessoal','ALTA — ponto forte (Diplomata), usar como alavanca',4,'E8F5E9',None,'Ativo existente'),
    ('Atendimento','Empatia','MÉDIA — boa naturalmente, manter e direcionar',3,'FFF3E0',None,'Ativo existente'),
    ('Desenvolvimento Pessoal','Vontade de Aprender','ALTA — nunca recusou atividade, busca entender',4,'E8F5E9',None,'Ativo existente'),
    ('Desenvolvimento Pessoal','Potencial de Crescimento','ALTA — liderança vê potencial real para promoção',3,'FFF3E0',None,'Ativo existente'),
    ('Desenvolvimento Pessoal','Comprometimento','ALTA — valoriza trabalhar na Vale, quer crescer',4,'E8F5E9',None,'Ativo existente'),
]

for i, (cat, comp, rel, nivel, nivel_color, rel_color, prio) in enumerate(comps_ref, 5):
    set_data_cell(ws7, i, 1, cat, align=data_align_l)
    set_data_cell(ws7, i, 2, comp, align=data_align_l)
    rel_font = Font(bold=True, size=10, color=rel_color) if rel_color else data_font
    set_data_cell(ws7, i, 3, rel, font=rel_font, align=data_align_l)
    set_data_cell(ws7, i, 4, nivel, fill=hfill(nivel_color), align=data_align_c)
    set_data_cell(ws7, i, 5, prio, align=data_align_c)

# ═══════════════════════════════════════════════
# SALVAR
# ═══════════════════════════════════════════════
output = '/home/headless/workspace/cih/operacional/execução/valeassessoria/aylalima/PDI_Ayla_Lima_ValeAssessoria.xlsx'
wb.save(output)
print(f'PDI salvo em: {output}')
